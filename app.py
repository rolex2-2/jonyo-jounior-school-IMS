from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, current_app
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from database import Fee, PaymentHistory, get_db, init_db, User, db, Student, Fee, Marks, LearningAreas, Messages, Announcements, TermInfo, PaymentHistory, AdminNotes, TeacherAssignments, LearningAreas, ClassTeachers, PerformanceLevels, ParentStudent, Mission, Vision, About, Contact
from forms import RegistrationForm, FeeFilterForm, LoginForm, TeacherRegistrationForm, StudentRegistrationForm, MarksForm, FeeForm, MessageForm, PerformanceLevelForm, TermInfoForm, BursarRegistrationForm, MarksFilterForm
from utils import generate_teacher_password, generate_excel_results, generate_report_card, generate_fee_statement, get_performance_levels, get_points, get_teacher_name, get_rank, get_total_fee, get_amount_paid, get_balance, get_class_teacher_comment, get_principal_comment, get_class_teacher_name, get_principal_name
from wtforms import StringField
from wtforms.validators import DataRequired
import pandas as pd
from datetime import datetime
from functools import wraps
from utils import generate_teacher_password, generate_excel_results, generate_report_card, generate_fee_statement, generate_fee_statement_excel, get_performance_levels, get_points, get_teacher_name, get_rank, get_total_fee, get_amount_paid, get_balance, get_class_teacher_comment, get_principal_comment, get_class_teacher_name, get_principal_name
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, Response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from forms import RegistrationForm, ParentRegistrationForm, LinkParentStudentForm, LoginForm, TeacherRegistrationForm, StudentRegistrationForm, MarksForm, FeeForm, MessageForm, PerformanceLevelForm, TermInfoForm, BursarRegistrationForm, MarksFilterForm, FeeFilterForm, ReportCardForm
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import zipfile
import io
import shutil
from utils import get_performance_levels
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from forms import RegistrationForm, AnnouncementsForm, LoginForm, TeacherRegistrationForm, StudentRegistrationForm, MarksForm, FeeForm, MessageForm, PerformanceLevelForm, TermInfoForm, BursarRegistrationForm, MarksFilterForm, FeeFilterForm, ReportCardForm
import pandas as pd
import os
import zipfile
import io
import shutil
from utils import generate_report_card, generate_excel_results, get_rank
import os, logging, io, shutil, zipfile
from flask import Flask, render_template, redirect, url_for, flash, send_file
from flask_login import login_required
from forms import ReportCardForm
import os, logging, io, shutil, zipfile
from flask import Flask, render_template, redirect, url_for, flash, send_file
from flask_login import login_required
from forms import ReportCardForm
from decorators import admin_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from utils import generate_student_password
from forms import BulkStudentUploadForm, ResultsFilterForm
from flask import Flask, render_template, flash, g, redirect, url_for
import timeout_decorator
import traceback
from flask import Blueprint, render_template, flash, redirect, url_for
import subprocess
import tempfile
from typing import List, Tuple
from flask import Flask, render_template, request, jsonify, flash, redirect, url_for, send_file
from flask_login import LoginManager, UserMixin, login_user, login_required, current_user
from flask_sqlalchemy import SQLAlchemy
import logging
import traceback
from io import BytesIO
from functools import wraps
from forms import ReportCardForm, NoteForm # Ensure this import is present
from utils import generate_excel_results, create_zipped_report_cards
import openpyxl
from utils import generate_individual_report_card  # Adjust import based on your project structure
from flask import Flask, render_template, flash, redirect, url_for, session, g
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from database import get_db, User, Student, TermInfo, Mission, Vision, About, Contact, ParentStudent, Announcements, Marks
from forms import RegistrationForm, LoginForm, FeeStatementForm, ReportCardForm, FeeForm, LinkParentStudentForm, StudentRegistrationForm
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
import logging
import traceback
import openpyxl
from io import BytesIO
from sqlalchemy.exc import OperationalError, ProgrammingError
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from reportlab.pdfgen import canvas
from database import User
from sqlalchemy import func
from flask import Flask, render_template, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_required, current_user
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func
from database import get_db, User, Student, Marks, Fee, TermInfo, PerformanceLevels, TeacherAssignments, ClassTeachers
import io
import zipfile
import logging
import traceback
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle
from io import BytesIO
from zipfile import ZipFile
import pdfkit
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from sqlalchemy import cast, Integer












app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] =   'postgresql://jonyo-jounior-school-IMS_owner:npg_wdnXJPo7Cj6m@ep-withered-boat-a919os0o-pooler.gwc.azure.neon.tech/jonyo-jounior-school-IMS?sslmode=require&channel_binding=require'
app.config['WTF_CSRF_ENABLED'] = False

# Initialize logging
logging.basicConfig(filename='app.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Initialize SQLAlchemy and bind to app
init_db(app)

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

GRADES = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')]
TERMS = [('Term 1', 'Term 1'), ('Term 2', 'Term 2'), ('Term 3', 'Term 3')]
EXAM_TYPES = [
    ('cat1', 'CAT 1'), ('cat2', 'CAT 2'), ('cat3', 'CAT 3'),
    ('rat1', 'RAT 1'), ('rat2', 'RAT 2'), ('rat3', 'RAT 3'),
    ('midterm', 'Mid Term'), ('endterm', 'End Term'),
    ('project1', 'Project 1'), ('project2', 'Project 2'), ('project3', 'Project 3')
]


@app.template_filter('datetimeformat')
def datetimeformat(value):
    try:
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, str):
            # Try parsing string to datetime
            return datetime.strptime(value, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
        return str(value)
    except (ValueError, TypeError):
        return str(value)
    
# Custom Jinja2 filter for currency formatting
def format_currency(value):
    try:
        return f"KSh {float(value):,.2f}"
    except (ValueError, TypeError):
        return value

app.jinja_env.filters['format_currency'] = format_currency




# Custom User class for Flask-Login
class CurrentUser(UserMixin):
    def __init__(self, id, admission_no, username, role, grade=None):
        self.id = id
        self.admission_no = admission_no
        self.username = username
        self.role = role
        self.grade = grade

@login_manager.user_loader
def load_user(user_id):
    try:
        logger.debug(f"Querying User with id: {user_id}")
        with app.app_context():
            user = db.session.get(User, int(user_id))
            if user:
                return CurrentUser(
                    id=user.id,
                    admission_no=user.admission_no,
                    username=user.username,
                    role=user.role,
                    grade=user.grade
                )
        return None
    except Exception as e:
        logger.error(f"Error in load_user: {str(e)}\n{traceback.format_exc()}")
        return None

# Role-based decorators
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'teacher':
            flash('Access denied!', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def bursar_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'bursar':
            flash('Bursar access required!', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def parent_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'parent':
            flash('Access denied: Parents only.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'student':
            flash('Access denied: Students only.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def admin_or_bursar_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['admin', 'bursar']:
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Utility functions
def get_grades():
    """Fetch distinct grades from Student and User tables (students only)."""
    try:
        with app.app_context():
            db_session = next(get_db())

            # Query grades from Student table
            student_grades = db_session.query(Student.grade).filter(Student.grade.isnot(None)).distinct().all()

            # Query grades from User table where role is student
            user_grades = db_session.query(User.grade).filter(
                User.role == 'student',
                User.grade.isnot(None)
            ).distinct().all()

            # Combine and deduplicate
            grades = set()
            for g in student_grades + user_grades:
                if g[0]:
                    grades.add(g[0])

            if not grades:
                logger.warning("No grades found in DB, using default fallback.")
                return [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')]

            # Return sorted choices
            sorted_grades = sorted(grades, key=lambda x: (len(x), x))
            logger.debug(f"Fetched grades from DB: {sorted_grades}")
            return [(grade, grade) for grade in sorted_grades]

    except Exception as e:
        logger.error(f"Error fetching grades: {str(e)}\n{traceback.format_exc()}")
        return [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')]  # Fallback default

def get_term_info():
    """Fetches the latest term information from the term_info table."""
    try:
        with app.app_context():
            db_session = next(get_db())
            term = db_session.query(TermInfo).order_by(TermInfo.year.desc(), TermInfo.term.desc()).first()
            if term:
                return {
                    'year': term.year or '2025',
                    'term': term.term or 'Term 1',
                    'principal': term.principal or 'Mr. Principal',
                    'start_date': term.start_date or '2025-01-01',
                    'end_date': term.end_date or '2025-04-01'
                }
            logger.debug("No term_info found in database")
            return {
                'year': '2025',
                'term': 'Term 1',
                'principal': 'Mr. Principal',
                'start_date': '2025-01-01',
                'end_date': '2025-04-01'
            }
    except Exception as e:
        logger.error(f"Error fetching term_info: {str(e)}\n{traceback.format_exc()}")
        return {
            'year': '2025',
            'term': 'Term 1',
            'principal': 'Mr. Principal',
            'start_date': '2025-01-01',
            'end_date': '2025-04-01'
        }

def fetch_common_data():
    db_session = next(get_db())
    try:
        # Fetch term_info
        term_data = db_session.query(TermInfo).filter_by(id=1).first()
        term_info = {
            'term': term_data.term if term_data and term_data.term else 'Term 1',
            'year': term_data.year if term_data and term_data.year else '2025',
            'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
            'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
            'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
        }

        # Fetch content_data
        content_types = ['mission', 'vision', 'about', 'contact']
        content_data = {}
        for content_type in content_types:
            Model = {'mission': Mission, 'vision': Vision, 'about': About, 'contact': Contact}.get(content_type)
            result = db_session.query(Model.content).filter_by(id=1).first()
            content_data[content_type] = result.content if result else ""

        return term_info, content_data
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in fetch_common_data: {str(e)}\n{traceback.format_exc()}")
        return {}, {}
    finally:
        db_session.close()

def generate_fee_statement_pdf(fees, grade, term, year):
    """Generate PDF file for fee statements."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"Fee Statement - {grade} {term} {year}", styles['Heading1'])
    elements.append(title)

    data = [['Admission No', 'Student Name', 'Total Fee', 'Amount Paid', 'Balance', 'Grade', 'Term', 'Year']]
    for fee in fees:
        data.append(list(fee))

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    return buffer


def migrate_db():
    """Migrate database schema for users, marks, and announcements tables."""
    try:
        with app.app_context():
            db_session = next(get_db())
            # Add phone_number column to users table
            try:
                db_session.execute('ALTER TABLE users ADD COLUMN phone_number TEXT')
                db_session.commit()
                logger.info("Successfully added phone_number column to users table")
            except (OperationalError, ProgrammingError) as e:
                if "duplicate column name" not in str(e).lower():
                    logger.error(f"Error migrating users table: {str(e)}\n{traceback.format_exc()}")
                    db_session.rollback()

            # Create marks table if it doesn't exist
            try:
                db_session.execute('''
                    CREATE TABLE IF NOT EXISTS marks (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        admission_no TEXT NOT NULL,
                        learning_area TEXT NOT NULL,
                        marks INTEGER NOT NULL,
                        exam_type TEXT NOT NULL,
                        total_marks INTEGER,
                        term TEXT NOT NULL,
                        year INTEGER NOT NULL,
                        grade TEXT NOT NULL,
                        FOREIGN KEY (admission_no) REFERENCES students(admission_no),
                        CONSTRAINT unique_marks UNIQUE (admission_no, learning_area, exam_type, term, year, grade)
                    )
                ''')
                db_session.commit()
                logger.info("Successfully created marks table")
            except (OperationalError, ProgrammingError) as e:
                if "already exists" not in str(e).lower():
                    logger.error(f"Error creating marks table: {str(e)}\n{traceback.format_exc()}")
                    db_session.rollback()

            # Migrate announcements table
            try:
                db_session.execute('ALTER TABLE announcements ADD COLUMN id INTEGER PRIMARY KEY AUTOINCREMENT')
                db_session.commit()
                logger.info("Successfully added id column to announcements table")
            except (OperationalError, ProgrammingError) as e:
                if "duplicate column name" not in str(e).lower():
                    logger.error(f"Error migrating announcements table: {str(e)}\n{traceback.format_exc()}")
                    db_session.rollback()
                    # Recreate announcements table if migration fails
                    db_session.execute('DROP TABLE IF EXISTS announcements')
                    db_session.execute('''
                        CREATE TABLE announcements (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            content TEXT NOT NULL,
                            date TEXT NOT NULL
                        )
                    ''')
                    db_session.commit()
                    logger.info("Recreated announcements table")
    except Exception as e:
        logger.error(f"Unexpected error in migrate_db: {str(e)}\n{traceback.format_exc()}")
        db_session.rollback()

def generate_fee_statement_excel(fees, grade, term=None, year=None):
    """Generate an Excel file from fees data."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Fee Statement - Grade {grade if grade != 'all' else 'All'}"

        # Define headers
        headers = ['Admission No', 'Student Name', 'Total Fee', 'Amount Paid', 'Balance', 'Grade', 'Term', 'Year']
        sheet.append(headers)

        # Style headers
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add fee data
        for fee in fees:
            sheet.append(list(fee))

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # Save to BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}\n{traceback.format_exc()}")
        return None

# Public Routes
@app.route('/')
def index():
    """Render the homepage with announcements and school information."""
    try:
        with app.app_context():
            db_session = next(get_db())
            # Fetch announcements
            announcements = db_session.query(Announcements).order_by(Announcements.date.desc()).all()
            announcements = [{'id': a.id, 'content': a.content, 'date': a.date} for a in announcements]

            # Fetch term_info and content_data using helper function
            term_info, content_data = fetch_common_data()

            return render_template(
                'index.html',
                announcements=announcements,
                term_info=term_info,
                content_data=content_data
            )
    except Exception as e:
        logger.error(f"Error in index route: {str(e)}\n{traceback.format_exc()}")
        flash('An error occurred while loading the homepage.', 'danger')
        return render_template(
            'index.html',
            announcements=[],
            term_info={},
            content_data={'mission': '', 'vision': '', 'about': '', 'contact': ''}
        )
    finally:
        if 'db_session' in locals():
            db_session.close()

@app.route('/about')
def about():
    """Render the about page with school information."""
    try:
        with app.app_context():
            db_session = next(get_db())
            # Fetch content for mission, vision, about, contact
            content = {}
            for model, content_type in [(Mission, 'mission'), (Vision, 'vision'), (About, 'about'), (Contact, 'contact')]:
                result = db_session.query(model).filter_by(id=1).first()
                content[content_type + '_content'] = result.content if result else ""
                logger.debug(f"Fetched {content_type}_content: {content[content_type + '_content']}")

            # Fetch term_info
            term_info = get_term_info()

            return render_template('about.html', term_info=term_info, **content)
    except Exception as e:
        logger.error(f"Error in about route: {str(e)}\n{traceback.format_exc()}")
        flash('An error occurred while loading the about page.', 'danger')
        return render_template('about.html', term_info={}, mission_content="", vision_content="", about_content="", contact_content="")

@app.route('/contact')
def contact():
    """Render the contact page with school information."""
    try:
        with app.app_context():
            db_session = next(get_db())
            # Fetch content for mission, vision, about, contact
            content = {}
            for model, content_type in [(Mission, 'mission'), (Vision, 'vision'), (About, 'about'), (Contact, 'contact')]:
                result = db_session.query(model).filter_by(id=1).first()
                content[content_type + '_content'] = result.content if result else ""
                logger.debug(f"Fetched {content_type}_content: {content[content_type + '_content']}")

            # Fetch term_info
            term_info = get_term_info()

            return render_template('contact.html', term_info=term_info, **content)
    except Exception as e:
        logger.error(f"Error in contact route: {str(e)}\n{traceback.format_exc()}")
        flash('An error occurred while loading the contact page.', 'danger')
        return render_template('contact.html', term_info={}, mission_content="", vision_content="", about_content="", contact_content="")

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Register a new admin or parent user."""
    form = RegistrationForm()
    try:
        with db.session() as db_session:
            # Prepare content_data for template
            mission = db_session.query(Mission).first()
            vision = db_session.query(Vision).first()
            about = db_session.query(About).first()
            content_data = {
                'mission': mission.content if mission else "To provide quality education for all students",
                'vision': vision.content if vision else "To be a leading institution in academic excellence",
                'about': about.content if about else "Jonyo Junior School is dedicated to fostering holistic education"
            }

            if request.method == 'POST' and form.validate_on_submit():
                # Check if username exists
                if db_session.query(User).filter_by(username=form.username.data).first():
                    flash('Username already exists. Please choose a different username.', 'error')
                    return render_template('register.html', form=form, content_data=content_data)

                # Block parent registration
                if form.role.data == 'parent':
                    flash('Cannot register as parent. Please contact the system admin for registration.', 'error')
                    return render_template('register.html', form=form, content_data=content_data)

                # Check admin count (limit to 3)
                admin_count = db_session.query(User).filter_by(role='admin').count()
                if form.role.data == 'admin' and admin_count >= 3:
                    flash('Maximum number of admins (3) reached. Contact support for assistance.', 'error')
                    return render_template('register.html', form=form, content_data=content_data)

                # Create new user
                user = User(
                    username=form.username.data,
                    password_hash=generate_password_hash(form.password.data),
                    role=form.role.data,
                    grade=None,
                    admission_no=None,
                    phone_number=None
                )
                db_session.add(user)
                db_session.commit()
                logger.info(f"User {form.username.data} registered successfully with role {form.role.data}")
                flash('Registration successful! Please log in.', 'success')
                return redirect(url_for('login'))
            
            return render_template('register.html', form=form, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in register route: {str(e)}\n{traceback.format_exc()}")
        flash(f'Registration failed: {str(e)}. Please try again or contact support.', 'error')
        return render_template('register.html', form=form, content_data={
            'mission': "To provide quality education for all students",
            'vision': "To be a leading institution in academic excellence",
            'about': "Jonyo Junior School is dedicated to fostering holistic education"
        })
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login."""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = LoginForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

    if form.validate_on_submit():
        db_session = next(get_db())
        try:
            user = db_session.query(User).filter_by(username=form.username.data, role=form.role.data).first()
            if user and check_password_hash(user.password_hash, form.password.data):
                login_user(CurrentUser(
                    id=user.id,
                    admission_no=user.admission_no,
                    username=user.username,
                    role=user.role,
                    grade=user.grade
                ))
                flash('Login successful!', 'success')
                logger.debug(f"User {user.username} logged in with role {user.role}")
                return redirect(url_for('dashboard'))
            else:
                flash('Invalid username, password, or role.', 'danger')
        except SQLAlchemyError as e:
            db_session.rollback()
            logger.error(f"Database error in login: {str(e)}\n{traceback.format_exc()}")
            flash('An error occurred while logging in.', 'danger')
        finally:
            db_session.close()
    
    return render_template('login.html', form=form, term_info=term_info, content_data=content_data)
@app.route('/logout')
@login_required
def logout():
    """Handle user logout."""
    logout_user()
    flash('Logged out successfully!', 'success')
    return redirect(url_for('index'))

# Dashboard Routes
from flask import render_template, request, flash, redirect, url_for
from flask_login import login_required, current_user
from datetime import datetime
import logging
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func

logger = logging.getLogger(__name__)

# Assuming these are defined in your app.py or imported
GRADES = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')]
TERMS = [('Term 1', 'Term 1'), ('Term 2', 'Term 2'), ('Term 3', 'Term 3')]
EXAM_TYPES = [
    ('cat1', 'CAT 1'), ('cat2', 'CAT 2'), ('cat3', 'CAT 3'),
    ('rat1', 'RAT 1'), ('rat2', 'RAT 2'), ('rat3', 'RAT 3'),
    ('midterm', 'Mid Term'), ('endterm', 'End Term'),
    ('project1', 'Project 1'), ('project2', 'Project 2'), ('project3', 'Project 3')
]

@app.route('/dashboard', methods=['GET', 'POST'])  # Updated to handle POST for form submissions
@login_required
def dashboard():
    """Render role-specific dashboards."""
    db_session = next(get_db())
    try:
        logger.debug(f"Accessing dashboard for user: id={current_user.id}, username={current_user.username}, role={current_user.role}")

        templates = {
            'admin': 'admin_dashboard.html',
            'teacher': 'teacher_dashboard.html',
            'student': 'student_dashboard.html',
            'parent': 'parent_dashboard.html',
            'bursar': 'bursar_dashboard.html'
        }

        template = templates.get(current_user.role)
        if not template:
            logger.error(f"User {current_user.username} has invalid role: {current_user.role}")
            flash('Access denied: Invalid role.', 'danger')
            return redirect(url_for('logout'))

        # Fetch term_info and content_data
        term_info, content_data = fetch_common_data()

        # Handle student-specific logic
        if current_user.role == 'student':
            student = db_session.query(Student).filter_by(admission_no=current_user.admission_no).first()
            if not student:
                flash('Student record not found.', 'danger')
                logger.warning(f"No student found for admission_no={current_user.admission_no}")
                return redirect(url_for('logout'))

            student_data = {
                'admission_no': student.admission_no,
                'name': student.name,
                'grade': student.grade
            }

            # Fetch grades for form choices
            grades = db_session.query(Marks.grade).filter_by(admission_no=student.admission_no).distinct().order_by(Marks.grade).all()
            grades = [(g[0], g[0]) for g in grades if g[0] in [grade[0] for grade in GRADES]] or GRADES

            # Initialize forms
            report_form = ReportCardForm(admission_no=student.admission_no, grade=student.grade)
            report_form.admission_no.choices = [(student.admission_no, f"{student.name} ({student.admission_no})")]
            report_form.admission_no.data = student.admission_no
            report_form.grade.choices = grades
            report_form.grade.data = student.grade if student.grade in [g[0] for g in grades] else grades[0][0]
            report_form.term.choices = TERMS
            report_form.term.data = term_info['term'] if term_info['term'] in [t[0] for t in TERMS] else TERMS[0][0]
            report_form.exam_type.choices = EXAM_TYPES
            report_form.exam_type.data = 'endterm' if 'endterm' in [e[0] for e in EXAM_TYPES] else EXAM_TYPES[0][0]
            report_form.year.data = int(term_info['year']) if term_info['year'].isdigit() else datetime.now().year

            fee_form = FeeStatementForm(admission_no=student.admission_no, grade=student.grade)
            fee_form.admission_no.choices = [(student.admission_no, f"{student.name} ({student.admission_no})")]
            fee_form.admission_no.data = student.admission_no
            fee_form.grade.choices = grades
            fee_form.grade.data = student.grade if student.grade in [g[0] for g in grades] else grades[0][0]
            fee_form.term.choices = TERMS
            fee_form.term.data = term_info['term'] if term_info['term'] in [t[0] for t in TERMS] else TERMS[0][0]
            fee_form.year.data = term_info['year'] if term_info['year'].isdigit() else str(datetime.now().year)

            view_results_form = ResultsFilterForm(admission_no=student.admission_no, grade=student.grade)
            view_results_form.grade.choices = grades
            view_results_form.grade.data = student.grade if student.grade in [g[0] for g in grades] else grades[0][0]
            view_results_form.term.choices = TERMS
            view_results_form.term.data = term_info['term'] if term_info['term'] in [t[0] for t in TERMS] else TERMS[0][0]
            view_results_form.exam_type.choices = EXAM_TYPES
            view_results_form.exam_type.data = 'endterm' if 'endterm' in [e[0] for e in EXAM_TYPES] else EXAM_TYPES[0][0]
            view_results_form.year.data = int(term_info['year']) if term_info['year'].isdigit() else datetime.now().year

            # Handle form submissions
            if request.method == 'POST':
                logger.debug(f"POST data: {request.form}")
                if report_form.submit.data and report_form.validate_on_submit():
                    logger.info(f"Redirecting to download_report_card: admission_no={report_form.admission_no.data}, grade={report_form.grade.data}, term={report_form.term.data}, year={report_form.year.data}, exam_type={report_form.exam_type.data}")
                    return redirect(url_for('student_download_report_card',
                                            admission_no=student.admission_no,
                                            grade=report_form.grade.data,
                                            term=report_form.term.data,
                                            year=report_form.year.data,
                                            exam_type=report_form.exam_type.data))
                elif fee_form.submit.data and fee_form.validate_on_submit():
                    logger.info(f"Redirecting to download_fee_statement: admission_no={fee_form.admission_no.data}, grade={fee_form.grade.data}, term={fee_form.term.data}, year={fee_form.year.data}")
                    return redirect(url_for('student_download_fee_statement',
                                            admission_no=student.admission_no,
                                            grade=fee_form.grade.data,
                                            term=fee_form.term.data,
                                            year=fee_form.year.data))
                elif view_results_form.submit.data and view_results_form.validate_on_submit():
                    logger.info(f"Redirecting to view_results: admission_no={student.admission_no}, grade={view_results_form.grade.data}, term={view_results_form.term.data}, year={view_results_form.year.data}, exam_type={view_results_form.exam_type.data}")
                    return redirect(url_for('student_view_results',
                                            admission_no=student.admission_no,
                                            grade=view_results_form.grade.data,
                                            term=view_results_form.term.data,
                                            year=view_results_form.year.data,
                                            exam_type=view_results_form.exam_type.data))
                else:
                    for form, form_name in [(report_form, 'ReportCardForm'), (fee_form, 'FeeStatementForm'), (view_results_form, 'ResultsFilterForm')]:
                        for field, errors in form.errors.items():
                            for error in errors:
                                logger.error(f"{form_name} error in {field}: {error}, submitted value: {request.form.get(field, 'None')}")
                                flash(f"{form_name} error: {error}", 'danger')

            # Fetch announcements
            recent_announcements = db_session.query(Announcements).order_by(Announcements.date.desc()).limit(5).all()
            recent_announcements = [{'content': a.content, 'date': a.date} for a in recent_announcements]

            return render_template(
                template,
                student=student_data,
                recent_announcements=recent_announcements,
                report_form=report_form,
                fee_form=fee_form,
                view_results_form=view_results_form,
                term_info=term_info,
                content_data=content_data,
                parent_view=False
            )

        # Handle parent-specific logic
        elif current_user.role == 'parent':
            report_form = ReportCardForm()
            linked_students = db_session.query(Student.admission_no, Student.name, Student.grade).\
                join(ParentStudent, Student.admission_no == ParentStudent.admission_no).\
                filter(ParentStudent.parent_id == current_user.id).all()
            report_form.admission_no.choices = [
                (s.admission_no, f"{s.name} ({s.admission_no})") for s in linked_students
            ] if linked_students else []
            fee_form = FeeStatementForm()
            fee_form.admission_no.choices = report_form.admission_no.choices
            link_form = LinkParentStudentForm()
            view_results_form = ResultsFilterForm()

            # Handle form submissions
            if request.method == 'POST':
                logger.debug(f"POST data: {request.form}")
                if report_form.submit.data and report_form.validate_on_submit():
                    logger.info(f"Parent redirecting to download_report_card: admission_no={report_form.admission_no.data}")
                    return redirect(url_for('student_download_report_card',
                                            admission_no=report_form.admission_no.data,
                                            grade=report_form.grade.data,
                                            term=report_form.term.data,
                                            year=report_form.year.data,
                                            exam_type=report_form.exam_type.data))
                elif fee_form.submit.data and fee_form.validate_on_submit():
                    logger.info(f"Parent redirecting to download_fee_statement: admission_no={fee_form.admission_no.data}")
                    return redirect(url_for('student_download_fee_statement',
                                            admission_no=fee_form.admission_no.data,
                                            grade=fee_form.grade.data,
                                            term=fee_form.term.data,
                                            year=fee_form.year.data))
                elif view_results_form.submit.data and view_results_form.validate_on_submit():
                    logger.info(f"Parent redirecting to view_results: admission_no={view_results_form.admission_no.data}")
                    return redirect(url_for('student_view_results',
                                            admission_no=report_form.admission_no.data,
                                            grade=view_results_form.grade.data,
                                            term=view_results_form.term.data,
                                            year=view_results_form.year.data,
                                            exam_type=view_results_form.exam_type.data))
                elif link_form.validate_on_submit():
                    # Handle parent-student linking logic (assumed to be defined elsewhere)
                    pass

            return render_template(
                template,
                report_form=report_form,
                link_form=link_form,
                fee_form=fee_form,
                view_results_form=view_results_form,
                term_info=term_info,
                content_data=content_data,
                parent_view=True
            )

        # Default for other roles (admin, teacher, bursar)
        return render_template(
            template,
            term_info=term_info,
            content_data=content_data
        )

    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in dashboard: {str(e)}\n{traceback.format_exc()}")
        flash('An unexpected error occurred while loading the dashboard.', 'danger')
        return render_template(
            template or 'index.html',
            term_info={},
            content_data={'mission': '', 'vision': '', 'about': '', 'contact': ''},
            report_form=ReportCardForm(),
            fee_form=FeeStatementForm(),
            view_results_form=ResultsFilterForm(),
            link_form=LinkParentStudentForm() if current_user.role == 'parent' else None,
            student=None,
            marks=[],
            recent_announcements=[],
            parent_view=current_user.role == 'parent'
        )
    finally:
        db_session.close()

@app.route('/admin_dashboard')
@login_required
@admin_required
def admin_dashboard():
    """Render the admin dashboard with school information and recent announcements."""
    try:
        with app.app_context():
            db_session = next(get_db())
            content_types = ['mission', 'vision', 'about', 'contact']
            content_data = {}
            for model, content_type in [(Mission, 'mission'), (Vision, 'vision'), (About, 'about'), (Contact, 'contact')]:
                result = db_session.query(model).filter_by(id=1).first()
                content_data[content_type] = result.content if result else ""
                logger.debug(f"Fetched {content_type}_content: {content_data[content_type]}")

            recent_announcements = db_session.query(Announcements).order_by(Announcements.date.desc()).limit(5).all()
            recent_announcements = [{'id': a.id, 'content': a.content, 'date': a.date} for a in recent_announcements]

            return render_template(
                'admin_dashboard.html',
                content_data=content_data,  # Pass content_data to the template
                mission_content=content_data['mission'],
                vision_content=content_data['vision'],
                about_content=content_data['about'],
                contact_content=content_data['contact'],
                recent_announcements=recent_announcements
            )
    except Exception as e:
        logger.error(f"Error in admin_dashboard: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error loading dashboard: {str(e)}', 'danger')
        return render_template(
            'admin_dashboard.html',
            content_data={  # Pass content_data in the error case as well
                'mission': 'Error loading content.',
                'vision': 'Error loading content.',
                'about': 'Error loading content.',
                'contact': 'Error loading content.'
            },
            mission_content="Error loading content.",
            vision_content="Error loading content.",
            about_content="Error loading content.",
            contact_content="Error loading content.",
            recent_announcements=[]
        )
@app.route('/parent_dashboard', methods=['GET'])
@login_required
def parent_dashboard():
    """Render the parent dashboard showing all linked students."""
    if current_user.role != 'parent':
        logger.warning(f"Unauthorized access to parent_dashboard by user {current_user.id} with role {current_user.role}")
        flash('You are not authorized to access this page.', 'danger')
        return redirect(url_for('index'))

    db_session = next(get_db())
    try:
        # Log user details for debugging
        parent_name = getattr(current_user, 'name', 'Parent') or 'Parent'
        logger.info(f"Fetching data for parent ID: {current_user.id}, role: {current_user.role}, name: {parent_name}")

        # Verify parent exists in users table
        parent = db_session.query(User).filter_by(id=current_user.id, role='parent').first()
        if not parent:
            logger.error(f"No user found with ID {current_user.id} and role 'parent'")
            flash('User account not found. Please contact the school administration.', 'danger')
            return redirect(url_for('index'))

        # Update parent_name if available in database
        parent_name = parent.name if parent.name else 'Parent'
        logger.debug(f"Parent name from database: {parent_name}")

        # Fetch all students linked to the parent
        linked_students = db_session.query(Student.admission_no, Student.name, Student.grade).\
            join(ParentStudent, Student.admission_no == ParentStudent.admission_no).\
            filter(ParentStudent.parent_id == current_user.id).all()
        
        # Log raw query result
        logger.debug(f"Raw query result for linked students: {linked_students}")

        # Prepare student data for template
        linked_students = [
            {
                'admission_no': s.admission_no,
                'name': s.name if s.name else 'Unknown',
                'grade': s.grade if s.grade else 'N/A'
            }
            for s in linked_students
        ]

        # Log processed linked students
        logger.info(f"Processed linked students: {linked_students}")

        if not linked_students:
            # Check if parent_id exists in ParentStudent table
            parent_exists = db_session.query(ParentStudent).filter_by(parent_id=current_user.id).first()
            logger.info(f"Parent ID {current_user.id} exists in ParentStudent: {bool(parent_exists)}")
            if parent_exists:
                # Check if admission_no exists in Student table
                linked_admission_nos = db_session.query(ParentStudent.admission_no).filter_by(parent_id=current_user.id).all()
                logger.debug(f"Linked admission numbers: {linked_admission_nos}")
                for admission_no in linked_admission_nos:
                    student_exists = db_session.query(Student).filter_by(admission_no=admission_no[0]).first()
                    logger.debug(f"Student exists for admission_no {admission_no[0]}: {bool(student_exists)}")
                flash('No valid students found linked to your account. Please verify with the school administration.', 'warning')
            else:
                logger.warning(f"No ParentStudent records found for parent ID {current_user.id}")
                flash('No students are linked to your account. Please contact the school administration to link your account with a student.', 'warning')
        else:
            # Auto-redirect if only one student is linked
            if len(linked_students) == 1:
                logger.info(f"Auto-redirecting to single linked student dashboard for admission_no: {linked_students[0]['admission_no']}")
                return redirect(url_for('student_dashboard', admission_no=linked_students[0]['admission_no']))

        return render_template(
            'parent_dashboard.html',
            linked_students=linked_students,
            parent_name=parent_name
        )
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in parent_dashboard for user {current_user.id}: {str(e)}\n{traceback.format_exc()}")
        flash('Database error: Unable to load dashboard. Please try again later.', 'danger')
        return render_template(
            'parent_dashboard.html',
            linked_students=[],
            parent_name=parent_name
        )
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in parent_dashboard for user {current_user.id}: {str(e)}\n{traceback.format_exc()}")
        flash('An unexpected error occurred. Please try again later.', 'danger')
        return render_template(
            'parent_dashboard.html',
            linked_students=[],
            parent_name=parent_name
        )
    finally:
        db_session.close()
from flask import render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from sqlalchemy.exc import SQLAlchemyError
import traceback
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

@app.route('/student_dashboard', methods=['GET', 'POST'])
@app.route('/student_dashboard/<admission_no>', methods=['GET', 'POST'])
@login_required
def student_dashboard(admission_no=None):
    """Render the student dashboard for students or parents, with academic and fee information."""
    db_session = next(get_db())
    try:
        # Determine user role and admission number
        if current_user.role == 'student':
            if admission_no and admission_no != current_user.admission_no:
                logger.warning(f"Student {current_user.id} attempted to access dashboard for admission_no={admission_no}")
                flash('You can only view your own dashboard.', 'danger')
                return redirect(url_for('student_dashboard'))
            admission_no = current_user.admission_no
            parent_view = False
        elif current_user.role == 'parent':
            if not admission_no:
                logger.warning(f"Parent {current_user.id} accessed /student_dashboard without admission_no")
                flash('Please select a linked student.', 'danger')
                return redirect(url_for('parent_dashboard'))
            parent_student = db_session.query(ParentStudent).filter_by(
                parent_id=current_user.id,
                admission_no=admission_no
            ).first()
            if not parent_student:
                logger.warning(f"Parent {current_user.id} not linked to admission_no={admission_no}")
                flash('You are not authorized to view this student’s dashboard.', 'danger')
                return redirect(url_for('parent_dashboard'))
            parent_view = True
        else:
            logger.error(f"Unauthorized role {current_user.role} for user {current_user.id}")
            flash('You are not authorized to access this page.', 'danger')
            return redirect(url_for('index'))

        # Fetch student data
        student = db_session.query(Student).filter_by(admission_no=admission_no).first()
        if not student:
            logger.warning(f"No student found for admission_no={admission_no}")
            flash('Student profile not found.', 'danger')
            return redirect(url_for('logout' if current_user.role == 'student' else 'parent_dashboard'))

        student_data = {
            'admission_no': student.admission_no,
            'name': student.name,
            'grade': student.grade
        }

        # Fetch marks
        marks = db_session.query(Marks).filter_by(admission_no=student.admission_no).order_by(Marks.year.desc(), Marks.term.desc()).limit(10).all()
        marks = [{
            'learning_area': m.learning_area,
            'total_marks': m.total_marks,
            'exam_type': m.exam_type,
            'term': m.term,
            'year': m.year,
            'grade': m.grade
        } for m in marks]

        # Fetch term_info
        term_data = db_session.query(TermInfo).filter_by(id=1).first()
        term_info = {
            'term': term_data.term if term_data and term_data.term else 'Term 1',
            'year': term_data.year if term_data and term_data.year else '2025',
            'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
            'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
            'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
        }
        content_data = fetch_common_data()[1]

        # Fetch grades for form choices
        grades = db_session.query(Marks.grade).filter_by(admission_no=admission_no).distinct().order_by(Marks.grade).all()
        grades = [(g[0], g[0]) for g in grades if g[0] in [grade[0] for grade in GRADES]] or GRADES

        # Fetch linked students for parent view
        linked_students = []
        if parent_view:
            linked_students = db_session.query(Student.admission_no, Student.name, Student.grade).\
                join(ParentStudent, Student.admission_no == ParentStudent.admission_no).\
                filter(ParentStudent.parent_id == current_user.id).all()
            linked_students = [
                {
                    'admission_no': s.admission_no,
                    'name': s.name if s.name else 'Unknown',
                    'grade': s.grade if s.grade else 'N/A'
                }
                for s in linked_students
            ]

        # Initialize forms
        report_form = ReportCardForm(admission_no=student.admission_no, grade=student.grade)
        report_form.admission_no.choices = [(student.admission_no, f"{student.name} ({student.admission_no})")]
        report_form.admission_no.data = student.admission_no
        report_form.grade.choices = grades
        report_form.grade.data = student.grade if student.grade in [g[0] for g in grades] else grades[0][0]
        report_form.term.choices = TERMS
        report_form.term.data = term_info['term'] if term_info['term'] in [t[0] for t in TERMS] else TERMS[0][0]
        report_form.exam_type.choices = EXAM_TYPES
        report_form.exam_type.data = 'endterm' if 'endterm' in [e[0] for e in EXAM_TYPES] else EXAM_TYPES[0][0]
        report_form.year.data = int(term_info['year']) if term_info['year'].isdigit() else datetime.now().year

        fee_form = FeeStatementForm(admission_no=student.admission_no, grade=student.grade)
        if parent_view:
            fee_form.admission_no.choices = [
                (s.admission_no, f"{s.name} ({s.admission_no})")
                for s in linked_students
            ]
            fee_form.admission_no.data = admission_no
        else:
            fee_form.admission_no.choices = [(student.admission_no, f"{student.name} ({student.admission_no})")]
            fee_form.admission_no.data = student.admission_no
        fee_form.grade.choices = grades
        fee_form.grade.data = student.grade if student.grade in [g[0] for g in grades] else grades[0][0]
        fee_form.term.choices = TERMS
        fee_form.term.data = term_info['term'] if term_info['term'] in [t[0] for t in TERMS] else TERMS[0][0]
        fee_form.year.data = term_info['year'] if term_info['year'].isdigit() else str(datetime.now().year)

        view_results_form = ResultsFilterForm(admission_no=student.admission_no, grade=student.grade)
        view_results_form.grade.choices = grades
        view_results_form.admission_no.choices = [(student.admission_no, f"{student.name} ({student.admission_no})")]
        view_results_form.admission_no.data = student.admission_no
        view_results_form.grade.data = student.grade if student.grade in [g[0] for g in grades] else grades[0][0]
        view_results_form.term.choices = TERMS
        view_results_form.term.data = term_info['term'] if term_info['term'] in [t[0] for t in TERMS] else TERMS[0][0]
        view_results_form.exam_type.choices = EXAM_TYPES
        view_results_form.exam_type.data = 'endterm' if 'endterm' in [e[0] for e in EXAM_TYPES] else EXAM_TYPES[0][0]
        view_results_form.year.data = int(term_info['year']) if term_info['year'].isdigit() else datetime.now().year

        # Handle form submissions
        if request.method == 'POST':
            logger.debug(f"POST data: {request.form}")
            if report_form.submit.data and report_form.validate_on_submit():
                logger.info(f"Redirecting to download_report_card: admission_no={report_form.admission_no.data}, grade={report_form.grade.data}, term={report_form.term.data}, year={report_form.year.data}, exam_type={report_form.exam_type.data}")
                return redirect(url_for('student_download_report_card',
                                        admission_no=report_form.admission_no.data,
                                        grade=report_form.grade.data,
                                        term=report_form.term.data,
                                        year=report_form.year.data,
                                        exam_type=report_form.exam_type.data))
            elif fee_form.submit.data and fee_form.validate_on_submit():
                logger.info(f"Redirecting to download_fee_statement: admission_no={fee_form.admission_no.data}, grade={fee_form.grade.data}, term={fee_form.term.data}, year={fee_form.year.data}")
                return redirect(url_for('student_download_fee_statement',
                                        admission_no=fee_form.admission_no.data,
                                        grade=fee_form.grade.data,
                                        term=fee_form.term.data,
                                        year=fee_form.year.data))
            elif view_results_form.submit.data and view_results_form.validate_on_submit():
                logger.info(f"Redirecting to view_results: admission_no={view_results_form.admission_no.data}, grade={view_results_form.grade.data}, term={view_results_form.term.data}, year={view_results_form.year.data}, exam_type={view_results_form.exam_type.data}")
                return redirect(url_for('student_view_results',
                                        admission_no=view_results_form.admission_no.data,
                                        grade=view_results_form.grade.data,
                                        term=view_results_form.term.data,
                                        year=view_results_form.year.data,
                                        exam_type=view_results_form.exam_type.data))
            else:
                for form, form_name in [(report_form, 'ReportCardForm'), (fee_form, 'FeeStatementForm'), (view_results_form, 'ResultsFilterForm')]:
                    for field, errors in form.errors.items():
                        for error in errors:
                            logger.error(f"{form_name} error in {field}: {error}, submitted value: {request.form.get(field, 'None')}")
                            flash(f"{form_name} error: {error}", 'danger')

        # Fetch announcements
        recent_announcements = db_session.query(Announcements).order_by(Announcements.date.desc()).limit(5).all()
        recent_announcements = [{'content': a.content, 'date': a.date} for a in recent_announcements]

        # Render dashboard
        return render_template(
            'student_dashboard.html',
            student=student_data,
            marks=marks,
            recent_announcements=recent_announcements,
            report_form=report_form,
            fee_form=fee_form,
            view_results_form=view_results_form,
            term_info=term_info,
            content_data=content_data,
            parent_view=parent_view,
            linked_students=linked_students
        )
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in student_dashboard for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        flash('Database error: Unable to load dashboard. Please try again later.', 'danger')
        return render_template(
            'student_dashboard.html',
            student=None,
            marks=[],
            recent_announcements=[],
            report_form=ReportCardForm(),
            fee_form=FeeStatementForm(),
            view_results_form=ResultsFilterForm(),
            term_info={},
            content_data={'mission': '', 'vision': '', 'about': '', 'contact': ''},
            parent_view=parent_view,
            linked_students=[]
        )
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in student_dashboard for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        flash('An unexpected error occurred. Please try again later.', 'danger')
        return render_template(
            'student_dashboard.html',
            student=None,
            marks=[],
            recent_announcements=[],
            report_form=ReportCardForm(),
            fee_form=FeeStatementForm(),
            view_results_form=ResultsFilterForm(),
            term_info={},
            content_data={'mission': '', 'vision': '', 'about': '', 'contact': ''},
            parent_view=parent_view,
            linked_students=[]
        )
    finally:
        db_session.close()

@app.route('/parent/link_student', methods=['GET', 'POST'])
@login_required
def link_student():
    """Link a parent to a student using admission number."""
    if current_user.role != 'parent':
        logger.error(f"Unauthorized access to link_student by user {getattr(current_user, 'id', 'N/A')} with role {getattr(current_user, 'role', 'N/A')}")
        flash('You are not authorized to link students.', 'danger')
        return redirect(url_for('index'))

    # Initialize form with pre-filled parent username
    form = LinkParentStudentForm(parent_id=current_user.username)
    term_info, content_data = fetch_common_data()
    parent_name = current_user.username or f'Parent_{current_user.id}'
    db_session = next(get_db())
    try:
        if request.method == 'POST':
            logger.debug(f"Form submitted with parent_id={form.parent_id.data}, admission_no={form.admission_no.data}, user_username={current_user.username}, user_id={current_user.id}")

            if form.validate_on_submit():
                admission_no = form.admission_no.data.strip()

                # Verify parent_id matches current_user.username
                if form.parent_id.data != current_user.username:
                    logger.warning(f"Parent username mismatch: submitted={form.parent_id.data}, expected={current_user.username}")
                    flash('Invalid parent username. Please use your own username.', 'danger')
                    return render_template('link_student.html', form=form, term_info=term_info, content_data=content_data, parent_name=parent_name)

                # Fetch parent by username to get their ID
                parent = db_session.query(User).filter_by(username=current_user.username, role='parent').first()
                if not parent:
                    logger.warning(f"Parent not found for username={current_user.username}")
                    flash('Parent account not found.', 'danger')
                    return render_template('link_student.html', form=form, term_info=term_info, content_data=content_data, parent_name=parent_name)

                # Check if student exists
                student = db_session.query(Student).filter_by(admission_no=admission_no).first()
                if not student:
                    logger.warning(f"Student not found for admission_no={admission_no}")
                    flash('No student found with this admission number.', 'danger')
                    return render_template('link_student.html', form=form, term_info=term_info, content_data=content_data, parent_name=parent_name)

                # Check if already linked
                existing_link = db_session.query(ParentStudent).filter_by(
                    parent_id=parent.id,
                    admission_no=admission_no
                ).first()
                if existing_link:
                    logger.warning(f"Existing link found for parent_id={parent.id}, admission_no={admission_no}")
                    flash('This student is already linked to your account.', 'warning')
                    return redirect(url_for('parent_dashboard'))

                # Create new link
                new_link = ParentStudent(parent_id=parent.id, admission_no=admission_no)
                db_session.add(new_link)
                db_session.commit()
                logger.info(f"Linked parent_id={parent.id} to admission_no={admission_no}")
                flash(f'Successfully linked student {student.name} to your account.', 'success')
                return redirect(url_for('parent_dashboard'))

            else:
                logger.debug(f"Form validation failed: {form.errors}")
                flash('Form validation failed. Please check your input.', 'danger')

        return render_template('link_student.html', form=form, term_info=term_info, content_data=content_data, parent_name=parent_name)
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in link_student for parent_id={current_user.id}: {str(e)}\n{traceback.format_exc()}")
        flash('Database error: Unable to link student. Please try again later.', 'danger')
        return render_template('link_student.html', form=form, term_info=term_info, content_data=content_data, parent_name=parent_name)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in link_student for parent_id={current_user.id}: {str(e)}\n{traceback.format_exc()}")
        flash('An unexpected error occurred. Please try again later.', 'danger')
        return render_template('link_student.html', form=form, term_info=term_info, content_data=content_data, parent_name=parent_name)
    finally:
        db_session.close()

@app.route('/register_student', methods=['GET', 'POST'])
@login_required
@admin_required
def register_student():
    """Register a new student and create a user account."""
    form = StudentRegistrationForm()
    db_session = next(get_db())
    try:
        # Fetch term_info and content_data
        term_info, content_data = fetch_common_data()

        if form.validate_on_submit():
            admission_no = form.admission_no.data.strip()
            name = form.name.data.strip()
            grade = form.grade.data
            username = name.lower().replace(' ', '_')
            password = generate_student_password(name, admission_no)
            password_hash = generate_password_hash(password)

            try:
                new_student = Student(admission_no=admission_no, name=name, grade=grade)
                new_user = User(
                    username=username,
                    password_hash=password_hash,
                    role='student',
                    admission_no=admission_no,
                    grade=grade
                )
                db_session.add(new_student)
                db_session.add(new_user)
                db_session.commit()
                flash(f'Student {name} registered successfully! Username: {username}, Password: {password}', 'success')
                logger.debug(f"Registered student: {name}, admission_no: {admission_no}, username: {username}")
                return redirect(url_for('dashboard'))
            except IntegrityError as e:
                db_session.rollback()
                error_msg = str(e).lower()
                if "admission_no" in error_msg:
                    flash('Admission number already exists!', 'danger')
                elif "username" in error_msg:
                    flash(f'Username {username} already exists! Try a different name or modify it.', 'danger')
                else:
                    flash(f'Error registering student: {str(e)}', 'danger')
                logger.error(f"Error registering student {name}: {str(e)}\n{traceback.format_exc()}")
            except Exception as e:
                db_session.rollback()
                flash(f'Error registering student: {str(e)}', 'danger')
                logger.error(f"Error registering student {name}: {str(e)}\n{traceback.format_exc()}")

        return render_template('register_student.html', form=form, term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in register_student: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error loading student registration: {str(e)}', 'danger')
        return render_template(
            'register_student.html',
            form=form,
            term_info={},
            content_data={'mission': '', 'vision': '', 'about': '', 'contact': ''}
        )
    finally:
        db_session.close()



@app.route('/admin/register_teacher', methods=['GET', 'POST'])
@login_required
@admin_required
def register_teacher():
    """Register a new teacher with a generated password."""
    form = TeacherRegistrationForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

    if form.validate_on_submit():
        db_session = next(get_db())
        try:
            username = form.username.data.strip()
            password = generate_teacher_password()
            grade = form.grade.data

            teacher_count = db_session.query(User).filter_by(role='teacher').count()
            if teacher_count >= 500:
                flash('Maximum number of teachers (500) reached!', 'danger')
                return redirect(url_for('dashboard'))

            new_user = User(
                username=username,
                password_hash=generate_password_hash(password),
                role='teacher',
                grade=grade
            )
            db_session.add(new_user)
            db_session.commit()
            flash(f'Teacher registered successfully! Username: {username}, Password: {password}', 'success')
            logger.debug(f"Registered teacher: {username}, grade: {grade}")
            return redirect(url_for('dashboard'))
        except IntegrityError:
            db_session.rollback()
            flash('Username already exists!', 'danger')
            logger.error(f"Username {username} already exists")
            return render_template('register_teacher.html', form=form, term_info=term_info, content_data=content_data)
        except Exception as e:
            db_session.rollback()
            flash(f'Error registering teacher: {str(e)}', 'danger')
            logger.error(f"Error registering teacher: {str(e)}\n{traceback.format_exc()}")
            return render_template('register_teacher.html', form=form, term_info=term_info, content_data=content_data)
        finally:
            db_session.close()

    return render_template('register_teacher.html', form=form, term_info=term_info, content_data=content_data)



@app.route('/admin/register_bursar', methods=['GET', 'POST'])
@login_required
@admin_required
def register_bursar():
    """Register a new bursar with a generated password."""
    form = BursarRegistrationForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

    if form.validate_on_submit():
        db_session = next(get_db())
        try:
            username = form.username.data.strip()
            password = generate_teacher_password()

            bursar_count = db_session.query(User).filter_by(role='bursar').count()
            if bursar_count >= 1:
                flash('Only one bursar can be registered!', 'danger')
                return redirect(url_for('dashboard'))

            new_user = User(
                username=username,
                password_hash=generate_password_hash(password),
                role='bursar'
            )
            db_session.add(new_user)
            db_session.commit()
            flash(f'Bursar registered successfully! Username: {username}, Password: {password}', 'success')
            logger.debug(f"Registered bursar: {username}")
            return redirect(url_for('dashboard'))
        except IntegrityError:
            db_session.rollback()
            flash('Username already exists!', 'danger')
            logger.error(f"Username {username} already exists")
            return render_template('register_bursar.html', form=form, term_info=term_info, content_data=content_data)
        except Exception as e:
            db_session.rollback()
            flash(f'Error registering bursar: {str(e)}', 'danger')
            logger.error(f"Error registering bursar: {str(e)}\n{traceback.format_exc()}")
            return render_template('register_bursar.html', form=form, term_info=term_info, content_data=content_data)
        finally:
            db_session.close()

    return render_template('register_bursar.html', form=form, term_info=term_info, content_data=content_data)



@app.route('/register_parent', methods=['GET', 'POST'])
@login_required
@admin_required
def register_parent():
    """Register a new parent with optional phone numbers."""
    form = ParentRegistrationForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

    if form.validate_on_submit():
        db_session = next(get_db())
        try:
            phone_numbers = [num for num in [form.phone_number1.data, form.phone_number2.data, form.phone_number3.data] if num]
            phone_number_str = ','.join(phone_numbers) if phone_numbers else None
            password_hash = generate_password_hash(form.password.data)

            new_user = User(
                username=form.username.data,
                password_hash=password_hash,
                role='parent',
                phone_number=phone_number_str
            )
            db_session.add(new_user)
            db_session.commit()
            flash('Parent registered successfully!', 'success')
            logger.debug(f"Registered parent: {form.username.data}")
            return redirect(url_for('dashboard'))
        except IntegrityError:
            db_session.rollback()
            flash('Username already exists.', 'danger')
            logger.error(f"Username {form.username.data} already exists")
            return render_template('register_parent.html', form=form, term_info=term_info, content_data=content_data)
        except Exception as e:
            db_session.rollback()
            flash('An error occurred. Please try again.', 'danger')
            logger.error(f"Error registering parent: {str(e)}\n{traceback.format_exc()}")
            return render_template('register_parent.html', form=form, term_info=term_info, content_data=content_data)
        finally:
            db_session.close()

    return render_template('register_parent.html', form=form, term_info=term_info, content_data=content_data)




@app.route('/admin/upload_bulk_students', methods=['GET', 'POST'])
@login_required
@admin_required
def upload_bulk_students():
    """Upload multiple students from an Excel file."""
    form = BulkStudentUploadForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

    if form.validate_on_submit():
        db_session = next(get_db())
        try:
            file = form.file.data
            grade = form.grade.data
            logger.debug(f"Grade: {grade}, File size: {file.content_length} bytes")

            file.seek(0, 2)
            file_size = file.tell()
            file.seek(0)
            if file_size > 512 * 1024:
                flash('File too large. Maximum size is 512KB.', 'danger')
                logger.warning(f"File too large: {file_size} bytes")
                return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)

            logger.debug("Reading Excel file")
            df = pd.read_excel(file, engine='openpyxl', nrows=1000)
            logger.debug(f"Excel rows read: {len(df)}")

            required_columns = ['Admission No', 'Name']
            if not all(col in df.columns for col in required_columns):
                flash('Excel file must contain columns: Admission No, Name', 'danger')
                logger.error("Missing required columns")
                return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)

            logger.debug("Cleaning data")
            df = df.dropna(subset=['Admission No', 'Name'])
            df['Admission No'] = df['Admission No'].astype(str).str.strip()
            df['Name'] = df['Name'].astype(str).str.strip()
            df = df.sort_values(by='Admission No')  # Sort by Admission No
            if df.empty:
                flash('No valid data found in the Excel file.', 'warning')
                logger.warning("Empty dataframe after cleaning")
                return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)

            student_count = db_session.query(User).filter_by(role='student').count()
            max_students = 500
            if student_count >= max_students:
                flash(f'Maximum number of learners ({max_students}) reached.', 'danger')
                logger.warning(f"Student limit reached: {student_count}")
                return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)

            successes = []
            errors = []
            batch_size = 5
            logger.debug(f"Processing {len(df)} rows in batches of {batch_size}")

            for start_idx in range(0, len(df), batch_size):
                batch_df = df[start_idx:start_idx + batch_size]
                logger.debug(f"Processing batch {start_idx//batch_size + 1}")
                with db_session.no_autoflush:
                    for index, row in batch_df.iterrows():
                        admission_no = row['Admission No']
                        name = row['Name']
                        logger.debug(f"Row {index+2}: admission_no={admission_no}, name={name}")
                        if not admission_no or not name:
                            errors.append(f"Row {index+2}: Admission No or Name is empty")
                            continue

                        # Check if student already exists
                        existing_student = db_session.query(Student).filter_by(admission_no=admission_no).first()
                        if existing_student:
                            logger.debug(f"Skipping existing student with admission_no={admission_no}")
                            continue

                        username = name.lower().replace(' ', '_')
                        counter = 1
                        orig_username = username
                        while True:
                            if not db_session.query(User).filter_by(username=username).first():
                                break
                            username = f"{orig_username}_{counter}"
                            counter += 1
                            if counter > 5:
                                errors.append(f"Row {index+2}: Cannot generate unique username for {name}")
                                logger.error(f"Username generation failed for {name}, admission_no: {admission_no}")
                                break
                        if counter > 5:
                            continue

                        password = generate_student_password(name, admission_no)
                        password_hash = generate_password_hash(password)
                        logger.debug(f"Generated username={username}, password={password}")

                        try:
                            new_student = Student(admission_no=admission_no, name=name, grade=grade)
                            new_user = User(
                                username=username,
                                password_hash=password_hash,
                                role='student',
                                admission_no=admission_no,
                                grade=grade
                            )
                            db_session.add_all([new_student, new_user])
                        except Exception as e:
                            errors.append(f"Row {index+2}: Database error: {str(e)}")
                            logger.error(f"Database error for row {index+2}: {str(e)}\n{traceback.format_exc()}")
                            continue

                    try:
                        db_session.commit()
                        logger.debug(f"Committed Student and User for batch {start_idx//batch_size + 1}")
                    except IntegrityError as e:
                        db_session.rollback()
                        error_msg = str(e).lower()
                        for index, row in batch_df.iterrows():
                            if not (row['Admission No'] and row['Name']):
                                continue
                            admission_no = row['Admission No']
                            name = row['Name']
                            if 'admission_no' in error_msg:
                                errors.append(f"Row {index+2}: Admission No {admission_no} already exists")
                            elif 'username' in error_msg:
                                errors.append(f"Row {index+2}: Username conflict for {name}")
                            else:
                                errors.append(f"Row {index+2}: Database error: {str(e)}")
                        logger.error(f"IntegrityError for batch {start_idx//batch_size + 1}: {str(e)}")
                        continue
                    except Exception as e:
                        db_session.rollback()
                        errors.append(f"Batch {start_idx//batch_size + 1}: Database error: {str(e)}")
                        logger.error(f"Batch error: {str(e)}\n{traceback.format_exc()}")
                        continue

                    # Add Fee records after committing Student and User
                    for index, row in batch_df.iterrows():
                        if not (row['Admission No'] and row['Name']):
                            continue
                        admission_no = row['Admission No']
                        # Skip if student was not added due to existing record
                        if db_session.query(Student).filter_by(admission_no=admission_no).first():
                            try:
                                new_fee = Fee(
                                    admission_no=admission_no,
                                    total_fee=0,
                                    amount_paid=0,
                                    balance=0,
                                    grade=grade,
                                    term=term_info['term'],
                                    year=term_info['year']
                                )
                                db_session.add(new_fee)
                                successes.append(f"{row['Name']} (Admission No: {row['Admission No']})")
                            except Exception as e:
                                errors.append(f"Row {index+2}: Database error for Fee: {str(e)}")
                                logger.error(f"Database error for Fee in row {index+2}: {str(e)}\n{traceback.format_exc()}")
                                continue

                    try:
                        db_session.commit()
                        logger.debug(f"Committed Fee for batch {start_idx//batch_size + 1}")
                    except IntegrityError as e:
                        db_session.rollback()
                        error_msg = str(e).lower()
                        for index, row in batch_df.iterrows():
                            if not (row['Admission No'] and row['Name']):
                                continue
                            admission_no = row['Admission No']
                            name = row['Name']
                            errors.append(f"Row {index+2}: Fee database error: {str(e)}")
                        logger.error(f"IntegrityError for Fee in batch {start_idx//batch_size + 1}: {str(e)}")
                        continue
                    except Exception as e:
                        db_session.rollback()
                        errors.append(f"Batch {start_idx//batch_size + 1}: Fee database error: {str(e)}")
                        logger.error(f"Fee batch error: {str(e)}\n{traceback.format_exc()}")
                        continue

            if successes:
                flash(f'Successfully uploaded {len(successes)} students.', 'success')
                logger.info(f"Uploaded {len(successes)} students")
            if errors:
                flash(f'Failed to upload {len(errors)} students: {"; ".join(errors)}', 'danger')
                logger.error(f"Upload errors: {errors}")
            if not successes and not errors:
                flash('No valid data found in the Excel file.', 'warning')
                logger.warning("No data processed")

            return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)
        except pd.errors.ParserError:
            flash('Invalid Excel file format. Please upload a valid .xlsx file.', 'danger')
            logger.error("Excel parsing error")
            return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)
        except Exception as e:
            logger.error(f"Unexpected error in upload_bulk_students: {str(e)}\n{traceback.format_exc()}")
            flash(f'Error processing file: {str(e)}', 'danger')
            return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)
        finally:
            db_session.close()

    return render_template('upload_bulk_students.html', form=form, term_info=term_info, content_data=content_data)

@app.route('/admin/promote_learners', methods=['POST'])
@login_required
@admin_required
def promote_learners():
    """Promote students to the next grade."""
    db_session = next(get_db())
    try:
        db_session.query(User).filter(User.grade == 'Grade 7', User.role == 'student').update({User.grade: 'Grade 8'})
        db_session.query(User).filter(User.grade == 'Grade 8', User.role == 'student').update({User.grade: 'Grade 9'})
        db_session.query(Fee).filter(Fee.grade == 'Grade 7').update({Fee.grade: 'Grade 8'})
        db_session.query(Fee).filter(Fee.grade == 'Grade 8').update({Fee.grade: 'Grade 9'})
        db_session.commit()
        flash('Learners promoted successfully!', 'success')
        logger.info("Learners promoted successfully")
    except Exception as e:
        db_session.rollback()
        flash(f'Error promoting learners: {str(e)}', 'danger')
        logger.error(f"Error promoting learners: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('dashboard'))



@app.route('/search_students', methods=['GET'])
def search_students():
    """Search students by name for autocomplete."""
    query = request.args.get('q', '').strip()
    if not query:
        logger.debug("Empty query in search_students")
        return jsonify([])

    db_session = next(get_db())
    try:
        students = db_session.query(User).filter(
            User.username.ilike(f'%{query}%'),
            User.role == 'student'
        ).limit(10).all()
        logger.debug(f"Search students: query={query}, found={len(students)} students")
        return jsonify([{'username': student.username, 'admission_no': student.admission_no} for student in students])
    except Exception as e:
        logger.error(f"Error searching students: {str(e)}\n{traceback.format_exc()}")
        return jsonify([])
    finally:
        db_session.close()






@app.route('/manage_fees', methods=['GET', 'POST'])
@login_required
@admin_or_bursar_required
def manage_fees():
    """Manage student fees for admin or bursar."""
    filter_form = FeeFilterForm()
    filter_form.grade.choices = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')]
    update_form = FeeForm(role=current_user.role)
    students = []
    fees_data = []
    payment_history = []
    term_info, content_data = fetch_common_data()

    db_session = next(get_db())
    try:
        if request.method == 'POST' and filter_form.validate_on_submit():
            grade = filter_form.grade.data
            term = filter_form.term.data
            year = filter_form.year.data
            logger.debug(f"Filter form submitted: grade={grade}, term={term}, year={year}")

            # Fetch students for the selected grade
            students_query = db_session.query(User.admission_no, User.username).filter(
                User.role == 'student',
                User.grade.ilike(grade)  # Case-insensitive matching
            )
            students = students_query.order_by(User.admission_no).all()
            logger.debug(f"Fetched {len(students)} students for grade={grade}")

            if not students:
                flash(f'No students found for grade {grade}.', 'warning')
                logger.warning(f"No students found for grade={grade}")
            else:
                # Fetch fee records
                for student in students:
                    admission_no = student[0]
                    fee = db_session.query(Fee).filter_by(
                        admission_no=admission_no,
                        grade=grade,
                        term=term,
                        year=str(year)
                    ).first()
                    total_fee = float(fee.total_fee or 0.0) if fee else 0.0
                    amount_paid = float(fee.amount_paid or 0.0) if fee else 0.0
                    balance = total_fee - amount_paid  # Simplified balance calculation
                    fees_data.append({
                        'admission_no': admission_no,
                        'name': student[1],
                        'total_fee': total_fee,
                        'amount_paid': amount_paid,
                        'balance': balance,
                        'grade': grade,
                        'term': term,
                        'year': year
                    })

                # Fetch payment history
                payment_history = db_session.query(PaymentHistory).filter(
                    PaymentHistory.term == term,
                    PaymentHistory.year == str(year),
                    PaymentHistory.admission_no.in_([s[0] for s in students])
                ).order_by(PaymentHistory.date.desc()).all()
                logger.debug(f"Fetched {len(payment_history)} payment history records")
                for payment in payment_history:
                    logger.debug(f"PaymentHistory: admission_no={payment.admission_no}, date={payment.date}, type={type(payment.date)}")

                if not fees_data:
                    flash(f'No fee records found for grade={grade}, term={term}, year={year}.', 'warning')
                    logger.warning(f"No fee records found for grade={grade}, term={term}, year={year}")
                else:
                    flash(f'Successfully fetched {len(fees_data)} fee records.', 'success')

        elif filter_form.errors:
            logger.error(f"Filter form errors: {filter_form.errors}")
            flash(f'Filter form errors: {filter_form.errors}', 'danger')

        if update_form.validate_on_submit():
            admission_no = update_form.admission_no.data
            total_fee = float(update_form.total_fee.data or 0.0)
            amount_paid = float(update_form.amount_paid.data or 0.0)
            grade = update_form.grade.data
            term = update_form.term.data
            year = str(update_form.year.data)
            logger.debug(f"Update form submitted: admission_no={admission_no}, total_fee={total_fee}, amount_paid={amount_paid}, grade={grade}, term={term}, year={year}")

            # Check for existing fee record
            fee_record = db_session.query(Fee).filter_by(
                admission_no=admission_no,
                grade=grade,
                term=term,
                year=year
            ).first()

            if not fee_record:
                logger.debug(f"Inserting new fee record for {admission_no}")
                balance = total_fee - amount_paid
                fee_record = Fee(
                    admission_no=admission_no,
                    total_fee=total_fee,
                    amount_paid=amount_paid,
                    balance=balance,
                    grade=grade,
                    term=term,
                    year=year
                )
                db_session.add(fee_record)
            else:
                logger.debug(f"Updating existing fee record for {admission_no}")
                fee_record.total_fee = total_fee
                fee_record.amount_paid = float(fee_record.amount_paid or 0.0) + amount_paid
                fee_record.balance = total_fee - fee_record.amount_paid

            # Record payment history
            if amount_paid > 0:
                db_session.add(PaymentHistory(
                    admission_no=admission_no,
                    amount=amount_paid,
                    date=datetime.now(),
                    term=term,
                    year=year
                ))

            db_session.commit()
            flash(f'Fee updated successfully for student {admission_no}.', 'success')
            return redirect(url_for('manage_fees'))

        elif update_form.errors:
            logger.error(f"Update form errors: {update_form.errors}")
            flash(f'Form errors: {update_form.errors}', 'danger')

        return render_template(
            'manage_fees.html',
            filter_form=filter_form,
            update_form=update_form,
            students=students,
            fees_data=fees_data,
            payment_history=payment_history,
            term_info=term_info,
            content_data=content_data
        )
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in manage_fees: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error loading fee management: {str(e)}', 'danger')
        return render_template(
            'manage_fees.html',
            filter_form=filter_form,
            update_form=update_form,
            students=[],
            fees_data=[],
            payment_history=[],
            term_info=term_info,
            content_data=content_data
        )
    finally:
        db_session.close()
        
        
@app.route('/bursar/dashboard', methods=['GET'])
@login_required
@bursar_required
def bursar_dashboard():
    """Render the bursar dashboard."""
    logger.debug("Entering bursar_dashboard route")
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    try:
        logger.debug("Rendering bursar_dashboard.html")
        return render_template('bursar_dashboard.html', term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in bursar_dashboard: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error fetching term info: {str(e)}', 'danger')
        return render_template('bursar_dashboard.html', term_info=term_info, content_data=content_data)
    
    
    
    
    
@app.route('/bursar/manage_fees', methods=['GET', 'POST'])
@login_required
@bursar_required
def bursar_manage_fees():
    """Manage student fees for bursar."""
    filter_form = FeeFilterForm()
    update_form = FeeForm()
    students = []
    fees_data = []
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

    if filter_form.validate_on_submit():
        db_session = next(get_db())
        try:
            grade = filter_form.grade.data
            term = filter_form.term.data
            year = filter_form.year.data
            logger.debug(f"Filter form submitted: grade={grade}, term={term}, year={year}")

            students_query = db_session.query(User.admission_no, User.username).filter(
                User.role == 'student',
                User.grade == grade
            ).order_by(User.admission_no)
            students = students_query.all()
            logger.debug(f"Fetched {len(students)} students for grade={grade}")

            if not students:
                flash(f'No students found for grade {grade}.', 'warning')
                logger.warning(f"No students found for grade={grade}")

            for student in students:
                admission_no = student[0]
                fee = db_session.query(Fee).filter_by(
                    admission_no=admission_no,
                    grade=grade,
                    term=term,
                    year=str(year)
                ).first()
                logger.debug(f"Processing fee for admission_no={admission_no}, fee={fee.__dict__ if fee else None}")

                total_fee = 0.0
                amount_paid = 0.0
                balance = 0.0
                if fee:
                    total_fee = float(fee.total_fee or 0.0)
                    amount_paid = float(fee.amount_paid or 0.0)
                    balance = float(fee.balance or 0.0)
                    if fee.total_fee is None:
                        logger.warning(f"total_fee is None for admission_no={admission_no}")
                    if fee.amount_paid is None:
                        logger.warning(f"amount_paid is None for admission_no={admission_no}")
                    if fee.balance is None:
                        logger.warning(f"balance is None for admission_no={admission_no}")

                fees_data.append({
                    'admission_no': admission_no,
                    'name': student[1],
                    'total_fee': total_fee,
                    'amount_paid': amount_paid,
                    'balance': balance
                })
            logger.debug(f"Fetched {len(fees_data)} fee records")
            if not fees_data:
                flash(f'No fee records found for grade={grade}, term={term}, year={year}.', 'warning')
                logger.warning(f"No fee records found for grade={grade}, term={term}, year={year}")
            else:
                flash(f'Successfully fetched {len(fees_data)} fee records.', 'success')
        except Exception as e:
            db_session.rollback()
            logger.error(f"Error fetching data: {str(e)}\n{traceback.format_exc()}")
            flash(f'Error fetching data: {str(e)}', 'danger')
        finally:
            db_session.close()

    elif filter_form.errors:
        logger.error(f"Filter form errors: {filter_form.errors}")
        flash(f'Filter form errors: {filter_form.errors}', 'danger')

    if update_form.validate_on_submit():
        db_session = next(get_db())
        try:
            learner_name = update_form.learner_name.data
            total_fee = float(update_form.total_fee.data or 0.0)
            amount_paid = float(update_form.amount_paid.data or 0.0)
            grade = update_form.grade.data
            term = update_form.term.data
            year = update_form.year.data
            logger.debug(f"Update form submitted: learner_name={learner_name}, total_fee={total_fee}, amount_paid={amount_paid}, grade={grade}, term={term}, year={year}")

            student = db_session.query(User).filter(
                User.username.ilike(f'%{learner_name}%'),
                User.role == 'student'
            ).first()
            if not student:
                flash(f'No student found with name: {learner_name}', 'danger')
                logger.error(f"No student found for learner_name: {learner_name}")
                return render_template('manage_fees.html', filter_form=filter_form, update_form=update_form, students=students, fees_data=fees_data, term_info=term_info, content_data=content_data)

            admission_no = student.admission_no
            logger.debug(f"Resolved learner_name={learner_name} to admission_no={admission_no}")

            fee_record = db_session.query(Fee).filter_by(admission_no=admission_no, grade=grade, term=term, year=str(year)).first()
            if not fee_record:
                logger.debug(f"Inserting new fee record for {admission_no}")
                balance = total_fee - amount_paid
                fee_record = Fee(
                    admission_no=admission_no,
                    total_fee=total_fee,
                    amount_paid=amount_paid,
                    balance=balance,
                    grade=grade,
                    term=term,
                    year=str(year)
                )
                db_session.add(fee_record)
            else:
                logger.debug(f"Updating existing fee record for {admission_no}")
                new_paid = float(fee_record.amount_paid or 0.0) + amount_paid
                fee_record.total_fee = total_fee
                fee_record.amount_paid = new_paid
                fee_record.balance = total_fee - new_paid

            db_session.add(PaymentHistory(
                admission_no=admission_no,
                amount=amount_paid,
                date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                term=term,
                year=str(year)
            ))
            db_session.commit()
            flash('Fee updated successfully!', 'success')
            logger.debug(f"Updated fee for admission_no: {admission_no}, grade: {grade}, term: {term}, year: {year}")
            return redirect(url_for('bursar_manage_fees'))
        except Exception as e:
            db_session.rollback()
            flash(f'Error updating fee: {str(e)}', 'danger')
            logger.error(f"Error updating fee: {str(e)}\n{traceback.format_exc()}")
        finally:
            db_session.close()

    elif update_form.errors:
        logger.error(f"Update form errors: {update_form.errors}")
        flash(f'Form errors: {update_form.errors}', 'danger')

    return render_template('manage_fees.html', filter_form=filter_form, update_form=update_form, students=students, fees_data=fees_data, term_info=term_info, content_data=content_data)




@app.route('/download_fee_statement_excel', methods=['GET', 'POST'])
@login_required
@admin_or_bursar_required
def download_fee_statement_excel():
    """Download fee statements as Excel for admin or bursar."""
    form = FeeFilterForm()
    form.grade.choices = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9'), ('all', 'All')]
    form.term.choices = [('Term 1', 'Term 1'), ('Term 2', 'Term 2'), ('Term 3', 'Term 3')]
    form.year.data = str(datetime.now().year)
    term_info, content_data = fetch_common_data()

    db_session = next(get_db())
    try:
        if request.method == 'POST' and form.validate_on_submit():
            grade = form.grade.data
            term = form.term.data
            year = form.year.data
            logger.debug(f"Download Excel fee statement: grade={grade}, term={term}, year={year}")

            # Build query
            query = db_session.query(
                Fee.admission_no,
                User.username.label('student_name'),
                Fee.total_fee,
                Fee.amount_paid,
                Fee.balance,
                Fee.grade,
                Fee.term,
                Fee.year
            ).join(User, Fee.admission_no == User.admission_no).filter(
                User.role == 'student'
            )
            if grade != 'all':
                query = query.filter(Fee.grade.ilike(grade, escape='/'))
            if term:
                query = query.filter(Fee.term == term)
            if year:
                query = query.filter(Fee.year == year)

            fees = [(f.admission_no, f.student_name, f.total_fee, f.amount_paid, f.balance, f.grade, f.term, f.year) for f in query.all()]
            if not fees:
                logger.warning(f"No fee data available for grade={grade}, term={term}, year={year}")
                flash(f'No fee data available for grade {grade}, term {term}, year {year}.', 'warning')
                return render_template('select_fee_statement.html', form=form, role=current_user.role, term_info=term_info, content_data=content_data)

            # Generate Excel file
            buffer = generate_fee_statement_excel(fees, grade, term, year)
            if not buffer or len(buffer.getvalue()) == 0:
                logger.error(f"Empty or invalid Excel buffer for grade={grade}, term={term}, year={year}")
                flash('Error generating fee statement Excel.', 'danger')
                return render_template('select_fee_statement.html', form=form, role=current_user.role, term_info=term_info, content_data=content_data)

            buffer.seek(0)
            filename = f'fee_statement_{grade}_{term}_{year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'.replace(' ', '_')
            return send_file(
                buffer,
                download_name=filename,
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        elif form.errors:
            logger.error(f"Form errors: {form.errors}")
            flash(f'Form errors: {form.errors}', 'danger')

        return render_template('select_fee_statement.html', form=form, role=current_user.role, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in download_fee_statement_excel: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error generating Excel fee statement: {str(e)}', 'danger')
        return render_template('select_fee_statement.html', form=form, role=current_user.role, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()

def generate_fee_statement_excel(fees, grade, term, year):
    """Generate Excel file for fee statements."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Fee Statement {grade} {term} {year}"
    ws.append(['Admission No', 'Student Name', 'Total Fee', 'Amount Paid', 'Balance', 'Grade', 'Term', 'Year'])
    for fee in fees:
        ws.append(fee)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer

def generate_fee_statement_excel(fees, grade, term=None, year=None):
    """Generate an Excel file for fee statements."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Fee Statement - Grade {grade if grade != 'all' else 'All'}"
        
        headers = ['Admission No', 'Student Name', 'Total Fee', 'Amount Paid', 'Balance', 'Grade', 'Term', 'Year']
        sheet.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        for fee in fees:
            sheet.append(fee)
        
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
        
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}\n{traceback.format_exc()}")
        return None
    
    

@app.route('/get_payment_history/<admission_no>')
@login_required
def get_payment_history(admission_no):
    """Retrieve payment history for a student."""
    if current_user.role not in ['admin', 'bursar']:
        return jsonify({'error': 'Access denied'}), 403
    db_session = next(get_db())
    try:
        history = db_session.query(PaymentHistory.amount, PaymentHistory.date).filter_by(admission_no=admission_no).order_by(PaymentHistory.date.desc()).all()
        return jsonify([{'amount': h.amount, 'date': h.date} for h in history])
    except SQLAlchemyError as e:
        logger.error(f"Database error in get_payment_history: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    finally:
        db_session.close()



@app.route('/student_download_fee_statement', methods=['GET', 'POST'])
@login_required
def student_download_fee_statement(admission_no=None, grade=None, term=None, year=None):
    """Download fee statement for a student as PDF, accessible by both parents and students."""
    form = FeeStatementForm()
    term_info, content_data = fetch_common_data()
    db_session = next(get_db())
    linked_students = []
    
    try:
        # Initialize grade choices for all cases
        form.grade.choices = [(g, g) for g in GRADES]  # GRADES is assumed to be a list of valid grades

        if current_user.role == 'parent':
            # Fetch linked students for parent
            linked_students = db_session.query(Student.admission_no, Student.name, Student.grade).\
                join(ParentStudent, Student.admission_no == ParentStudent.admission_no).\
                filter(ParentStudent.parent_id == current_user.id).all()
            
            form.admission_no.choices = [
                (s.admission_no, f"{s.name} ({s.admission_no})")
                for s in linked_students
            ] if linked_students else []
            
            if not linked_students:
                logger.warning(f"No linked students found for parent_id: {current_user.id}")
                flash('No linked students found. Please link a student.', 'danger')
                return redirect(url_for('parent_dashboard'))
            
            # Get admission_no from form, args, or None
            admission_no = admission_no or (form.admission_no.data if form.validate_on_submit() else request.args.get('admission_no'))
            
            if not admission_no:
                logger.warning(f"No admission_no provided for parent_id: {current_user.id}")
                flash('Please select a student to download their fee statement.', 'danger')
                return redirect(url_for('parent_dashboard'))
            
            # Verify parent-student linkage
            parent_student = db_session.query(ParentStudent).filter_by(
                parent_id=current_user.id,
                admission_no=admission_no
            ).first()
            
            if not parent_student:
                logger.warning(f"Parent {current_user.id} not linked to admission_no: {admission_no}")
                flash('You are not authorized to view this student’s fee statement.', 'danger')
                return redirect(url_for('parent_dashboard'))
            
            student_data = db_session.query(Student.admission_no, Student.name, Student.grade).filter_by(admission_no=admission_no).first()
            if not student_data:
                logger.error(f"Student not found: {admission_no}")
                flash('Student profile not found.', 'danger')
                return redirect(url_for('parent_dashboard'))
            
            # Set default grade from student data or form
            form.grade.data = form.grade.data or student_data.grade

        elif current_user.role == 'student':
            # Handle student role
            admission_no = getattr(current_user, 'admission_no', None)
            if not admission_no:
                logger.error(f"No admission number for user {current_user.id}")
                flash('No admission number associated with your account. Please contact the administrator.', 'danger')
                return redirect(url_for('student_dashboard'))
            
            student_data = db_session.query(Student.admission_no, Student.name, Student.grade).filter_by(admission_no=admission_no).first()
            if not student_data:
                logger.error(f"Student not found: {admission_no}")
                flash('Student profile not found.', 'danger')
                return redirect(url_for('student_dashboard'))
            
            form.admission_no.choices = [(student_data.admission_no, f"{student_data.name} ({student_data.admission_no})")]
            form.admission_no.data = student_data.admission_no
            form.grade.data = form.grade.data or student_data.grade

        else:
            logger.error(f"Unauthorized role {current_user.role} for user {current_user.id}")
            flash('You are not authorized to access this page.', 'danger')
            return redirect(url_for('index'))

        # Set term and year choices
        form.term.choices = [(t, t) for t in TERMS]  # TERMS is assumed to be a list of valid terms
        form.year.data = str(datetime.now().year)

        if request.method == 'POST' and form.validate_on_submit():
            grade = form.grade.data
            term = form.term.data
            year = form.year.data
            
            # Query fee data with flexible filters
            query = db_session.query(
                Fee.admission_no,
                Student.name.label('student_name'),
                Fee.total_fee,
                Fee.amount_paid,
                Fee.balance,
                Fee.grade,
                Fee.term,
                Fee.year
            ).join(Student, Fee.admission_no == Student.admission_no).filter(Fee.admission_no == admission_no)
            
            if term:
                query = query.filter(Fee.term == term)
            if year:
                query = query.filter(Fee.year == year)
            if grade:
                query = query.filter(Fee.grade.ilike(grade, escape='/'))
            
            fees = query.all()
            
            if not fees:
                logger.warning(f"No fee statement available for admission_no={admission_no}, grade={grade}, term={term}, year={year}")
                flash(f'No fee statement available for grade {grade}, term {term}, year {year}.', 'warning')
                return render_template(
                    'student_dashboard.html',
                    student={'admission_no': student_data.admission_no, 'name': student_data.name, 'grade': student_data.grade},
                    recent_announcements=[],
                    report_form=ReportCardForm(admission_no=student_data.admission_no, grade=student_data.grade),
                    fee_form=form,
                    view_results_form=ResultsFilterForm(admission_no=student_data.admission_no, grade=student_data.grade),
                    term_info=term_info,
                    content_data=content_data,
                    parent_view=current_user.role == 'parent',
                    linked_students=linked_students
                )

            # Generate PDF
            buffer = generate_fee_statement_pdf([(f.admission_no, f.student_name, f.total_fee, f.amount_paid, f.balance, f.grade, f.term, f.year) for f in fees], grade, term, year)
            if not buffer or len(buffer.getvalue()) == 0:
                logger.error(f"Empty or invalid fee statement PDF buffer for admission_no={admission_no}")
                flash('Error generating fee statement PDF.', 'danger')
                return render_template(
                    'student_dashboard.html',
                    student={'admission_no': student_data.admission_no, 'name': student_data.name, 'grade': student_data.grade},
                    recent_announcements=[],
                    report_form=ReportCardForm(admission_no=student_data.admission_no, grade=student_data.grade),
                    fee_form=form,
                    view_results_form=ResultsFilterForm(admission_no=student_data.admission_no, grade=student_data.grade),
                    term_info=term_info,
                    content_data=content_data,
                    parent_view=current_user.role == 'parent',
                    linked_students=linked_students
                )

            buffer.seek(0)
            filename = f'fee_statement_{admission_no}_{grade}_{term}_{year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'.replace(' ', '_')
            logger.info(f"Fee statement generated for {admission_no}, size: {len(buffer.getvalue())} bytes")
            return send_file(
                buffer,
                download_name=filename,
                as_attachment=True,
                mimetype='application/pdf'
            )

        # Render template for GET request or form validation failure
        return render_template(
            'student_dashboard.html',
            student={'admission_no': student_data.admission_no, 'name': student_data.name, 'grade': student_data.grade},
            recent_announcements=[],
            report_form=ReportCardForm(admission_no=student_data.admission_no, grade=student_data.grade),
            fee_form=form,
            view_results_form=ResultsFilterForm(admission_no=student_data.admission_no, grade=student_data.grade),
            term_info=term_info,
            content_data=content_data,
            parent_view=current_user.role == 'parent',
            linked_students=linked_students
        )
    
    except SQLAlchemyError as e:
        logger.error(f"Database error for admission_no={admission_no or 'unknown'}: {str(e)}\n{traceback.format_exc()}")
        flash('Database error: Unable to fetch fee statement data. Please try again later.', 'danger')
        return render_template(
            'student_dashboard.html',
            student={'admission_no': '', 'name': '', 'grade': ''},
            recent_announcements=[],
            report_form=ReportCardForm(),
            fee_form=form,
            view_results_form=ResultsFilterForm(),
            term_info=term_info,
            content_data=content_data,
            parent_view=current_user.role == 'parent',
            linked_students=linked_students
        )
    
    except Exception as e:
        logger.error(f"Unexpected error in student_download_fee_statement: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template(
            'student_dashboard.html',
            student={'admission_no': '', 'name': '', 'grade': ''},
            recent_announcements=[],
            report_form=ReportCardForm(),
            fee_form=form,
            view_results_form=ResultsFilterForm(),
            term_info=term_info,
            content_data=content_data,
            parent_view=current_user.role == 'parent',
            linked_students=linked_students
        )
    
    finally:
        db_session.close()
        

        
        
@app.route('/admin/notes', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_notes():
    """Manage admin notes."""
    form = NoteForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        if form.validate_on_submit():
            content = form.content.data
            note_id = request.form.get('note_id')
            if note_id:
                note = db_session.query(AdminNotes).filter_by(id=note_id).first()
                if note:
                    note.content = content
                    note.created_by = current_user.username
                    note.created_at = datetime.now()
                    flash('Note updated successfully.', 'success')
                    logger.info(f"Note updated: id={note_id}, user={current_user.username}")
                else:
                    flash('Note not found.', 'danger')
                    logger.warning(f"Note not found: id={note_id}")
                    return render_template('admin_notes.html', form=form, notes=[], term_info=term_info, content_data=content_data)
            else:
                new_note = AdminNotes(content=content, created_by=current_user.username, created_at=datetime.now())
                db_session.add(new_note)
                flash('Note added successfully.', 'success')
                logger.info(f"Note added by user={current_user.username}")
            db_session.commit()
            return redirect(url_for('admin_notes'))

        notes = db_session.query(AdminNotes.id, AdminNotes.content, AdminNotes.created_by, AdminNotes.created_at).order_by(AdminNotes.created_at.desc()).all()
        return render_template('admin_notes.html', form=form, notes=notes, term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in admin_notes: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error processing notes: {str(e)}', 'danger')
        return render_template('admin_notes.html', form=form, notes=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/admin/notes/delete/<int:note_id>', methods=['POST'])
@login_required
@admin_required
def delete_note(note_id):
    """Delete an admin note."""
    try:
        note = db.session.query(AdminNotes).filter_by(id=note_id).first()
        if note:
            db.session.delete(note)
            db.session.commit()
            flash('Note deleted successfully.', 'success')
        else:
            flash('Note not found.', 'danger')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Database error deleting note: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error deleting note: {str(e)}', 'danger')
    return redirect(url_for('admin_notes'))

@app.route('/admin/enter_marks', methods=['GET', 'POST'])
@login_required
def enter_marks():
    """Allow admin and teacher to enter and submit student marks."""
    if current_user.role not in ['admin', 'teacher']:
        flash('Access denied!', 'danger')
        logger.warning(f'Access denied for user {current_user.id} with role {current_user.role}')
        return redirect(url_for('dashboard'))

    form = MarksForm()
    learners = []
    show_table = False
    term_info = {}
    content_data = {}

    try:
        with db.session() as db_session:
            # Fetch term info
            term_data = db_session.query(TermInfo).first()
            term_info = {
                'term': term_data.term if term_data and term_data.term else 'Term 1',
                'year': term_data.year if term_data and term_data.year else '2025',
                'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
                'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
                'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
            }
            logger.debug(f"Term info: {term_info}")

            # Fetch mission, vision, about
            mission = db_session.query(Mission).first()
            vision = db_session.query(Vision).first()
            about = db_session.query(About).first()
            content_data = {
                'mission': mission.content if mission else "To provide quality education for all students",
                'vision': vision.content if vision else "To be a leading institution in academic excellence",
                'about': about.content if about else "Jonyo Junior School is dedicated to fostering holistic education"
            }
            logger.debug(f"Content data: {content_data}")

            # Populate form choices
            if current_user.role == 'teacher':
                assignments = db_session.query(TeacherAssignments.learning_area, TeacherAssignments.grade).filter_by(teacher_id=current_user.id).all()
                form.learning_area.choices = [(f"{a.learning_area}|{a.grade}", f"{a.learning_area} ({a.grade})") for a in assignments] or [('', 'No subjects assigned')]
                if not assignments:
                    flash('You are not assigned to any subjects.', 'warning')
                    logger.warning(f"No assignments for teacher {current_user.id}")
            else:
                learning_areas = db_session.query(LearningAreas.name, LearningAreas.grade).all()
                form.learning_area.choices = [(f"{la.name}|{la.grade}", f"{la.name} ({la.grade})") for la in learning_areas]
                logger.debug(f"Admin learning areas: {[(la.name, la.grade) for la in learning_areas]}")

            # Populate other form choices
            grades = db_session.query(Student.grade).distinct().all()
            form.grade.choices = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')] or [('Grade 7', 'Grade 7')]
            form.term.choices = [('Term 1', 'Term 1'), ('Term 2', 'Term 2'), ('Term 3', 'Term 3')]
            form.exam_type.choices = [
                ('cat1', 'CAT 1'), ('cat2', 'CAT 2'), ('cat3', 'CAT 3'),
                ('rat1', 'RAT 1'), ('rat2', 'RAT 2'), ('rat3', 'RAT 3'),
                ('midterm', 'Mid Term'), ('endterm', 'End Term'),
                ('project1', 'Project 1'), ('project2', 'Project 2'), ('project3', 'Project 3')
            ]
            form.year.choices = [(str(y), str(y)) for y in range(2020, 2030)]
            logger.debug(f"Form choices - grades: {form.grade.choices}, learning_areas: {form.learning_area.choices}")

            if request.method == 'POST':
                logger.debug(f"POST data: {request.form}")

                # Extract form data
                learning_area_grade = form.learning_area.data or request.form.get('learning_area')
                grade = form.grade.data or request.form.get('grade')
                exam_type = form.exam_type.data or request.form.get('exam_type')
                term = form.term.data or request.form.get('term')
                year = form.year.data or request.form.get('year')

                # If learning_area includes grade, extract it
                if learning_area_grade and '|' in learning_area_grade:
                    learning_area, grade_from_learning_area = learning_area_grade.split('|')
                else:
                    learning_area = learning_area_grade

                # Prefer grade from form, fallback to grade from learning_area
                grade = grade or grade_from_learning_area

                logger.debug(f"Form values: grade={grade}, learning_area={learning_area}, exam_type={exam_type}, term={term}, year={year}")

                if request.form.get('submit_marks') != 'true':  # Handle "Show Students" (fetch_students)
                    # Validate inputs
                    if not all([grade, learning_area, exam_type, term, year]):
                        flash('All fields (grade, learning area, exam type, term, year) are required.', 'danger')
                        logger.warning(f"Missing form fields: grade={grade}, learning_area={learning_area}, exam_type={exam_type}, term={term}, year={year}")
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    try:
                        year = int(year)
                        if not (2000 <= year <= 2100):
                            raise ValueError("Year out of range")
                    except ValueError:
                        flash('Invalid year input. Must be between 2000 and 2100.', 'danger')
                        logger.error(f"Invalid year: {year}")
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    # Validate learning area
                    if not db_session.query(LearningAreas).filter_by(name=learning_area, grade=grade).first():
                        flash(f'Invalid learning area "{learning_area}" for grade "{grade}".', 'danger')
                        logger.error(f"Invalid learning area: {learning_area} for grade {grade}")
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    # Fetch students in ascending order of admission number
                    students = db_session.query(Student.admission_no, Student.name)\
    .filter_by(grade=grade)\
    .order_by(cast(Student.admission_no, Integer).asc())\
    .all()
                    logger.debug(f"Fetched {len(students)} students for grade {grade}: {[(s.admission_no, s.name) for s in students[:5]]}")

                    if not students:
                        flash(f'No students found for grade {grade}. Please add students to the database.', 'warning')
                        logger.warning(f"No students found for grade {grade}")
                        show_table = True  # Show table with "No students found" message
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    # Prepare learners list with existing marks
                    for student in students:
                        mark = db_session.query(Marks).filter_by(
                            admission_no=student.admission_no,
                            learning_area=learning_area,
                            exam_type=exam_type,
                            term=term,
                            year=year,
                            grade=grade
                        ).first()
                        learners.append({
                            'admission_no': student.admission_no,
                            'name': student.name,
                            'marks': str(mark.marks) if mark and mark.marks is not None else ''
                        })
                    show_table = True
                    logger.debug(f"Prepared {len(learners)} learners: {learners[:5]}")

                elif request.form.get('submit_marks') == 'true':  # Handle "Save Marks"
                    if not all([grade, learning_area, exam_type, term, year]):
                        flash('All fields (grade, learning area, exam type, term, year) are required.', 'danger')
                        logger.warning(f"Missing form fields on submit: grade={grade}, learning_area={learning_area}, exam_type={exam_type}, term={term}, year={year}")
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    try:
                        year = int(year)
                        if not (2000 <= year <= 2100):
                            raise ValueError("Year out of range")
                    except ValueError:
                        flash('Invalid year input. Must be between 2000 and 2100.', 'danger')
                        logger.error(f"Invalid year: {year}")
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    # Validate learning area
                    if not db_session.query(LearningAreas).filter_by(name=learning_area, grade=grade).first():
                        flash(f'Invalid learning area "{learning_area}" for grade "{grade}".', 'danger')
                        logger.error(f"Invalid learning area: {learning_area} for grade {grade}")
                        return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

                    saved, errors = 0, []
                    admission_nos = request.form.getlist('admission_no')
                    logger.debug(f"Submitting marks for {len(admission_nos)} students")

                    for adm_no in admission_nos:
                        marks_val = request.form.get(f'marks_{adm_no}')
                        if not marks_val:
                            logger.debug(f"No marks provided for {adm_no}")
                            continue
                        try:
                            marks = float(marks_val)  # Changed to float to match template's step="0.1"
                            if not (0 <= marks <= 100):
                                raise ValueError("Marks out of range")
                        except ValueError:
                            errors.append(f'Invalid marks for {adm_no}: {marks_val}')
                            logger.error(f"Invalid marks for {adm_no}: {marks_val}")
                            continue

                        student = db_session.query(Student).filter_by(admission_no=adm_no, grade=grade).first()
                        if not student:
                            errors.append(f'Invalid admission number: {adm_no}')
                            logger.error(f"Invalid admission number: {adm_no}")
                            continue

                        rec = db_session.query(Marks).filter_by(
                            admission_no=adm_no,
                            learning_area=learning_area,
                            exam_type=exam_type,
                            term=term,
                            year=year,
                            grade=grade
                        ).first()

                        pct = (marks / 100.0) * 100
                        if rec:
                            rec.marks = marks
                            rec.total_marks = marks
                            logger.debug(f"Updated marks for {adm_no}: {marks}")
                        else:
                            db_session.add(Marks(
                                admission_no=adm_no,
                                learning_area=learning_area,
                                exam_type=exam_type,
                                marks=marks,
                                total_marks=marks,
                                term=term,
                                year=year,
                                grade=grade,
                            ))
                            logger.debug(f"Added new marks for {adm_no}: {marks}")
                        saved += 1

                    if saved:
                        db_session.commit()
                        flash(f'Successfully saved {saved} marks.', 'success' if not errors else 'warning')
                        logger.info(f"Saved {saved} marks for grade={grade}, learning_area={learning_area}, term={term}, year={year}, exam_type={exam_type}")
                    else:
                        db_session.rollback()
                        flash('No marks saved due to errors.', 'danger')
                        logger.warning("No marks saved due to errors")

                    for err in errors:
                        flash(err, 'danger')

                    # Reset form and learners after submission
                    form = MarksForm()
                    form.grade.choices = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')] or [('Grade 7', 'Grade 7')]
                    form.term.choices = [('Term 1', 'Term 1'), ('Term 2', 'Term 2'), ('Term 3', 'Term 3')]
                    form.exam_type.choices = [
                        ('cat1', 'CAT 1'), ('cat2', 'CAT 2'), ('cat3', 'CAT 3'),
                        ('rat1', 'RAT 1'), ('rat2', 'RAT 2'), ('rat3', 'RAT 3'),
                        ('midterm', 'Mid Term'), ('endterm', 'End Term'),
                        ('project1', 'Project 1'), ('project2', 'Project 2'), ('project3', 'Project 3')
                    ]
                    form.year.choices = [(str(y), str(y)) for y in range(2020, 2030)]
                    if current_user.role == 'teacher':
                        assignments = db_session.query(TeacherAssignments.learning_area, TeacherAssignments.grade).filter_by(teacher_id=current_user.id).all()
                        form.learning_area.choices = [(f"{a.learning_area}|{a.grade}", f"{a.learning_area} ({a.grade})") for a in assignments] or [('', 'No subjects assigned')]
                    else:
                        learning_areas = db_session.query(LearningAreas.name, LearningAreas.grade).all()
                        form.learning_area.choices = [(f"{la.name}|{la.grade}", f"{la.name} ({la.grade})") for la in learning_areas]
                    learners = []
                    show_table = False

            return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)

    except SQLAlchemyError as e:
        logger.error(f"Database error: {str(e)}", exc_info=True)
        flash(f'Database error: {str(e)}', 'danger')
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        flash(f'Error: {str(e)}', 'danger')

    return render_template('enter_marks.html', form=form, learners=learners, show_table=show_table, term_info=term_info, content_data=content_data)
@app.route('/view_marks', methods=['GET', 'POST'])
@login_required
def view_marks():
    """View student marks for admin or teacher."""
    if current_user.role not in ['admin', 'teacher']:
        flash('Access denied!', 'danger')
        logger.warning(f'Access denied for user {current_user.id} with role {current_user.role}')
        return redirect(url_for('dashboard'))

    form = MarksFilterForm()
    form.term.data = form.term.data or 'Term 1'
    form.year.data = form.year.data or 2025
    marks_data = []
    learning_areas = [
        'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
        'Agriculture and Nutrition', 'Social Studies', 'Creative Arts', 'CRE'
    ] if current_user.role == 'admin' else []
    teacher_grades = []
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())

    try:
        # For teachers, fetch assigned grades and learning areas
        if current_user.role == 'teacher':
            assignments = db_session.query(TeacherAssignments.learning_area, TeacherAssignments.grade).filter_by(teacher_id=current_user.id).all()
            learning_areas = sorted([assignment.learning_area for assignment in assignments])
            teacher_grades = sorted(list(set([assignment.grade for assignment in assignments])))
            form.grade.choices = [(g, g) for g in teacher_grades]
            logger.debug(f'Teacher {current_user.id} assigned to grades: {teacher_grades}, subjects: {learning_areas}')
            
            if not teacher_grades:
                flash('You are not assigned to any grades or subjects.', 'warning')
                return render_template('view_marks.html', form=form, marks_data=marks_data, learning_areas=learning_areas,
                                      edit_route='edit_marks', role=current_user.role, term_info=term_info, content_data=content_data)

        if form.validate_on_submit():
            grade = form.grade.data
            term = form.term.data
            year = form.year.data
            exam_type = form.exam_type.data
            logger.debug(f'Viewing marks for grade={grade}, term={term}, year={year}, exam_type={exam_type}')

            if current_user.role == 'teacher' and grade not in teacher_grades:
                flash('You are not assigned to this grade!', 'danger')
                return render_template('view_marks.html', form=form, marks_data=marks_data, learning_areas=learning_areas,
                                      edit_route='edit_marks', role=current_user.role, term_info=term_info, content_data=content_data)

            students = db_session.query(User.admission_no, User.username).filter_by(
                role='student', grade=grade
            ).order_by(User.admission_no.asc()).all()
            logger.debug(f'Found {len(students)} students for grade {grade}')

            if not students:
                flash('No students found for the selected grade.', 'warning')
                return render_template('view_marks.html', form=form, marks_data=marks_data, learning_areas=learning_areas,
                                      edit_route='edit_marks', role=current_user.role, term_info=term_info, content_data=content_data)

            marks_query = db_session.query(Marks).filter_by(
                grade=grade, term=term, year=str(year), exam_type=exam_type
            ).all()
            marks_dict = {(m.admission_no, m.learning_area): m for m in marks_query}

            for student in students:
                admission_no, name = student
                row = {'admission_no': admission_no, 'name': name}
                total = 0
                for la in learning_areas:
                    mark = marks_dict.get((admission_no, la))
                    if mark and mark.total_marks is not None:
                        row[la] = mark.total_marks
                        total += mark.total_marks
                    else:
                        row[la] = '-'
                row['total'] = total if total > 0 else '-'
                marks_data.append(row)
                logger.debug(f'Student {admission_no} total: {total}')

        edit_route = 'edit_marks'
        return render_template('view_marks.html', form=form, marks_data=marks_data, learning_areas=learning_areas,
                              edit_route=edit_route, role=current_user.role, term_info=term_info, content_data=content_data)

    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in view_marks: {str(e)}\n{traceback.format_exc()}")
        flash('Error retrieving marks: Please try again.', 'danger')
        return render_template('view_marks.html', form=form, marks_data=marks_data, learning_areas=learning_areas,
                              edit_route='edit_marks', role=current_user.role, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/edit_marks/<admission_no>/<learning_area>/<exam_type>/<term>/<year>', methods=['GET', 'POST'])
@login_required
def edit_marks(admission_no, learning_area, exam_type, term, year):
    """Edit marks for a student as admin or teacher."""
    if current_user.role not in ['admin', 'teacher']:
        flash('Access denied!', 'danger')
        logger.warning(f'Access denied for user {current_user.id} with role {current_user.role}')
        return redirect(url_for('view_marks'))

    form = MarksForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())

    try:
        # Restrict teachers to their assigned grades and subjects
        if current_user.role == 'teacher':
            assignments = db_session.query(TeacherAssignments.learning_area, TeacherAssignments.grade).filter_by(teacher_id=current_user.id).all()
            teacher_subjects = [a.learning_area for a in assignments]
            teacher_grades = [a.grade for a in assignments]
            if learning_area not in teacher_subjects:
                flash('You are not authorized to edit marks for this subject.', 'danger')
                logger.warning(f'Teacher {current_user.id} attempted to edit unassigned subject {learning_area}')
                return redirect(url_for('view_marks'))

        # Fetch existing marks
        marks = db_session.query(Marks).filter_by(
            admission_no=admission_no,
            learning_area=learning_area,
            exam_type=exam_type,
            term=term,
            year=year
        ).first()

        # Fetch student and check grade authorization
        student = db_session.query(User).filter_by(admission_no=admission_no, role='student').first()
        if not student:
            flash(f'Invalid admission number: {admission_no}', 'danger')
            return redirect(url_for('view_marks'))
        if current_user.role == 'teacher' and student.grade not in teacher_grades:
            flash('You are not authorized to edit marks for this student.', 'danger')
            logger.warning(f'Teacher {current_user.id} attempted to edit marks for unassigned grade {student.grade}')
            return redirect(url_for('view_marks'))

        # Populate form on GET request
        if request.method == 'GET':
            form.grade.data = student.grade
            form.learning_area.data = learning_area
            form.exam_type.data = exam_type
            form.term.data = term
            form.year.data = int(year)
            form.marks.data = marks.total_marks if marks else None
            if not marks:
                flash('No existing marks found. Enter new marks.', 'info')

        if form.validate_on_submit():
            total_marks = form.marks.data
            if not marks:
                marks = Marks(
                    admission_no=admission_no,
                    grade=form.grade.data,
                    learning_area=learning_area,
                    exam_type=exam_type,
                    term=term,
                    year=str(year),
                    total_marks=total_marks
                )
                db_session.add(marks)
            else:
                marks.total_marks = total_marks
            db_session.commit()
            flash('Marks updated successfully!', 'success')
            logger.info(f'Marks updated for admission_no={admission_no}, learning_area={learning_area}, exam_type={exam_type}, term={term}, year={year}')
            return redirect(url_for('view_marks'))

        return render_template('edit_marks.html', form=form, admission_no=admission_no,
                              learning_area=learning_area, exam_type=exam_type, term=term, year=year,
                              term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in edit_marks: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error processing marks: {str(e)}', 'danger')
        return render_template('edit_marks.html', form=form, admission_no=admission_no,
                              learning_area=learning_area, exam_type=exam_type, term=term, year=year,
                              term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
        

        
@app.route('/admin/view_users', methods=['GET', 'POST'])
@login_required
@admin_required
def view_users():
    """View all users grouped by role and grade, sorted by username for non-students and admission_no for students."""
    db_session = next(get_db())
    try:
        if request.method == 'POST':
            user_id = request.form.get('user_id')
            new_username = request.form.get('username')
            new_admission_no = request.form.get('admission_no')
            
            user = db_session.query(User).filter_by(id=user_id).first()
            if user:
                # Update username if provided and different
                if new_username and new_username != user.username:
                    user.username = new_username
                
                # Update admission number if provided and different (only for students)
                if new_admission_no and user.role == 'student' and new_admission_no != user.admission_no:
                    user.admission_no = new_admission_no
                
                db_session.commit()
                flash('User details updated successfully', 'success')
            else:
                flash('User not found', 'danger')
            
            return redirect(url_for('view_users'))

        users = db_session.query(User.id, User.username, User.role, User.grade, User.admission_no, User.password_hash).order_by(User.role, User.grade).all()
        term_info, content_data = fetch_common_data()  # Fetch term_info and content_data

        # Group users by role and grade
        grouped_users = {
            'admin': [u for u in users if u.role == 'admin'],
            'teacher': [u for u in users if u.role == 'teacher'],
            'parent': [u for u in users if u.role == 'parent'],
            'bursar': [u for u in users if u.role == 'bursar'],
            'learners_grade7': [u for u in users if u.role == 'student' and u.grade == 'Grade 7'],
            'learners_grade8': [u for u in users if u.role == 'student' and u.grade == 'Grade 8'],
            'learners_grade9': [u for u in users if u.role == 'student' and u.grade == 'Grade 9']
        }

        # Sort non-student groups by username
        for role in ['admin', 'teacher', 'parent', 'bursar']:
            grouped_users[role] = sorted(grouped_users[role], key=lambda u: u.username.lower())

        # Sort student groups by admission_no (numeric order)
        for grade in ['learners_grade7', 'learners_grade8', 'learners_grade9']:
            grouped_users[grade] = sorted(grouped_users[grade], key=lambda u: int(u.admission_no) if u.admission_no else float('inf'))

        return render_template('view_users.html', grouped_users=grouped_users, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in view_users: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error retrieving users: {str(e)}', 'danger')
        return render_template('view_users.html', grouped_users={}, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Delete a user and related data."""
    db_session = next(get_db())
    try:
        user = db_session.query(User).filter_by(id=user_id).first()
        if not user:
            flash('User not found!', 'danger')
            return redirect(url_for('view_users'))
        if user.role == 'admin' and current_user.id == user_id:
            flash('Cannot delete your own account!', 'danger')
        else:
            db_session.query(TeacherAssignments).filter_by(teacher_id=user_id).delete()
            db_session.query(Marks).filter(Marks.admission_no.in_(db_session.query(User.admission_no).filter_by(id=user_id))).delete(synchronize_session=False)
            db_session.query(Fee).filter(Fee.admission_no.in_(db_session.query(User.admission_no).filter_by(id=user_id))).delete(synchronize_session=False)
            db_session.delete(user)
            db_session.commit()
            flash('User deleted successfully!', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
        logger.error(f"Database error in delete_user: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('view_users'))

@app.route('/admin/assign_teachers', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_teachers():
    """Assign teachers to learning areas and grades."""
    db_session = next(get_db())
    try:
        term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
        if request.method == 'POST':
            teacher_id = request.form.get('teacher_id')
            learning_area = request.form.get('learning_area')
            grade = request.form.get('grade')
            try:
                assignment = TeacherAssignments(teacher_id=teacher_id, learning_area=learning_area, grade=grade)
                db_session.add(assignment)
                db_session.commit()
                flash('Teacher assigned successfully!', 'success')
                logger.info(f"Teacher assigned: teacher_id={teacher_id}, learning_area={learning_area}, grade={grade}")
            except IntegrityError:
                db_session.rollback()
                flash('Teacher assignment already exists!', 'danger')
                logger.warning(f"Duplicate assignment: teacher_id={teacher_id}, learning_area={learning_area}, grade={grade}")
            except Exception as e:
                db_session.rollback()
                flash(f'Error assigning teacher: {str(e)}', 'danger')
                logger.error(f"Error in assign_teachers: {str(e)}\n{traceback.format_exc()}")
            return redirect(url_for('assign_teachers'))

        teachers = db_session.query(User.id, User.username).filter_by(role='teacher').all()
        learning_areas = db_session.query(LearningAreas.name, LearningAreas.grade).all()
        assignments = db_session.query(TeacherAssignments.teacher_id, User.username, TeacherAssignments.learning_area, TeacherAssignments.grade).join(User, TeacherAssignments.teacher_id == User.id).all()
        return render_template('assign_teachers.html', teachers=teachers, learning_areas=learning_areas, assignments=assignments, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in assign_teachers: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template('assign_teachers.html', teachers=[], learning_areas=[], assignments=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()

@app.route('/admin/delete_teacher_assignment/<int:teacher_id>/<learning_area>/<grade>', methods=['POST'])
@login_required
@admin_required
def delete_teacher_assignment(teacher_id, learning_area, grade):
    """Delete a teacher assignment."""
    db_session = next(get_db())
    try:
        assignment = db_session.query(TeacherAssignments).filter_by(teacher_id=teacher_id, learning_area=learning_area, grade=grade).first()
        if assignment:
            db_session.delete(assignment)
            db_session.commit()
            flash('Teacher assignment deleted successfully!', 'success')
        else:
            flash('Assignment not found!', 'danger')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting assignment: {str(e)}', 'danger')
        logger.error(f"Database error in delete_teacher_assignment: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('assign_teachers'))

@app.route('/admin/assign_class_teacher', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_class_teacher():
    """Assign class teachers to grades."""
    db_session = next(get_db())
    try:
        term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
        if request.method == 'POST':
            teacher_id = request.form.get('teacher_id')
            grade = request.form.get('grade')
            try:
                assignment = ClassTeachers(teacher_id=teacher_id, grade=grade)
                db_session.add(assignment)
                db_session.commit()
                flash('Class teacher assigned successfully!', 'success')
                logger.info(f"Class teacher assigned: teacher_id={teacher_id}, grade={grade}")
            except IntegrityError:
                db_session.rollback()
                flash('Class teacher assignment already exists!', 'danger')
                logger.warning(f"Duplicate class teacher assignment: teacher_id={teacher_id}, grade={grade}")
            except Exception as e:
                db_session.rollback()
                flash(f'Error assigning class teacher: {str(e)}', 'danger')
                logger.error(f"Error in assign_class_teacher: {str(e)}\n{traceback.format_exc()}")
            return redirect(url_for('assign_class_teacher'))

        teachers = db_session.query(User.id, User.username).filter_by(role='teacher').all()
        class_assignments = db_session.query(ClassTeachers.teacher_id, User.username, ClassTeachers.grade).join(User, ClassTeachers.teacher_id == User.id).all()
        return render_template('assign_class_teacher.html', teachers=teachers, class_assignments=class_assignments, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in assign_class_teacher: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template('assign_class_teacher.html', teachers=[], class_assignments=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
@app.route('/admin/delete_class_teacher_assignment/<int:teacher_id>/<grade>', methods=['POST'])
@login_required
@admin_required
def delete_class_teacher_assignment(teacher_id, grade):
    """Delete a class teacher assignment."""
    db_session = next(get_db())
    try:
        assignment = db_session.query(ClassTeachers).filter_by(teacher_id=teacher_id, grade=grade).first()
        if assignment:
            db_session.delete(assignment)
            db_session.commit()
            flash('Class teacher assignment deleted successfully!', 'success')
        else:
            flash('Assignment not found!', 'danger')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting class teacher assignment: {str(e)}', 'danger')
        logger.error(f"Database error in delete_class_teacher_assignment: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('assign_class_teacher'))

@app.route('/admin/manage_learning_areas', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_learning_areas():
    """Manage learning areas."""
    term_info, content_data = fetch_common_data()
    db_session = next(get_db())
    try:
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            grade = request.form.get('grade', '').strip()
            if not name or not grade:
                flash('Learning area name and grade are required!', 'danger')
                return redirect(url_for('manage_learning_areas'))
            if grade not in ['Grade 7', 'Grade 8', 'Grade 9']:
                flash('Invalid grade selected!', 'danger')
                return redirect(url_for('manage_learning_areas'))
            try:
                learning_area = LearningAreas(name=name, grade=grade)
                db_session.add(learning_area)
                db_session.commit()
                flash('Learning area added successfully!', 'success')
            except IntegrityError:
                db_session.rollback()
                flash('Learning area already exists for this grade!', 'danger')
            except SQLAlchemyError as e:
                db_session.rollback()
                flash(f'Error adding learning area: {str(e)}', 'danger')
            return redirect(url_for('manage_learning_areas'))

        learning_areas = db_session.query(LearningAreas.id, LearningAreas.name, LearningAreas.grade).order_by(LearningAreas.grade, LearningAreas.name).all()
        return render_template('manage_learning_areas.html', learning_areas=learning_areas, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Database error in manage_learning_areas: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template('manage_learning_areas.html', learning_areas=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/admin/delete_learning_area/<int:area_id>', methods=['POST'])
@login_required
@admin_required
def delete_learning_area(area_id):
    """Delete a learning area and related data."""
    db_session = next(get_db())
    try:
        learning_area = db_session.query(LearningAreas).filter_by(id=area_id).first()
        if not learning_area:
            flash('Learning area not found!', 'danger')
            return redirect(url_for('manage_learning_areas'))
        learning_area_name = learning_area.name
        db_session.query(TeacherAssignments).filter_by(learning_area=learning_area_name).delete()
        db_session.query(Marks).filter_by(learning_area=learning_area_name).delete()
        db_session.delete(learning_area)
        db_session.commit()
        flash('Learning area and related assignments/marks deleted successfully!', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting learning area: {str(e)}', 'danger')
        logger.error(f"Database error in delete_learning_area: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('manage_learning_areas'))

@app.route('/admin/delete_all_learning_areas', methods=['POST'])
@login_required
@admin_required
def delete_all_learning_areas():
    """Delete all learning areas and related data."""
    db_session = next(get_db())
    try:
        learning_areas = db_session.query(LearningAreas).all()
        if not learning_areas:
            flash('No learning areas found to delete!', 'danger')
            return redirect(url_for('manage_learning_areas'))
        for area in learning_areas:
            db_session.query(TeacherAssignments).filter_by(learning_area=area.name).delete()
            db_session.query(Marks).filter_by(learning_area=area.name).delete()
        db_session.query(LearningAreas).delete()
        db_session.commit()
        flash('All learning areas and related assignments/marks deleted successfully!', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting all learning areas: {str(e)}', 'danger')
        logger.error(f"Database error in delete_all_learning_areas: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('manage_learning_areas'))

@app.route('/admin/manage_performance_levels', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_performance_levels():
    """Manage performance levels."""
    form = PerformanceLevelForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        if form.validate_on_submit():
            min_marks = form.min_marks.data
            max_marks = form.max_marks.data
            level = form.level.data.strip()
            points = form.points.data
            comment = form.comment.data.strip()
            type_ = form.type.data.strip()

            if min_marks < 0 or max_marks < 0:
                flash('Marks cannot be negative!', 'danger')
                return render_template('manage_performance_levels.html', form=form, performance_levels=[], term_info=term_info, content_data=content_data)
            if min_marks > max_marks:
                flash('Minimum marks cannot exceed maximum marks!', 'danger')
                return render_template('manage_performance_levels.html', form=form, performance_levels=[], term_info=term_info, content_data=content_data)
            if type_ not in ['grade', 'subject']:
                flash('Invalid type selected!', 'danger')
                return render_template('manage_performance_levels.html', form=form, performance_levels=[], term_info=term_info, content_data=content_data)

            overlap = db_session.query(PerformanceLevels).filter(
                PerformanceLevels.type == type_,
                ((PerformanceLevels.min_marks <= max_marks) & (PerformanceLevels.max_marks >= min_marks) |
                 (PerformanceLevels.min_marks <= min_marks) & (PerformanceLevels.max_marks >= max_marks))
            ).first()
            if overlap:
                flash('Performance level range overlaps with an existing level!', 'danger')
                return render_template('manage_performance_levels.html', form=form, performance_levels=[], term_info=term_info, content_data=content_data)

            try:
                performance_level = PerformanceLevels(
                    min_marks=min_marks,
                    max_marks=max_marks,
                    level=level,
                    points=points,
                    comment=comment,
                    type=type_
                )
                db_session.add(performance_level)
                db_session.commit()
                flash('Performance level added/updated successfully!', 'success')
                logger.info(f"Performance level added/updated: {level} ({type_}, {min_marks}-{max_marks}) by admin {current_user.username}")
            except IntegrityError:
                db_session.rollback()
                flash('Performance level already exists for this range!', 'danger')
            except Exception as e:
                db_session.rollback()
                flash(f'Error updating performance level: {str(e)}', 'danger')
                logger.error(f"Error updating performance level: {str(e)}\n{traceback.format_exc()}")

        performance_levels = db_session.query(PerformanceLevels.id, PerformanceLevels.min_marks, PerformanceLevels.max_marks, PerformanceLevels.level, PerformanceLevels.points, PerformanceLevels.comment, PerformanceLevels.type).order_by(PerformanceLevels.type, PerformanceLevels.min_marks).all()
        return render_template('manage_performance_levels.html', form=form, performance_levels=performance_levels, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in manage_performance_levels: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template('manage_performance_levels.html', form=form, performance_levels=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/admin/delete_performance_level/<int:level_id>', methods=['POST'])
@login_required
@admin_required
def delete_performance_level(level_id):
    """Delete a performance level."""
    db_session = next(get_db())
    try:
        level_data = db_session.query(PerformanceLevels).filter_by(id=level_id).first()
        if not level_data:
            flash('Performance level not found!', 'danger')
            return redirect(url_for('manage_performance_levels'))
        db_session.delete(level_data)
        db_session.commit()
        flash(f'Performance level "{level_data.level}" ({level_data.type}) deleted successfully!', 'success')
        logger.info(f"Performance level ID {level_id} ({level_data.level}, {level_data.type}) deleted by admin {current_user.username}")
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting performance level: {str(e)}', 'danger')
        logger.error(f"Database error deleting performance level ID {level_id}: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('manage_performance_levels'))

@app.route('/admin/delete_all_performance_levels', methods=['POST'])
@login_required
@admin_required
def delete_all_performance_levels():
    """Delete all performance levels."""
    db_session = next(get_db())
    try:
        performance_levels = db_session.query(PerformanceLevels).all()
        if not performance_levels:
            flash('No performance levels found to delete!', 'danger')
            return redirect(url_for('manage_performance_levels'))
        db_session.query(PerformanceLevels).delete()
        db_session.commit()
        flash('All performance levels deleted successfully!', 'success')
        logger.info(f"All performance levels deleted by admin {current_user.username}")
    except SQLAlchemyError as e:
        db_session.rollback()
        flash(f'Error deleting all performance levels: {str(e)}', 'danger')
        logger.error(f"Database error deleting all performance levels: {str(e)}\n{traceback.format_exc()}")
    finally:
        db_session.close()
    return redirect(url_for('manage_performance_levels'))



@app.route('/admin/update_term_info', methods=['GET', 'POST'])
@login_required
@admin_required
def update_term_info():
    """Update term information as admin."""
    form = TermInfoForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        if form.validate_on_submit():
            term = form.term.data
            year = form.year.data
            principal = form.principal.data
            start_date = form.start_date.data
            end_date = form.end_date.data

            term_data = db_session.query(TermInfo).filter_by(id=1).first()
            if term_data:
                term_data.term = term
                term_data.year = year
                term_data.principal = principal
                term_data.start_date = start_date
                term_data.end_date = end_date
            else:
                term_data = TermInfo(
                    id=1,
                    term=term,
                    year=year,
                    principal=principal,
                    start_date=start_date,
                    end_date=end_date
                )
                db_session.add(term_data)
            db_session.commit()
            flash('Term information updated successfully!', 'success')
            logger.info(f"Term info updated by admin {current_user.username}")
            return redirect(url_for('dashboard'))

        return render_template('update_term_info.html', form=form, term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        flash(f'Error updating term information: {str(e)}', 'danger')
        logger.error(f"Error updating term information: {str(e)}\n{traceback.format_exc()}")
        return render_template('update_term_info.html', form=form, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
@app.route('/admin/edit_content', defaults={'content_type': None}, methods=['GET', 'POST'])
@app.route('/admin/edit_content/<content_type>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_content(content_type):
    """Edit mission, vision, about, or contact content."""
    content_types = ['mission', 'vision', 'about', 'contact']
    db_session = next(get_db())
    try:
        # Fetch term_info and content_data
        term_info, content_data = fetch_common_data()

        if request.method == 'POST':
            action = request.form.get('action')
            content_type = request.form.get('content_type')

            if content_type not in content_types:
                flash('Invalid content type!', 'danger')
                return redirect(url_for('edit_content'))

            Model = {'mission': Mission, 'vision': Vision, 'about': About, 'contact': Contact}.get(content_type)
            if action == 'delete':
                db_session.query(Model).filter_by(id=1).delete()
                db_session.commit()
                flash(f'{content_type.capitalize()} content deleted successfully!', 'success')
            else:
                content = request.form.get('content')
                existing = db_session.query(Model).filter_by(id=1).first()
                if existing:
                    existing.content = content
                else:
                    new_content = Model(id=1, content=content)
                    db_session.add(new_content)
                db_session.commit()
                flash(f'{content_type.capitalize()} updated successfully!', 'success')
            return redirect(url_for('admin_dashboard'))

        if not content_type:
            return render_template('edit_content.html', content_types=content_types, content=None, term_info=term_info, content_data=content_data)

        Model = {'mission': Mission, 'vision': Vision, 'about': About, 'contact': Contact}.get(content_type)
        result = db_session.query(Model.content).filter_by(id=1).first()
        content = result.content if result else ""
        return render_template('edit_content.html', content_types=content_types, content=content, content_type=content_type, term_info=term_info, content_data=content_data)
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in edit_content ({content_type}): {str(e)}\n{traceback.format_exc()}")
        flash(f'Error processing {content_type}: {str(e)}', 'danger')
        return render_template(
            'edit_content.html',
            content_types=content_types,
            content=None,
            content_type=content_type,
            term_info=term_info,
            content_data=content_data
        )
    finally:
        db_session.close()
    return render_template('edit_content.html', content_type=content_type, content=content, content_types=content_types, term_info=term_info)
@app.route('/announcements', methods=['GET'])
@login_required
def announcements():
    """Display announcements."""
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        announcements = db_session.query(Announcements.id, Announcements.content, Announcements.date).order_by(Announcements.date.desc()).all()
        return render_template('announcements.html', announcements=[{'id': a.id, 'content': a.content, 'date': a.date} for a in announcements], term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in announcements: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error fetching announcements: {str(e)}', 'danger')
        return render_template('announcements.html', announcements=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/add_announcement', methods=['GET', 'POST'])
@login_required
@admin_required
def add_announcement():
    """Add a new announcement."""
    form = AnnouncementsForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        recent_announcements = db_session.query(Announcements.id, Announcements.content, Announcements.date).order_by(Announcements.date.desc()).limit(5).all()
        if form.validate_on_submit():
            content = form.content.data
            date = datetime.now()
            announcement = Announcements(content=content, date=date)
            db_session.add(announcement)
            db_session.commit()
            flash('Announcement added successfully!', 'success')
            logger.info(f"Announcement added by admin {current_user.username}")
            return redirect(url_for('announcements'))

        return render_template('add_announcement.html', form=form, recent_announcements=[{'id': a.id, 'content': a.content, 'date': a.date} for a in recent_announcements], term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error adding announcement: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error adding announcement: {str(e)}', 'danger')
        return render_template('add_announcement.html', form=form, recent_announcements=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
@app.route('/delete_announcement', methods=['POST'])
@login_required
@admin_required
def delete_announcement():
    """Delete an announcement."""
    db_session = next(get_db())
    try:
        announcement_id = request.form.get('announcement_id')
        if not announcement_id:
            flash('Invalid request.', 'danger')
            return redirect(url_for('announcements'))
        result = db_session.query(Announcements).filter_by(id=announcement_id).delete()
        if result == 0:
            flash('No announcement found to delete.', 'warning')
        else:
            db_session.commit()
            flash('Announcements deleted successfully.', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in delete_announcement: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error deleting announcement: {str(e)}', 'danger')
    finally:
        db_session.close()
    return redirect(url_for('announcements'))

@app.route('/admin/send_message', methods=['GET', 'POST'])
@login_required
def send_message():
    """Send a message as admin or teacher."""
    if current_user.role not in ['admin', 'teacher']:
        flash('Access denied!', 'danger')
        logger.warning(f"Unauthorized access attempt by user_id={current_user.id}, role={current_user.role}")
        return redirect(url_for('dashboard'))
    
    form = MessageForm()
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        if form.validate_on_submit():
            content = form.content.data
            recipient_role = form.recipient_role.data
            message = Messages(
                sender_id=current_user.id,
                content=content,
                date=datetime.now(),
                recipient_role=recipient_role
            )
            db_session.add(message)
            db_session.commit()
            flash('Message sent successfully!', 'success')
            logger.info(f"Message sent by {current_user.id} to {recipient_role}")
            return redirect(url_for('dashboard'))
        
        return render_template('send_message.html', form=form, term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        flash(f'Error sending message: {str(e)}', 'danger')
        logger.error(f"Error in send_message: {str(e)}\n{traceback.format_exc()}")
        return render_template('send_message.html', form=form, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/admin/view_messages', methods=['GET'])
@login_required
@admin_required
def view_messages():
    """View messages sent to parents."""
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        messages = db_session.query(Messages.content, Messages.date, User.username).join(User, Messages.sender_id == User.id).filter(Messages.recipient_role == 'parent').order_by(Messages.date.desc()).all()
        messages = [{'content': m.content, 'date': m.date, 'username': m.username} for m in messages]
        return render_template('view_messages.html', messages=messages, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in view_messages: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error fetching messages: {str(e)}', 'danger')
        return render_template('view_messages.html', messages=[], term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
        
        
        
@app.route('/student_view_results', methods=['GET', 'POST'])
@login_required
def student_view_results(admission_no=None, grade=None, term=None, year=None, exam_type=None):
    """View results for a student, allowing selection of grade, term, year, and exam type."""
    db_session = next(get_db())
    term_info, content_data = fetch_common_data()
    try:
        # Validate user role
        if current_user.role not in ['student', 'parent']:
            logger.error(f"Unauthorized role {current_user.role} for user {current_user.id}")
            flash('You are not authorized to access this page.', 'danger')
            return redirect(url_for('index'))

        # Determine admission number and validate access
        parent_view = False
        if current_user.role == 'student':
            if admission_no and admission_no != current_user.admission_no:
                logger.warning(f"Student {current_user.id} attempted to access results for admission_no={admission_no}")
                flash('You can only view your own results.', 'danger')
                return redirect(url_for('student_dashboard'))
            admission_no = current_user.admission_no
        elif current_user.role == 'parent':
            if not admission_no:
                logger.warning(f"Parent {current_user.id} accessed /student_view_results without admission_no")
                flash('Please select a linked student.', 'danger')
                return redirect(url_for('parent_dashboard'))
            parent_student = db_session.query(ParentStudent).filter_by(
                parent_id=current_user.id,
                admission_no=admission_no
            ).first()
            if not parent_student:
                logger.warning(f"Parent {current_user.id} not linked to admission_no={admission_no}")
                flash('You are not authorized to view this student’s results.', 'danger')
                return redirect(url_for('parent_dashboard'))
            parent_view = True

        # Fetch student data
        student = db_session.query(Student.admission_no, Student.name, Student.grade).filter_by(admission_no=admission_no).first()
        if not student:
            logger.warning(f"Student profile not found for admission_no={admission_no}")
            flash('Student profile not found.', 'danger')
            return redirect(url_for('logout' if current_user.role == 'student' else 'parent_dashboard'))

        admission_no, student_name, student_grade = student

        # Fetch available grades from Marks
        grades = db_session.query(Marks.grade).filter_by(admission_no=admission_no).distinct().order_by(Marks.grade).all()
        grade_choices = [(g[0], g[0]) for g in grades if g[0] in [grade[0] for grade in GRADES if grade[0] != 'all']] or [(student_grade, student_grade)]

        # Initialize form
        form = ResultsFilterForm(
            admission_no=admission_no,
            grade=student_grade if student_grade in [g[0] for g in grade_choices] else grade_choices[0][0],
            term=term or (term_info.term if term_info and term_info.term else 'Term 1'),
            year=int(year or (term_info.year if term_info and term_info.year else str(datetime.now().year))),
            exam_type=exam_type or 'endterm'
        )

        # Set form choices
        form.admission_no.choices = [(admission_no, f"{student_name} ({admission_no})")]
        form.grade.choices = grade_choices
        form.term.choices = TERMS
        form.exam_type.choices = EXAM_TYPES

        marks = []
        fee = None
        if request.method == 'POST' and form.validate_on_submit():
            grade = form.grade.data
            term = form.term.data
            year = str(form.year.data)
            exam_type = form.exam_type.data.lower()

            # Validate grade for students
            if current_user.role == 'student' and grade != student_grade:
                logger.warning(f"Student {current_user.id} attempted to select invalid grade {grade}")
                flash('You can only view results for your current grade.', 'danger')
                return render_template(
                    'student_view_results.html',
                    form=form,
                    marks=[],
                    fee=None,
                    student={'admission_no': admission_no, 'name': student_name, 'grade': student_grade},
                    term_info=term_info,
                    content_data=content_data,
                    parent_view=parent_view
                )

            # Fetch filtered results
            marks = db_session.query(Marks.learning_area, Marks.total_marks, Marks.exam_type, Marks.term, Marks.year, Marks.grade).filter_by(
                admission_no=admission_no,
                grade=grade,
                term=term,
                year=year,
                exam_type=exam_type
            ).all()

            # Fetch fee information
            fee = db_session.query(Fee.total_fee, Fee.amount_paid, Fee.balance, Fee.grade, Fee.term, Fee.year).filter_by(
                admission_no=admission_no,
                grade=grade,
                term=term,
                year=year
            ).first()

            if not marks and not fee:
                exam_type_display = next((et[1] for et in EXAM_TYPES if et[0] == exam_type), exam_type)
                flash(f'No results or fees found for {grade} {term} {year} ({exam_type_display}).', 'warning')
        else:
            # Fetch recent results for GET request
            marks = db_session.query(Marks.learning_area, Marks.total_marks, Marks.exam_type, Marks.term, Marks.year, Marks.grade).filter_by(
                admission_no=admission_no,
                grade=student_grade
            ).order_by(Marks.year.desc(), Marks.term.desc()).limit(10).all()
            fee = db_session.query(Fee.total_fee, Fee.amount_paid, Fee.balance, Fee.grade, Fee.term, Fee.year).filter_by(
                admission_no=admission_no,
                grade=student_grade
            ).order_by(Fee.year.desc(), Fee.term.desc()).first()

        # Format marks and fee for template
        marks = [{'learning_area': m.learning_area, 'total_marks': m.total_marks, 'exam_type': m.exam_type, 'term': m.term, 'year': m.year, 'grade': m.grade} for m in marks]
        fee = {'total_fee': fee.total_fee, 'amount_paid': fee.amount_paid, 'balance': fee.balance, 'grade': fee.grade, 'term': fee.term, 'year': fee.year} if fee else None

        return render_template(
            'student_view_results.html',
            form=form,
            marks=marks,
            fee=fee,
            student={'admission_no': admission_no, 'name': student_name, 'grade': student_grade},
            term_info=term_info,
            content_data=content_data,
            parent_view=parent_view
        )

    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in student_view_results for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        flash('Database error: Unable to retrieve results. Please try again later.', 'danger')
        return render_template(
            'student_view_results.html',
            form=form,
            marks=[],
            fee=None,
            student={'admission_no': admission_no or '', 'name': student_name or '', 'grade': student_grade or ''},
            term_info=term_info,
            content_data=content_data,
            parent_view=parent_view
        )
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in student_view_results for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        flash('An unexpected error occurred. Please try again later.', 'danger')
        return render_template(
            'student_view_results.html',
            form=form,
            marks=[],
            fee=None,
            student={'admission_no': admission_no or '', 'name': student_name or '', 'grade': student_grade or ''},
            term_info=term_info,
            content_data=content_data,
            parent_view=parent_view
        )
    finally:
        db_session.close()
        
        
@app.route('/parent/view_student', methods=['POST'])
@login_required
@parent_required
def view_student():
    """View student details for a parent."""
    term_info, content_data = fetch_common_data()  # Fetch term_info and content_data
    db_session = next(get_db())
    try:
        admission_no = request.form.get('admission_no', '').strip()
        if not admission_no:
            flash('Please enter an admission number.', 'danger')
            logger.warning(f"No admission number provided by parent_id={current_user.id}")
            return redirect(url_for('parent_dashboard'))

        linked_student = db_session.query(ParentStudent.admission_no).filter_by(parent_id=current_user.id, admission_no=admission_no).first()
        if not linked_student:
            flash('You are not authorized to view this student or student not found.', 'danger')
            logger.warning(f"Parent {current_user.id} not authorized for admission_no: {admission_no}")
            return redirect(url_for('parent_dashboard'))

        student = db_session.query(Student.admission_no, Student.name, Student.grade).filter_by(admission_no=admission_no).first()
        if not student:
            flash('No learner found with that admission number.', 'danger')
            logger.warning(f"Student not found for admission_no: {admission_no}")
            return redirect(url_for('parent_dashboard'))

        logger.debug(f"Redirecting to student_dashboard for admission_no: {admission_no}")
        return redirect(url_for('student_dashboard', admission_no=admission_no))
    except Exception as e:
        logger.error(f"Error in view_student: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('parent_dashboard'))
    finally:
        db_session.close()
        
        
@app.route('/parent/get_student_details/<admission_no>')
@login_required
@parent_required
def get_student_details(admission_no):
    """Get student details for a parent via JSON."""
    db_session = next(get_db())
    try:
        linked_student = db_session.query(ParentStudent.admission_no).filter_by(parent_id=current_user.id, admission_no=admission_no).first()
        if not linked_student:
            logger.warning(f"Parent {current_user.id} not authorized for admission_no: {admission_no}")
            return jsonify({'error': 'Unauthorized access'}), 403

        marks_data = db_session.query(Marks.admission_no, Marks.learning_area, Marks.marks, Marks.exam_type, Marks.term, Marks.year).filter_by(admission_no=admission_no).all()
        fees_data = db_session.query(Fee.total_fee, Fee.amount_paid, Fee.balance).filter_by(admission_no=admission_no).first()

        data = {
            'marks': [{'learning_area': m.learning_area, 'marks': m.marks, 'exam_type': m.exam_type, 'term': m.term, 'year': m.year} for m in marks_data],
            'fees': {'total_fee': fees_data.total_fee, 'amount_paid': fees_data.amount_paid, 'balance': fees_data.balance} if fees_data else {'total_fee': 0, 'amount_paid': 0, 'balance': 0}
        }
        return jsonify(data)
    except Exception as e:
        logger.error(f"Error in get_student_details: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'error': f'Error fetching student details: {str(e)}'}), 500
    finally:
        db_session.close()

def get_available_grades():
    """Get available grades."""
    db_session = next(get_db())
    try:
        grades = db_session.query(Student.grade).distinct().all()
        return [(g.grade, g.grade) for g in grades]
    finally:
        db_session.close()

def generate_report_card(students, marks, fees, term, year, exam_type, rank, total_students, grade):
    """Generate a report card PDF."""
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 750, f"Report Card - {term} {year}")
    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer
from flask import render_template, request, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from datetime import datetime
import logging
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func
import traceback

logger = logging.getLogger(__name__)



@app.route('/student_download_report_card', methods=['GET', 'POST'])
@login_required
def student_download_report_card(admission_no=None, grade=None, term=None, year=None, exam_type=None):
    """Download report card for a student as PDF, accessible by both parents and students."""
    db_session = next(get_db())
    try:
        # Validate user role
        if current_user.role not in ['student', 'parent']:
            logger.error(f"Unauthorized role {current_user.role} for user {current_user.id}")
            flash('You are not authorized to access this page.', 'danger')
            return redirect(url_for('index'))

        # Determine admission number and validate access
        parent_view = False
        if current_user.role == 'student':
            if admission_no and admission_no != current_user.admission_no:
                logger.warning(f"Student {current_user.id} attempted to access report card for admission_no={admission_no}")
                flash('You can only download your own report card.', 'danger')
                return redirect(url_for('student_dashboard'))
            admission_no = current_user.admission_no
        elif current_user.role == 'parent':
            if not admission_no:
                logger.warning(f"Parent {current_user.id} accessed /student_download_report_card without admission_no")
                flash('Please select a linked student.', 'danger')
                return redirect(url_for('parent_dashboard'))
            parent_student = db_session.query(ParentStudent).filter_by(
                parent_id=current_user.id,
                admission_no=admission_no
            ).first()
            if not parent_student:
                logger.warning(f"Parent {current_user.id} not linked to admission_no={admission_no}")
                flash('You are not authorized to view this student’s report card.', 'danger')
                return redirect(url_for('parent_dashboard'))
            parent_view = True

        # Fetch student data
        student = db_session.query(Student).filter_by(admission_no=admission_no).first()
        if not student:
            logger.warning(f"No student found for admission_no={admission_no}")
            flash('Student profile not found.', 'danger')
            return redirect(url_for('logout' if current_user.role == 'student' else 'parent_dashboard'))

        # Fetch term info for default values
        term_data = db_session.query(TermInfo).filter_by(id=1).first()
        default_term = term_data.term if term_data and term_data.term else 'Term 1'
        default_year = term_data.year if term_data and term_data.year else str(datetime.now().year)
        default_exam_type = 'endterm'

        # Fetch marks for display in the template
        marks = db_session.query(Marks).filter_by(admission_no=admission_no).order_by(Marks.year.desc(), Marks.term.desc()).all()
        marks_data = [{
            'learning_area': mark.learning_area,
            'marks': mark.total_marks,
            'exam_type': mark.exam_type,
            'total_marks': mark.total_marks,
            'term': mark.term,
            'year': mark.year,
            'grade': mark.grade
        } for mark in marks]

        # Initialize form
        report_form = ReportCardForm(
            admission_no=admission_no,
            grade=student.grade,
            term=default_term,
            year=int(default_year) if default_year.isdigit() else datetime.now().year,
            exam_type=default_exam_type
        )

        # Set form choices
        report_form.admission_no.choices = [(student.admission_no, f"{student.name} ({student.admission_no})")]
        report_form.grade.choices = [(g[0], g[1]) for g in GRADES if g[0] != 'all']  # Exclude 'all' for students/parents
        report_form.term.choices = TERMS
        report_form.exam_type.choices = EXAM_TYPES

        # Handle form submission
        if request.method == 'POST' and report_form.validate_on_submit():
            admission_no = report_form.admission_no.data
            grade = report_form.grade.data
            term = report_form.term.data
            year = str(report_form.year.data)
            exam_type = report_form.exam_type.data

            logger.debug(f"Form submitted: admission_no={admission_no}, grade={grade}, term={term}, year={year}, exam_type={exam_type}")

            # Generate report card
            pdf_buffer = generate_individual_report_card(admission_no, term, year, exam_type)
            if not pdf_buffer:
                logger.error(f"Failed to generate report card for admission_no={admission_no}")
                flash('Unable to generate report card. Please try again or contact support.', 'danger')
                return redirect(url_for('student_dashboard'))

            # Prepare response
            filename = f"Report_Card_{admission_no}_{grade.replace(' ', '_')}_{term}_{year}_{exam_type}.pdf"
            response = Response(
                pdf_buffer.getvalue(),
                mimetype='application/pdf',
                headers={
                    'Content-Disposition': f'attachment; filename={filename}',
                    'Content-Length': len(pdf_buffer.getvalue())
                }
            )
            pdf_buffer.close()
            logger.info(f"Report card downloaded for admission_no={admission_no}, filename={filename}")
            return response
        elif request.method == 'POST':
            # Log form errors
            for field, errors in report_form.errors.items():
                for error in errors:
                    logger.error(f"ReportCardForm error in {field}: {error}, submitted value: {request.form.get(field, 'None')}")
                    flash(f"Form error: {error}", 'danger')

        # Fetch recent announcements for display
        recent_announcements = db_session.query(Announcements).order_by(Announcements.date.desc()).limit(5).all()
        recent_announcements = [{'content': a.content, 'date': a.date} for a in recent_announcements]

        # Render form template for GET request
        student_data = {
            'admission_no': student.admission_no,
            'name': student.name,
            'grade': student.grade
        }
        return render_template(
            'student_download_report_card.html',
            form=report_form,
            student=student_data,
            marks=marks_data,
            recent_announcements=recent_announcements,
            parent_view=parent_view
        )

    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in student_download_report_card for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        flash('Database error: Unable to process request. Please try again later.', 'danger')
        return redirect(url_for('student_dashboard'))
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in student_download_report_card for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        flash('An unexpected error occurred. Please try again later.', 'danger')
        return redirect(url_for('student_dashboard'))
    finally:
        db_session.close()
        
        
        
def get_learning_areas(grade, db_session=None):
    try:
        if not isinstance(grade, str) or not grade.strip():
            logger.warning(f"Invalid grade for learning areas: {grade}")
            return [
                'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
                'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
            ]
        grade = grade.strip()
        db_session = db_session or next(get_db())
        try:
            learning_areas = db_session.query(Marks.learning_area).filter(
                func.lower(Marks.grade) == grade.lower()
            ).distinct().order_by(Marks.learning_area).all()
            areas = [la[0].strip() for la in learning_areas if la[0] and isinstance(la[0], str)]
            if not areas:
                logger.warning(f"No learning areas found in Marks for grade {grade}. Using default list.")
                areas = [
                    'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
                    'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
                ]
            logger.debug(f"Learning areas for grade {grade}: {areas}")
            return areas
        finally:
            if db_session and not hasattr(db_session, 'is_closed') or not db_session.is_closed:
                db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching learning areas for grade {grade}: {str(e)}")
        return [
            'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
            'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
        ]


def generate_excel_results(students, marks, fees, payment_history, grade, term=None, year=None, exam_type=None):
    """Generate Excel file with student results, fees, and payment history."""
    try:
        logger.debug(f"Generating Excel results for grade={grade}, term={term}, year={year}, exam_type={exam_type}")
        wb = Workbook()
        if wb['Sheet']:
            wb.remove(wb['Sheet'])

        # Fetch learning areas dynamically
        learning_areas = get_learning_areas(grade)
        if not learning_areas:
            logger.warning(f"No learning areas found for grade {grade}. Using default list.")
            learning_areas = [
                'Maths', 'English', 'Kiswahili', 'Integrated', 'Pre-Tech',
                'Social Studies', 'Agriculture', 'Creative Arts', 'CRE'
            ]

        # Marks Sheet
        ws_marks = wb.create_sheet(title=f"Results_{grade.replace(' ', '_')}")
        headers = ['Rank', 'Admission No', 'Name'] + learning_areas + ['Total']
        ws_marks.append(headers)

        processed_students = set()
        rank_data = {}
        if marks:
            db_session = next(get_db())
            try:
                rank_query = db_session.query(
                    Marks.admission_no,
                    func.sum(Marks.total_marks).label('total')
                ).filter(
                    Marks.grade.ilike(grade) if grade != 'all' else True,
                    Marks.term.ilike(term) if term else True,
                    Marks.year == year if year else True,
                    Marks.exam_type.ilike(exam_type) if exam_type else True,
                    Marks.learning_area.in_(learning_areas)
                ).group_by(Marks.admission_no).order_by(func.sum(Marks.total_marks).desc()).all()
                rank_data = {r[0]: i + 1 for i, r in enumerate(rank_query)}
            finally:
                db_session.close()

        for student in students:
            admission_no = str(student[3]).strip()
            name = str(student[1]).strip()
            student_grade = str(student[2]).strip()

            if grade != 'all' and student_grade.lower().strip() != grade.lower().strip():
                continue
            if admission_no in processed_students:
                continue
            processed_students.add(admission_no)

            student_marks = {
                m[1]: m[2] for m in marks
                if str(m[0]).strip() == admission_no and m[1] in learning_areas
            }
            row = [rank_data.get(admission_no, 'N/A'), admission_no, name]
            total_marks = 0
            for la in learning_areas:
                marks_value = student_marks.get(la, None)
                marks_display = int(marks_value) if marks_value is not None else ''
                row.append(marks_display)
                if marks_value is not None:
                    total_marks += float(marks_value)
            row.append(int(total_marks) if total_marks > 0 else '')
            ws_marks.append(row)

        if not processed_students:
            ws_marks.append(['No students found for this grade.'])

        # Fees Sheet
        ws_fees = wb.create_sheet(title="Fees")
        ws_fees.append(['Admission No', 'Name', 'Total Fee', 'Amount Paid', 'Balance', 'Grade', 'Term', 'Year'])
        for fee in fees:
            ws_fees.append([
                fee[0],
                next((s[1] for s in students if s[3] == fee[0]), 'Unknown'),
                float(fee[1]) if fee[1] is not None else 0.0,
                float(fee[2]) if fee[2] is not None else 0.0,
                float(fee[3]) if fee[3] is not None else 0.0,
                fee[4],
                fee[5],
                fee[6]
            ])
        if not fees:
            ws_fees.append(['No fees found for this selection.'])

        # Payment History Sheet
        ws_payments = wb.create_sheet(title="Payment History")
        ws_payments.append(['Admission No', 'Amount', 'Term', 'Year'])
        for payment in payment_history:
            ws_payments.append([
                payment[0],
                float(payment[1]) if payment[1] is not None else 0.0,
                payment[2],
                payment[3]
            ])
        if not payment_history:
            ws_payments.append(['No payment history found for this selection.'])

        # Apply styling
        header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for ws in [ws_marks, ws_fees, ws_payments]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.column > 3 and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error generating Excel results: {str(e)}\n{traceback.format_exc()}")
        return None



@app.route('/download_results', methods=['GET', 'POST'])
@login_required
@admin_required
def download_results():
    """Download student results as an Excel file for a grade."""
    form = ReportCardForm()
    term_info, content_data = fetch_common_data()
    db_session = next(get_db())
    try:
        if form.validate_on_submit():
            grade = str(form.grade.data).strip()
            term = str(form.term.data).strip()
            year = str(form.year.data).strip()
            exam_type = str(form.exam_type.data).lower().strip()

            # Fetch students
            students = db_session.query(Student.id, Student.name, Student.grade, Student.admission_no).filter(
                Student.grade.ilike(grade) if grade != 'all' else True
            ).order_by(Student.admission_no).all()
            if not students:
                flash(f"No students found for {grade}.", 'danger')
                return render_template('download_results.html', form=form, term_info=term_info, content_data=content_data)

            # Fetch learning areas
            learning_areas = get_learning_areas(grade)
            if not learning_areas:
                flash(f"No learning areas defined for {grade}. Please add learning areas.", 'danger')
                return render_template('download_results.html', form=form, term_info=term_info, content_data=content_data)

            # Fetch marks
            marks = db_session.query(
                Marks.admission_no, Marks.learning_area, Marks.total_marks, Marks.term, Marks.year, Marks.grade, Marks.exam_type
            ).filter(
                Marks.grade.ilike(grade) if grade != 'all' else True,
                Marks.term.ilike(term) if term else True,
                Marks.year == year if year else True,
                Marks.exam_type.ilike(exam_type) if exam_type else True,
                Marks.learning_area.in_(learning_areas)
            ).all()

            # Fetch fees
            fees = db_session.query(
                Fee.admission_no, Fee.total_fee, Fee.amount_paid, Fee.balance, Fee.grade, Fee.term, Fee.year
            ).filter(
                Fee.grade.ilike(grade) if grade != 'all' else True,
                Fee.term.ilike(term) if term else True,
                Fee.year == year if year else True
            ).all()

            # Fetch payment history (without payment_date)
            payment_history = db_session.query(
                Fee.admission_no, Fee.amount_paid, Fee.term, Fee.year
            ).filter(
                Fee.grade.ilike(grade) if grade != 'all' else True,
                Fee.term.ilike(term) if term else True,
                Fee.year == year if year else True,
                Fee.amount_paid > 0
            ).all()

            # Generate Excel file
            excel_buffer = generate_excel_results(students, marks, fees, payment_history, grade, term, year, exam_type)
            if not excel_buffer or len(excel_buffer.getvalue()) == 0:
                flash(f"Error generating Excel file for {grade}.", 'danger')
                return render_template('download_results.html', form=form, term_info=term_info, content_data=content_data)

            filename = f'Results_{grade.replace(" ", "_")}_{term}_{year}_{exam_type}.xlsx'
            return send_file(
                excel_buffer,
                download_name=filename,
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        return render_template('download_results.html', form=form, term_info=term_info, content_data=content_data)
    except Exception as e:
        logger.error(f"Error in download_results: {str(e)}\n{traceback.format_exc()}")
        flash(f"Error: {str(e)}", 'danger')
        return render_template('download_results.html', form=form, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()

# Helper functions for report card enhancements
def get_performance_levels(marks, type_='learning_area'):
    """Determine performance level based on marks and type."""
    try:
        if not isinstance(marks, (int, float)) or marks < 0 or marks > 100:
            logger.warning(f"Invalid marks for performance level: {marks}, type={type_}")
            return 'N/A'
        marks = float(marks)
        if type_ in ('learning_area', 'class_teacher', 'principal'):
            if marks >= 90:
                return 'EE1'
            elif marks >= 80:
                return 'EE2'
            elif marks >= 70:
                return 'ME1'
            elif marks >= 60:
                return 'ME2'
            elif marks >= 50:
                return 'AE1'
            else:
                return 'AE2'
        logger.warning(f"Invalid type for performance level: {type_}")
        return 'N/A'
    except (ValueError, TypeError) as e:
        logger.error(f"Error in get_performance_levels: marks={marks}, type={type_}, error={str(e)}")
        return 'N/A'

def get_points(marks, type_='learning_area'):
    """Calculate points based on marks."""
    try:
        if not isinstance(marks, (int, float)) or marks < 0 or marks > 100:
            logger.warning(f"Invalid marks for points calculation: {marks}, type={type_}")
            return 0.0
        marks = float(marks)
        if type_ in ('learning_area', 'class_teacher', 'principal'):
            if marks >= 90:
                return 4.0  # EE1
            elif marks >= 80:
                return 3.5  # EE2
            elif marks >= 70:
                return 3.0  # ME1
            elif marks >= 60:
                return 2.5  # ME2
            elif marks >= 50:
                return 2.0  # AE1
            else:
                return 1.0  # AE2
        logger.warning(f"Invalid type for points calculation: {type_}")
        return 0.0
    except (ValueError, TypeError) as e:
        logger.error(f"Error in get_points: marks={marks}, type={type_}, error={str(e)}")
        return 0.0

def get_teacher_comment(level):
    """Return teacher comment based on performance levels."""
    comments = {
        'EE1': 'Outstanding performance! Keep it up.',
        'EE2': 'Excellent work, consistently strong.',
        'ME1': 'Good effort, meets expectations.',
        'ME2': 'Satisfactory, room for improvement.',
        'AE1': 'Approaching expectations, needs more focus.',
        'AE2': 'Below expectations, seek extra support.',
        'N/A': 'No performance data available.'
    }
    comment = comments.get(level, 'No performance data available.')
    logger.debug(f"Teacher comment for level {level}: {comment}")
    return comment

def get_class_teacher_comment(total_marks, grade):
    """Generate class teacher comment based on total marks."""
    try:
        if not isinstance(grade, str) or not grade.strip() or not isinstance(total_marks, (int, float)) or total_marks <= 0:
            logger.warning(f"Invalid inputs for class teacher comment: total_marks={total_marks}, grade={grade}")
            return 'No performance data available.'
        level = get_performance_levels(total_marks, 'class_teacher')
        comment = get_teacher_comment(level)
        logger.debug(f"Class teacher comment for total_marks={total_marks}, grade={grade}: {comment}")
        return comment
    except Exception as e:
        logger.error(f"Error fetching class teacher comment: total_marks={total_marks}, grade={grade}, error={str(e)}")
        return 'No performance data available.'

def get_principal_comment(total_marks):
    """Generate principal comment based on total marks."""
    try:
        if not isinstance(total_marks, (int, float)) or total_marks <= 0:
            logger.warning(f"Invalid total_marks for principal comment: {total_marks}")
            return 'No performance data available.'
        level = get_performance_levels(total_marks, 'principal')
        comment = get_teacher_comment(level)
        logger.debug(f"Principal comment for total_marks={total_marks}: {comment}")
        return comment
    except Exception as e:
        logger.error(f"Error fetching principal comment: total_marks={total_marks}, error={str(e)}")
        return 'No performance data available.'

def get_teacher_name(learning_area, grade):
    """Fetch teacher name for a learning area and grade."""
    try:
        if not isinstance(learning_area, str) or not isinstance(grade, str) or not learning_area.strip() or not grade.strip():
            logger.warning(f"Invalid input: learning_area={learning_area}, grade={grade}")
            return 'Unknown Teacher'
        learning_area = learning_area.strip()
        grade = grade.strip()
        db_session = next(get_db())
        try:
            teacher = db_session.query(User.username).join(
                TeacherAssignments, User.id == TeacherAssignments.teacher_id
            ).filter(
                func.lower(TeacherAssignments.learning_area) == learning_area.lower(),
                func.lower(TeacherAssignments.grade) == grade.lower()
            ).first()
            teacher_name = teacher[0].strip() if teacher and teacher[0] else 'Unknown Teacher'
            logger.debug(f"Teacher for {learning_area}, {grade}: {teacher_name}")
            return teacher_name
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching teacher for {learning_area}, {grade}: {str(e)}")
        return 'Unknown Teacher'

def get_class_teacher_name(grade):
    """Fetch class teacher name for a grade."""
    try:
        if not isinstance(grade, str) or not grade.strip():
            logger.warning(f"Invalid grade: {grade}")
            return 'Unknown Class Teacher'
        grade = grade.strip()
        db_session = next(get_db())
        try:
            teacher = db_session.query(User.username).join(
                ClassTeachers, User.id == ClassTeachers.teacher_id
            ).filter(
                func.lower(ClassTeachers.grade) == grade.lower()
            ).first()
            teacher_name = teacher[0].strip() if teacher and teacher[0] else 'Unknown Class Teacher'
            logger.debug(f"Class teacher for grade {grade}: {teacher_name}")
            return teacher_name
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching class teacher for {grade}: {str(e)}")
        return 'Unknown Class Teacher'

def get_principal_name():
    """Fetch principal name from TermInfo."""
    try:
        db_session = next(get_db())
        try:
            principal = db_session.query(TermInfo.principal).filter_by(id=1).first()
            principal_name = principal[0].strip() if principal and principal[0] else 'School Principal'
            logger.debug(f"Principal name: {principal_name}")
            return principal_name
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching principal name: {str(e)}")
        return 'School Principal'

def get_rank(admission_no, grade, term, year, exam_type):
    """Calculate student rank based on total marks."""
    try:
        if not all(isinstance(x, str) for x in [admission_no, grade, term, exam_type, year]):
            logger.warning(f"Invalid inputs for rank: admission_no={admission_no}, grade={grade}, term={term}, year={year}, exam_type={exam_type}")
            return 'N/A'
        admission_no = admission_no.strip().lower()
        grade = grade.strip().lower()
        term = term.strip().lower()
        exam_type = exam_type.strip().lower()
        year = year.strip()
        db_session = next(get_db())
        try:
            ranks = db_session.query(
                Marks.admission_no,
                func.sum(Marks.total_marks).label('total')
            ).filter(
                func.lower(Marks.grade) == grade,
                func.lower(Marks.term) == term,
                Marks.year == year,
                func.lower(Marks.exam_type) == exam_type,
                Marks.total_marks.isnot(None),
                Marks.total_marks >= 0,
                Marks.total_marks <= 100
            ).group_by(Marks.admission_no).order_by(func.sum(Marks.total_marks).desc()).all()
            rank_list = [r[0].strip().lower() for r in rank]
            rank = rank_list.index(admission_no) + 1 if admission_no in rank_list else 'N/A'
            logger.debug(f"Rank for {admission_no} in {grade}, {term}, {year}, {exam_type}: {rank}")
            return str(rank)
        finally:
            db_session.close()
    except (SQLAlchemyError, ValueError) as e:
        logger.error(f"Error calculating rank for {admission_no}: {str(e)}")
        return 'N/A'

def get_total_fee(admission_no):
    """Fetch total fee for a student for the latest term."""
    try:
        if not isinstance(admission_no, str) or not admission_no.strip():
            logger.warning(f"Invalid admission_no: {admission_no}")
            return 0.0
        admission_no = admission_no.strip().lower()
        db_session = next(get_db())
        try:
            fee = db_session.query(Fee.total_fee).filter(
                func.lower(Fee.admission_no) == admission_no
            ).order_by(Fee.term.desc()).first()
            total_fee = float(fee[0]) if fee and fee[0] is not None else 0.0
            logger.debug(f"Total fee for {admission_no}: {total_fee}")
            return total_fee
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching total fee for {admission_no}: {str(e)}")
        return 0.0

def get_amount_paid(admission_no):
    """Fetch amount paid for a student for the latest term."""
    try:
        if not isinstance(admission_no, str) or not admission_no.strip():
            logger.warning(f"Invalid admission_no: {admission_no}")
            return 0.0
        admission_no = admission_no.strip().lower()
        db_session = next(get_db())
        try:
            paid = db_session.query(Fee.amount_paid).filter(
                func.lower(Fee.admission_no) == admission_no
            ).order_by(Fee.term.desc()).first()
            amount_paid = float(paid[0]) if paid and paid[0] is not None else 0.0
            logger.debug(f"Amount paid for {admission_no}: {amount_paid}")
            return amount_paid
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching amount paid for {admission_no}: {str(e)}")
        return 0.0

def get_balance(admission_no):
    """Fetch fee balance for a student for the latest term."""
    try:
        if not isinstance(admission_no, str) or not admission_no.strip():
            logger.warning(f"Invalid admission_no: {admission_no}")
            return 0.0
        admission_no = admission_no.strip().lower()
        db_session = next(get_db())
        try:
            balance = db_session.query(Fee.balance).filter(
                func.lower(Fee.admission_no) == admission_no
            ).order_by(Fee.term.desc()).first()
            fee_balance = float(balance[0]) if balance and balance[0] is not None else 0.0
            logger.debug(f"Balance for {admission_no}: {fee_balance}")
            return fee_balance
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching balance for {admission_no}: {str(e)}")
        return 0.0



def generate_report_card(students, marks, fees, term, year, exam_type, rank=None, total_students=None, grade=None, term_info=None):
    """Generate a PDF report card for multiple students."""
    try:
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        processed_students = set()
        has_valid_content = False

        if not isinstance(students, (list, tuple)):
            logger.error(f"Invalid students type: {type(students)}")
            raise ValueError("Students must be a list or tuple")
        if not isinstance(marks, (list, tuple)):
            logger.error(f"Invalid marks type: {type(marks)}")
            raise ValueError("Marks must be a list or tuple")
        if not isinstance(fees, (list, tuple)):
            logger.error(f"Invalid fees type: {type(fees)}")
            raise ValueError("Fees must be a list or tuple")
        if not all(isinstance(x, str) for x in [term, year, exam_type] if x is not None):
            logger.error(f"Invalid string inputs: term={type(term)}, year={type(year)}, exam_type={type(exam_type)}")
            raise ValueError("Term, year, and exam_type must be strings")

        term = term.strip().lower()
        year = year.strip()
        exam_type = exam_type.strip().lower().replace('cat 1', 'cat1').replace('cat 2', 'cat2').replace('cat 3', 'cat3') \
            .replace('rat 1', 'rat1').replace('rat 2', 'rat2').replace('rat 3', 'rat3') \
            .replace('mid term', 'midterm').replace('end term', 'endterm') \
            .replace('project 1', 'project1').replace('project 2', 'project2').replace('project 3', 'project3')
        grade = grade.strip().lower() if grade else None

        term_info = term_info or {
            'term': term.capitalize(),
            'year': year,
            'principal': get_principal_name(),
            'start_date': '2025-01-01',
            'end_date': '2025-04-01'
        }

        # Fetch learning areas
        learning_areas = get_learning_areas(grade) if grade else [
            'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
            'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
        ]
        logger.debug(f"Learning areas for grade {grade}: {learning_areas}")

        logger.debug(f"Processing students: {len(students)} students, {len(marks)} marks, term={term}, year={year}, exam_type={exam_type}, grade={grade}")

        for student in students:
            try:
                if isinstance(student, (list, tuple)) and len(student) >= 6:
                    admission_no = str(student[5]).strip().lower()
                    name = str(student[1]).strip()
                    student_grade = str(student[4]).strip().lower()
                elif isinstance(student, dict):
                    admission_no = str(student.get('admission_no', '')).strip().lower()
                    name = str(student.get('name', '')).strip()
                    student_grade = str(student.get('grade', '')).strip().lower()
                else:
                    logger.warning(f"Invalid student data: {student}")
                    continue

                if not all([admission_no, name, student_grade]):
                    logger.warning(f"Empty or invalid student data: admission_no={admission_no}, name={name}, grade={student_grade}")
                    continue
                if grade and student_grade != grade:
                    logger.debug(f"Skipping student {admission_no} with grade {student_grade} (expected {grade})")
                    continue
                if admission_no in processed_students:
                    logger.debug(f"Skipping duplicate admission_no: {admission_no}")
                    continue

                student_marks = [
                    m for m in marks
                    if (isinstance(m, dict) and str(m.get('admission_no', '')).strip().lower() == admission_no) or
                       (isinstance(m, (list, tuple)) and len(m) >= 3 and str(m[0]).strip().lower() == admission_no)
                ]
                logger.debug(f"Marks for {admission_no}: {len(student_marks)} entries: {[m[:3] for m in student_marks]}")

                processed_students.add(admission_no)
                has_valid_content = True

                # Header
                c.setFont("Helvetica-Bold", 16)
                c.drawCentredString(300, 750, "JONYO JUNIOR SECONDARY SCHOOL")
                c.setFont("Helvetica", 14)
                c.drawCentredString(300, 730, "REPORT CARD")
                c.setFont("Helvetica", 12)
                c.drawString(50, 700, f"Name: {name}")
                c.drawString(50, 680, f"Admission No: {admission_no.upper()}")
                c.drawString(50, 660, f"Grade: {student_grade.capitalize()}")
                formatted_exam_type = exam_type.replace('cat1', 'CAT 1').replace('cat2', 'CAT 2').replace('cat3', 'CAT 3') \
                    .replace('rat1', 'RAT 1').replace('rat2', 'RAT 2').replace('rat3', 'RAT 3') \
                    .replace('midterm', 'Mid Term').replace('endterm', 'End Term') \
                    .replace('project1', 'Project 1').replace('project2', 'Project 2').replace('project3', 'Project 3')
                c.drawString(50, 640, f"Term: {term_info['term']} {year} ({formatted_exam_type})")
                c.drawString(50, 620, f"School Year: {term_info['year']}")

                # Create table data
                table_data = [['Learning Area', 'Marks', 'Perf. Level', 'Points', 'Teacher Comment', 'Teacher']]
                total_marks = 0
                total_points = 0
                valid_subjects = 0
                subjects_covered = set()

                for learning_area in learning_areas:
                    mark_found = False
                    learning_area_normalized = learning_area.lower().strip()
                    for mark in student_marks:
                        mark_learning_area = str(mark[1] if isinstance(mark, (list, tuple)) else mark.get('learning_area', '')).strip().lower()
                        if mark_learning_area == learning_area_normalized and mark_learning_area not in subjects_covered:
                            try:
                                marks_value = float(mark[2] if isinstance(mark, (list, tuple)) else mark.get('total_marks', 0)) if (mark[2] if isinstance(mark, (list, tuple)) else mark.get('total_marks')) is not None else None
                                if marks_value is not None and 0 <= marks_value <= 100:
                                    total_marks += marks_value
                                    points = get_points(marks_value, 'learning_area')
                                    total_points += points
                                    valid_subjects += 1
                                    level = get_performance_levels(marks_value, 'learning_area')
                                    table_data.append([
                                        learning_area,
                                        f"{int(marks_value)}",
                                        level,
                                        f"{points:.2f}",
                                        get_teacher_comment(level)[:20],
                                        get_teacher_name(learning_area, student_grade)[:20]
                                    ])
                                    subjects_covered.add(mark_learning_area)
                                    mark_found = True
                                    logger.debug(f"Processed mark for {admission_no}, {learning_area}: {marks_value}, Points: {points}, Level: {level}")
                                else:
                                    logger.warning(f"Invalid marks value {marks_value} for {admission_no}, {learning_area}")
                            except (ValueError, TypeError) as e:
                                logger.error(f"Error processing mark for {admission_no}, {learning_area}: {str(e)}")
                    if not mark_found:
                        logger.debug(f"No marks found for {admission_no} in {learning_area}")
                        table_data.append([learning_area, '-', 'N/A', '0.00', 'No marks available', 'Unknown Teacher'])

                # Calculate average marks and points
                average_marks = total_marks / valid_subjects if valid_subjects > 0 else 0
                average_points = total_points / valid_subjects if valid_subjects > 0 else 0

                # Create table
                col_widths = [2.0*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1.5*inch, 1.2*inch]
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                table.wrapOn(c, 500, 400)
                table_height = table._height
                table_x = 30
                table_y = 600 - table_height
                if table_y < 100:
                    logger.warning(f"Table y-position {table_y} too low for {admission_no}, splitting table")
                    rows_per_page = max(5, int((600 - table_y) / 20))
                    for i in range(0, len(table_data), rows_per_page):
                        sub_table = Table(table_data[:1] + table_data[i+1:i+rows_per_page+1], colWidths=col_widths)
                        sub_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 1), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('LEFTPADDING', (0, 0), (-1, -1), 4),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                            ('TOPPADING', (0, 0), (-1, -1), 4),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ]))
                        sub_table.wrapOn(c, 500, 400)
                        sub_table.drawOn(c, table_x, 600 - sub_table._height)
                        c.showPage()
                        c.setFont("Helvetica-Bold", 16)
                        c.drawCentredString(300, 750, "JONYO JUNIOR SECONDARY SCHOOL")
                        c.setFont("Helvetica", 14)
                        c.drawCentredString(300, 730, "REPORT CARD")
                        c.setFont("Helvetica", 12)
                        c.drawString(50, 700, f"Name: {name}")
                        c.drawString(50, 680, f"Admission No: {admission_no.upper()}")
                        c.drawString(50, 660, f"Grade: {student_grade.capitalize()}")
                        c.drawString(50, 640, f"Term: {term_info['term']} {year} ({formatted_exam_type})")
                        c.drawString(50, 620, f"School Year: {term_info['year']}")
                    table = None
                else:
                    table.drawOn(c, table_x, table_y)

                # Footer details
                y = max(table_y - 20, 100) if table else 580
                student_fees = [
                    f for f in fees
                    if (isinstance(f, dict) and str(f.get('admission_no', '')).strip().lower() == admission_no) or
                       (isinstance(f, (list, tuple)) and len(f) >= 7 and str(f[0]).strip().lower() == admission_no)
                ]
                fee_info = student_fees[0] if student_fees else (admission_no, 0, 0, 0, grade, term, year)

                total_fee = float(fee_info[1] if isinstance(fee_info, (list, tuple)) else fee_info.get('total_fee', 0) or 0)
                balance = float(fee_info[3] if isinstance(fee_info, (list, tuple)) else fee_info.get('balance', 0) or 0)

                student_rank = rank.get(admission_no, 'N/A') if isinstance(rank, dict) else rank if rank else 'N/A'

                # Use average marks for performance level and comments
                performance_level = get_performance_levels(average_marks, 'class_teacher') if valid_subjects > 0 else 'N/A'
                class_teacher_comment = get_class_teacher_comment(average_marks, student_grade) if valid_subjects > 0 else 'No performance data available'
                principal_comment = get_principal_comment(average_marks) if valid_subjects > 0 else 'No performance data available'

                c.setFont("Helvetica", 10)
                c.drawString(30, y-20, f"Rank: {student_rank} out of {total_students if total_students else 'N/A'}")
                c.drawString(30, y-40, f"Total Marks: {int(total_marks)}" if total_marks > 0 else "Total Marks: N/A")
                c.drawString(30, y-60, f"Total Points: {total_points:.2f}" if total_points > 0 else "Total Points: N/A")
                c.drawString(30, y-80, f"Average Points: {average_points:.2f}" if valid_subjects > 0 else "Average Points: N/A")
                c.drawString(30, y-100, f"Performance Level: {performance_level}")
                c.drawString(30, y-120, f"Total Fee: {total_fee:,.2f}")
                c.drawString(30, y-140, f"Balance: {balance:,.2f}")
                c.drawString(30, y-160, f"Class Teacher Comment: {class_teacher_comment[:50]}")
                c.drawString(30, y-180, f"Principal Comment: {principal_comment[:50]}")
                c.drawString(30, y-200, f"Class Teacher: {get_class_teacher_name(student_grade)[:30]}")
                c.drawString(30, y-220, f"Principal: {get_principal_name()[:30]}")
                c.drawString(30, y-240, f"Start Date: {term_info['start_date']}")
                c.drawString(30, y-260, f"End Date: {term_info['end_date']}")
                c.drawString(30, y-280, "School Stamp: ____________________")

                # Watermark
                c.setFont("Helvetica", 50)
                c.setFillColor(colors.grey, alpha=0.2)
                c.rotate(45)
                c.drawCentredString(400, 200, "JONYO JSS")
                c.rotate(-45)
                c.setFillColor(colors.black)
                c.showPage()
            except Exception as e:
                logger.error(f"Error processing student {admission_no}: {str(e)}\n{traceback.format_exc()}")
                continue

        if not has_valid_content:
            logger.warning(f"No valid report cards generated for {grade}, {term}, {year}, {exam_type}")
            c.setFont("Helvetica", 12)
            c.drawString(50, 700, "No valid student data or marks available. Please contact your teacher.")
            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        c.save()
        buffer.seek(0)
        pdf_content = buffer.getvalue()
        if not pdf_content.startswith(b'%PDF-'):
            logger.error(f"Invalid PDF generated for {grade}, starts with: {pdf_content[:10]}")
            buffer.close()
            return None
        logger.debug(f"PDF report card generated for {grade}, size: {len(pdf_content)} bytes")
        return buffer
    except Exception as e:
        logger.error(f"Error in generate_report_card: {str(e)}\n{traceback.format_exc()}")
        if 'buffer' in locals():
            buffer.close()
        return None

import uuid

def create_zipped_report_cards(students, marks, fees, term, year, exam_type, rank, total_students, grade):
    """Generate a ZIP file containing report card PDFs for each student."""
    logger.debug(f"Creating zipped report cards for grade: {grade}, term: {term}, year: {year}, exam_type: {exam_type}")

    # Ensure all string inputs are strings and validate
    try:
        term = str(term).strip()
        exam_type = str(exam_type).strip()
        grade = str(grade).strip()
        year = str(year).strip()
    except Exception as e:
        logger.error(f"Failed to convert inputs to strings: term={term}, year={year}, exam_type={exam_type}, error={str(e)}")
        return None

    if not all(isinstance(x, str) for x in [term, year, exam_type, grade]):
        logger.error(f"Invalid input types. term={type(term)}, year={type(year)}, exam_type={type(exam_type)}, grade={type(grade)}")
        return None

    # Validate year
    try:
        year_int = int(year)
        if not (2000 <= year_int <= 2100):
            logger.error(f"Year {year} is out of valid range (2000-2100)")
            return None
    except ValueError:
        logger.error(f"Invalid year format: {year}")
        return None

    term_info = {
        'term': term.capitalize(),
        'year': year,
        'principal': get_principal_name(),
        'start_date': '2025-01-01',
        'end_date': '2025-04-01'
    }

    zip_buffer = io.BytesIO()
    report_cards_added = 0

    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zip_file:
            for student in students:
                try:
                    # Extract student information
                    if isinstance(student, (list, tuple)) and len(student) >= 6:
                        admission_no = str(student[5]).strip().lower()
                        name = str(student[1]).strip()
                        student_grade = str(student[4]).strip().lower()
                    elif isinstance(student, dict):
                        admission_no = str(student.get('admission_no', '')).strip().lower()
                        name = str(student.get('name', '')).strip()
                        student_grade = str(student.get('grade', '')).strip().lower()
                    else:
                        logger.warning(f"Invalid student format: {student}")
                        continue

                    # Validate student data
                    if not admission_no or not name or not student_grade:
                        logger.warning(f"Incomplete student data: admission_no={admission_no}, name={name}, grade={student_grade}")
                        continue

                    if student_grade != grade.lower():
                        logger.debug(f"Skipping {admission_no}: grade mismatch ({student_grade} != {grade})")
                        continue

                    # Filter student marks
                    student_marks = [
                        m for m in marks
                        if (isinstance(m, dict) and str(m.get('admission_no', '')).strip().lower() == admission_no) or
                           (isinstance(m, (list, tuple)) and len(m) >= 3 and str(m[0]).strip().lower() == admission_no)
                    ]

                    # Filter student fees
                    student_fees = [
                        f for f in fees
                        if (isinstance(f, dict) and str(f.get('admission_no', '')).strip().lower() == admission_no) or
                           (isinstance(f, (list, tuple)) and len(f) >= 7 and str(f[0]).strip().lower() == admission_no)
                    ] or [(admission_no, 0, 0, 0, grade, term, year)]

                    if not student_marks:
                        logger.warning(f"No marks found for {admission_no} ({name})")
                        continue

                    logger.debug(f"Generating report for {admission_no}: marks={len(student_marks)}, fees={len(student_fees)}")

                    # Generate PDF report card for a single student
                    pdf_buffer = generate_report_card(
                        students=[student],  # Pass single student
                        marks=student_marks,
                        fees=student_fees,
                        term=term,
                        year=year,
                        exam_type=exam_type,
                        rank=rank.get(admission_no.lower(), 'N/A') if isinstance(rank, dict) else rank,
                        total_students=total_students,
                        grade=grade,
                        term_info=term_info
                    )

                    if not pdf_buffer or pdf_buffer.getvalue() == b'':
                        logger.warning(f"generate_report_card returned None or empty for {admission_no}")
                        continue

                    # Verify PDF content
                    pdf_buffer.seek(0)
                    pdf_content = pdf_buffer.getvalue()
                    if not pdf_content.startswith(b'%PDF-'):
                        logger.warning(f"Corrupt or invalid PDF for {admission_no}. Starts with: {pdf_content[:10]}")
                        pdf_buffer.close()
                        continue

                    # Generate unique filename
                    safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '_')).replace(' ', '_')
                    filename = f'Report_Card_{safe_name}_{admission_no}_{grade.replace(" ", "_")}_{term}_{year}_{exam_type}_{uuid.uuid4().hex[:8]}.pdf'
                    zip_file.writestr(filename, pdf_content)
                    report_cards_added += 1
                    logger.debug(f"Added {filename}, size={len(pdf_content)} bytes")
                    pdf_buffer.close()

                except Exception as e:
                    logger.error(f"Error processing student {admission_no}: {str(e)}\n{traceback.format_exc()}")
                    continue

            # If no report cards were added, include an error message
            if report_cards_added == 0:
                missing_students = [
                    s[5] if isinstance(s, (list, tuple)) else s.get('admission_no', '')
                    for s in students
                    if (s[5] if isinstance(s, (list, tuple)) else s.get('admission_no', '')) not in
                       [m[0] if isinstance(m, (list, tuple)) else m.get('admission_no', '') for m in marks]
                ]
                error_message = (
                    f"No report cards generated for {grade}, {term}, {year}, {exam_type}.\n"
                    f"Missing marks for: {', '.join(map(str, missing_students))}.\n"
                    "Please ensure all required data is present in the system."
                )
                zip_file.writestr("no_report_cards.txt", error_message.encode('utf-8'))
                report_cards_added += 1
                logger.warning(error_message)

        # Ensure buffer is properly positioned
        zip_buffer.seek(0)
        zip_content = zip_buffer.getvalue()
        if not zip_content:
            logger.error("ZIP file is empty")
            return None

        # Verify ZIP file integrity
        try:
            with zipfile.ZipFile(zip_buffer, 'r') as test_zip:
                bad_file = test_zip.testzip()
                if bad_file:
                    logger.error(f"ZIP integrity check failed, bad file: {bad_file}")
                    return None
                logger.debug(f"ZIP integrity verified. Files: {test_zip.namelist()}")
        except zipfile.BadZipFile as e:
            logger.error(f"Bad ZIP file generated: {str(e)}\n{traceback.format_exc()}")
            return None

        logger.info(f"ZIP created with {report_cards_added} files for grade {grade}")
        return zip_buffer

    except Exception as e:
        logger.error(f"create_zipped_report_cards failed: {str(e)}\n{traceback.format_exc()}")
        if 'zip_buffer' in locals():
            zip_buffer.close()
        return None

@app.route('/download_report_card', methods=['GET', 'POST'])
@login_required
def download_report_card():
    """Download zipped report cards for a grade (admin only)."""
    if current_user.role != 'admin':
        logger.warning(f"Unauthorized access attempt by user {current_user.id} with role {current_user.role}")
        flash("Only admins can download report cards.", 'danger')
        return render_template('download_report_card.html', form=ReportCardForm(), term_info=fetch_common_data()[0])

    form = ReportCardForm()
    term_info, content_data = fetch_common_data()
    db_session = next(get_db())
    try:
        # Validate and convert term_info['year'] to integer
        try:
            default_year = int(term_info['year']) if term_info.get('year') else 2025
            if not (2000 <= default_year <= 2100):
                raise ValueError(f"Year {default_year} out of range (2000-2100)")
        except (ValueError, TypeError) as e:
            logger.error(f"Invalid term_info year: {term_info.get('year')}, error: {str(e)}")
            default_year = 2025
            flash("Invalid year in term information. Using default year 2025.", 'warning')

        # Populate form choices
        form.grade.choices = [('Grade 7', 'Grade 7'), ('Grade 8', 'Grade 8'), ('Grade 9', 'Grade 9')]
        form.term.choices = [('Term 1', 'Term 1'), ('Term 2', 'Term 2'), ('Term 3', 'Term 3')]
        form.exam_type.choices = EXAM_TYPES  # Ensure EXAM_TYPES is defined
        form.year.data = default_year

        if form.validate_on_submit():
            grade = str(form.grade.data).strip()
            term = str(form.term.data).strip()
            try:
                year = int(form.year.data)
                if not (2000 <= year <= 2100):
                    raise ValueError(f"Year {year} out of range (2000-2100)")
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid form year value: {form.year.data}, error: {str(e)}")
                flash("Invalid year format. Please enter a valid year (2000-2100).", 'danger')
                return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)
            exam_type = str(form.exam_type.data).strip().lower()

            logger.debug(f"Form submitted: role={current_user.role}, grade={grade}, term={term}, year={year}, exam_type={exam_type}")

            # Expanded variations for term and exam_type
            term_variations = [
                term.lower(), term.lower().replace('term ', 'term'), term.lower().replace(' ', ''),
                term.lower().replace('term', 'term '), f"{term.lower()} 1", term.lower().upper(),
                term.lower().capitalize()
            ]
            exam_type_variations = [
                exam_type, exam_type.upper(), exam_type.capitalize(),
                exam_type.replace('cat1', 'CAT 1').replace('cat2', 'CAT 2').replace('cat3', 'CAT 3'),
                exam_type.replace('cat', 'CAT ').replace('rat', 'RAT ').replace('midterm', 'Mid Term')
                .replace('endterm', 'End Term').replace('project1', 'Project 1')
                .replace('project2', 'Project 2').replace('project3', 'Project 3')
            ]

            # Pre-check marks availability
            learning_areas = get_learning_areas(grade, db_session)
            if not learning_areas:
                logger.warning(f"No learning areas for grade {grade}")
                flash(f"No learning areas defined for grade {grade}. Please contact the system admin.", 'danger')
                return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

            marks_check = db_session.query(Marks).filter(
                func.lower(Marks.grade) == grade.lower().strip(),
                func.lower(Marks.term).in_(term_variations),
                Marks.year == year,
                func.lower(Marks.exam_type).in_(exam_type_variations),
                func.lower(Marks.learning_area).in_([la.lower().strip() for la in learning_areas]),
                Marks.admission_no != None
            ).count()
            if marks_check == 0:
                logger.warning(f"No marks found for grade={grade}, term={term_variations}, year={year}, exam_type={exam_type_variations}")
                flash(f"No marks found for grade {grade}, term {term}, year {year}, exam type {exam_type}.", 'danger')
                return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

            # Fetch students
            students = db_session.query(
                Student.id, Student.name, Student.grade, Student.admission_no
            ).filter(
                func.lower(Student.grade) == grade.lower().strip()
            ).order_by(Student.admission_no).all()
            if not students:
                flash(f"No students found for grade {grade}.", 'warning')
                return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

            # Format students as 6-element tuples
            students_formatted = []
            for s in students:
                if not s.admission_no:
                    logger.warning(f"Skipping student with missing admission_no: {s}")
                    continue
                try:
                    student_tuple = (
                        s.id,  # student_id
                        s.name.strip() if s.name else 'Unknown',  # name
                        None,  # placeholder
                        None,  # placeholder
                        s.grade.strip() if s.grade else 'Unknown',  # grade
                        s.admission_no.strip()  # admission_no
                    )
                    students_formatted.append(student_tuple)
                except AttributeError as e:
                    logger.error(f"Error formatting student: {s}, error: {str(e)}")
                    continue
            if not students_formatted:
                flash(f"No valid students found for grade {grade}. Ensure student records are complete.", 'danger')
                return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

            logger.debug(f"Formatted {len(students_formatted)} students")

            # Fetch marks
            marks = db_session.query(
                Marks.admission_no, Marks.learning_area, Marks.marks, Marks.term, Marks.year, Marks.grade, Marks.exam_type
            ).filter(
                func.lower(Marks.grade) == grade.lower().strip(),
                func.lower(Marks.term).in_(term_variations),
                Marks.year == year,
                func.lower(Marks.exam_type).in_(exam_type_variations),
                func.lower(Marks.learning_area).in_([la.lower().strip() for la in learning_areas]),
                Marks.admission_no != None
            ).all()

            # Validate and convert marks
            valid_marks = []
            invalid_marks = []
            for m in marks:
                try:
                    marks_value = float(m.marks) if m.marks is not None else None
                    if marks_value is not None and 0 <= marks_value <= 100:
                        valid_marks.append((
                            m.admission_no.strip(),  # admission_no
                            m.learning_area.strip(),  # learning_area
                            marks_value,  # marks
                            m.term.strip(),  # term
                            m.year,  # year
                            m.grade.strip(),  # grade
                            m.exam_type.strip()  # exam_type
                        ))
                    else:
                        invalid_marks.append((m.admission_no, m.learning_area, m.marks))
                except (ValueError, TypeError):
                    invalid_marks.append((m.admission_no, m.learning_area, m.marks))
            if invalid_marks:
                logger.error(f"Invalid marks detected: {invalid_marks}")
                flash(f"Invalid marks data for some students: {', '.join([f'{m[0]} ({m[1]})' for m in invalid_marks])}. Contact the system admin.", 'warning')
                if not valid_marks:
                    return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

            # Fetch fees
            fees = db_session.query(
                Fee.admission_no, Fee.total_fee, Fee.amount_paid, Fee.balance, Fee.grade, Fee.term, Fee.year
            ).filter(
                func.lower(Fee.grade) == grade.lower().strip(),
                func.lower(Fee.term).in_(term_variations),
                Fee.year == year,
                Fee.admission_no != None
            ).all()

            # Validate and convert fees
            valid_fees = []
            invalid_fees = []
            for f in fees:
                try:
                    total_fee = float(f.total_fee) if f.total_fee is not None else 0.0
                    amount_paid = float(f.amount_paid) if f.amount_paid is not None else 0.0
                    balance = float(f.balance) if f.balance is not None else 0.0
                    if all(isinstance(x, (int, float)) and x >= 0 for x in [total_fee, amount_paid, balance]):
                        valid_fees.append((
                            f.admission_no.strip(),  # admission_no
                            total_fee,  # total_fee
                            amount_paid,  # amount_paid
                            balance,  # balance
                            f.grade.strip(),  # grade
                            f.term.strip(),  # term
                            f.year  # year
                        ))
                    else:
                        invalid_fees.append((f.admission_no, f.total_fee, f.amount_paid, f.balance))
                except (ValueError, TypeError):
                    invalid_fees.append((f.admission_no, f.total_fee, f.amount_paid, f.balance))
            if invalid_fees:
                logger.warning(f"Invalid fees detected: {invalid_fees}")
                flash(f"Invalid fee data for some students: {', '.join([f'{f[0]}' for f in invalid_fees])}. Using default values (0) for affected records.", 'warning')

            # Calculate ranks
            rank_query = db_session.query(
                Marks.admission_no, func.sum(Marks.marks).label('total')
            ).filter(
                func.lower(Marks.grade) == grade.lower().strip(),
                func.lower(Marks.term).in_(term_variations),
                Marks.year == year,
                func.lower(Marks.exam_type).in_(exam_type_variations),
                func.lower(Marks.learning_area).in_([la.lower().strip() for la in learning_areas]),
                Marks.admission_no != None
            ).group_by(Marks.admission_no).order_by(func.sum(Marks.marks).desc())
            rank_query_results = rank_query.all()
            ranks = {}
            for i, r in enumerate(rank_query_results):
                try:
                    total_marks = float(r.total) if r.total is not None else 0.0
                    if 0 <= total_marks <= (100 * len(learning_areas)):
                        ranks[r.admission_no.strip().lower()] = i + 1
                except (ValueError, TypeError):
                    logger.warning(f"Invalid total marks for {r.admission_no}: {r.total}")
                    continue
            total_students = len(ranks) or len(students_formatted)

            # Generate ZIP file
            try:
                zip_buffer = create_zipped_report_cards(
                    students=students_formatted,
                    marks=valid_marks,
                    fees=valid_fees,
                    term=term,
                    year=year,
                    exam_type=exam_type,
                    rank=ranks,
                    total_students=total_students,
                    grade=grade
                )
                if not zip_buffer or len(zip_buffer.getvalue()) == 0:
                    logger.error(f"Failed to generate zipped report cards for grade {grade}")
                    flash(f"Could not generate report cards for grade {grade}. Check database data or contact system admin jonyojss@gmail.com.", 'danger')
                    return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

                # Validate ZIP file integrity
                zip_buffer.seek(0)
                try:
                    with zipfile.ZipFile(zip_buffer, 'r') as test_zip:
                        bad_file = test_zip.testzip()
                        if bad_file:
                            logger.error(f"ZIP integrity check failed, bad file: {bad_file}")
                            flash(f"Generated ZIP file is corrupt. Please contact system admin jonyojss@gmail.com.", 'danger')
                            return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)
                        logger.debug(f"ZIP file contains: {test_zip.namelist()}")
                except zipfile.BadZipFile as e:
                    logger.error(f"Invalid ZIP file: {str(e)}\n{traceback.format_exc()}")
                    flash(f"Generated ZIP file is invalid. Please contact system admin jonyojss@gmail.com.", 'danger')
                    return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

                # Reset buffer position for sending
                zip_buffer.seek(0)
                filename = f'Report_Cards_{grade.replace(" ", "_")}_{term}_{year}_{exam_type}.zip'
                logger.info(f"Admin {current_user.id} downloaded zipped report cards for grade {grade}, size={len(zip_buffer.getvalue())} bytes")
                return send_file(
                    zip_buffer,
                    download_name=filename,
                    as_attachment=True,
                    mimetype='application/zip'
                )

            except Exception as e:
                logger.error(f"Error in create_zipped_report_cards: {str(e)}\n{traceback.format_exc()}")
                flash(f"Error generating report cards: {str(e)}. Contact the system admin jonyojss@gmail.com.", 'danger')
                return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)

        # Set default form values
        form.term.data = term_info['term']
        form.year.data = default_year
        return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in download_report_card: {str(e)}\n{traceback.format_exc()}")
        flash('Database error while generating report cards. Please contact the system admin jonyojss@gmail.com.', 'danger')
        return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in download_report_card: {str(e)}\n{traceback.format_exc()}")
        flash(f"Error generating report cards: {str(e)}. Please try again or contact system admin jonyojss@gmail.com.", 'danger')
        return render_template('download_report_card.html', form=form, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()
@app.route('/admin/link_parent_student', methods=['GET', 'POST'])
@login_required
@admin_required
def link_parent_student():
    """Link a parent to a student by typing username and admission number."""
    form = LinkParentStudentForm()
    db_session = next(get_db())
    try:
        # Fetch term_info
        term_data = db_session.query(TermInfo).first()
        term_info = {
            'term': term_data.term if term_data and term_data.term else 'Term 1',
            'year': term_data.year if term_data and term_data.year else '2025',
            'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
            'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
            'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
        }

        # Prepare content_data
        mission = db_session.query(Mission).first()
        vision = db_session.query(Vision).first()
        about = db_session.query(About).first()
        content_data = {
            'mission': mission.content if mission else "To provide quality education for all students",
            'vision': vision.content if vision else "To be a leading institution in academic excellence",
            'about': about.content if about else "Jonyo Junior School is dedicated to fostering holistic education"
        }

        if form.validate_on_submit():
            parent_username = form.parent_id.data
            admission_no = form.admission_no.data

            # Fetch parent_id from username
            parent = db_session.query(User).filter_by(username=parent_username, role='parent').first()
            if not parent:
                logger.warning(f"Parent not found for username: {parent_username}")
                flash('Parent username not found.', 'danger')
                return render_template('link_parent_student.html', form=form, term_info=term_info, content_data=content_data)

            # Verify student exists
            student = db_session.query(Student).filter_by(admission_no=admission_no).first()
            if not student:
                logger.warning(f"Student not found for admission_no: {admission_no}")
                flash('Student admission number not found.', 'danger')
                return render_template('link_parent_student.html', form=form, term_info=term_info, content_data=content_data)

            # Check if link already exists
            existing_link = db_session.query(ParentStudent).filter_by(parent_id=parent.id, admission_no=admission_no).first()
            if existing_link:
                logger.warning(f"Existing link found for parent_id: {parent.id}, admission_no: {admission_no}")
                flash('Parent is already linked to this student.', 'danger')
                return render_template('link_parent_student.html', form=form, term_info=term_info, content_data=content_data)

            # Create link
            link = ParentStudent(parent_id=parent.id, admission_no=admission_no)
            db_session.add(link)
            db_session.commit()
            logger.info(f"Linked parent_id {parent.id} to admission_no {admission_no}")
            flash('Parent linked to student successfully!', 'success')
            return redirect(url_for('dashboard'))

        return render_template('link_parent_student.html', form=form, term_info=term_info, content_data=content_data)
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in link_parent_student: {str(e)}\n{traceback.format_exc()}")
        flash(f'Database error: {str(e)}. Please try again.', 'danger')
        return render_template('link_parent_student.html', form=form, term_info=term_info, content_data=content_data)
    except Exception as e:
        db_session.rollback()
        logger.error(f"Error in link_parent_student: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}. Please try again or contact support.', 'danger')
        return render_template('link_parent_student.html', form=form, term_info=term_info, content_data=content_data)
    finally:
        db_session.close()

@app.route('/view_links', methods=['GET'])
@login_required
@admin_required
def view_links():
    """View parent-student links."""
    logger.debug(f"Accessing view_links for user: id={getattr(current_user, 'id', 'N/A')}, username={getattr(current_user, 'username', 'N/A')}, role={getattr(current_user, 'role', 'N/A')}")
    db_session = next(get_db())
    try:
        term_info, content_data = fetch_common_data()
        links = db_session.query(
            User.id.label('parent_id'),
            User.username,
            Student.name.label('student_name'),
            ParentStudent.admission_no,
            User.phone_number
        ).join(
            User, ParentStudent.parent_id == User.id
        ).join(
            Student, ParentStudent.admission_no == Student.admission_no
        ).all()
        links = [
            {
                'parent_id': link.parent_id,
                'username': link.username,
                'student_name': link.student_name,
                'admission_no': link.admission_no,
                'phone_number': link.phone_number
            } for link in links
        ]

        if not links:
            logger.debug(f"No parent-student links found for user {getattr(current_user, 'username', 'Unknown')}")
            flash('No parent-student links available.', 'info')

        return render_template('view_links.html', links=links, term_info=term_info, content_data=content_data, current_user=current_user)
    except Exception as e:
        logger.error(f"Error in view_links: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error fetching parent-student links: {str(e)}', 'danger')
        return render_template('view_links.html', links=[], term_info=term_info, content_data=content_data, current_user=current_user)
    finally:
        db_session.close()

@app.route('/delete_link', methods=['POST'])
@login_required
@admin_required
def delete_link():
    """Delete a parent-student link."""
    db_session = next(get_db())
    try:
        parent_username = request.form.get('parent_id')  # Form sends username
        admission_no = request.form.get('admission_no')
        if not parent_username or not admission_no:
            logger.warning("Invalid request: missing parent_id or admission_no")
            flash('Invalid request.', 'danger')
            return redirect(url_for('view_links'))

        # Fetch parent_id from username
        parent = db_session.query(User).filter_by(username=parent_username, role='parent').first()
        if not parent:
            logger.warning(f"Parent not found for username: {parent_username}")
            flash('Parent not found.', 'danger')
            return redirect(url_for('view_links'))

        # Delete link
        result = db_session.query(ParentStudent).filter_by(parent_id=parent.id, admission_no=admission_no).delete()
        if result == 0:
            logger.warning(f"No link found for parent_id: {parent.id}, admission_no: {admission_no}")
            flash('No link found to delete.', 'warning')
        else:
            db_session.commit()
            logger.info(f"Deleted link for parent_id: {parent.id}, admission_no: {admission_no}")
            flash('Parent-student link deleted successfully.', 'success')
    except SQLAlchemyError as e:
        db_session.rollback()
        logger.error(f"Database error in delete_link: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error deleting link: {str(e)}', 'danger')
    except Exception as e:
        db_session.rollback()
        logger.error(f"Unexpected error in delete_link: {str(e)}\n{traceback.format_exc()}")
        flash(f'Error: {str(e)}. Please try again or contact support.', 'danger')
    finally:
        db_session.close()
    return redirect(url_for('view_links'))


    
if __name__ == '__main__':
    app.run(debug=True)