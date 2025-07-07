import io
import zipfile
import logging
import random
import string
import traceback
from database import get_db, User, Student, Marks, Fee, TermInfo, PerformanceLevels, TeacherAssignments, ClassTeachers
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import Table, TableStyle
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Configure logging
logging.basicConfig(filename='app.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def generate_teacher_password():
    """Generate a random teacher password."""
    letter = random.choice(string.ascii_uppercase)
    numbers = ''.join(random.choice(string.digits) for _ in range(5))
    return f"{letter}-{numbers}"

def generate_student_password(name, admission_no):
    """Generate a student password based on name and admission number."""
    first_name = name.strip().split()[0] if name.strip() else 'S'
    initial = first_name[0].upper() if first_name else 'S'
    return f"{initial}-{admission_no}"

def get_learning_areas(grade):
    """Fetch learning areas for a specific grade from the database."""
    try:
        if not isinstance(grade, str):
            logger.warning(f"Invalid grade for learning areas: {grade}")
            return []
        grade = grade.strip()
        db_session = next(get_db())
        try:
            learning_areas = db_session.query(Marks.learning_area).filter(
                func.lower(Marks.grade) == grade.lower()
            ).distinct().order_by(Marks.learning_area).all()
            areas = [la[0].strip() for la in learning_areas if la[0]]
            if not areas:
                logger.warning(f"No learning areas found in Marks for grade {grade}. Using default list.")
                areas = [
                    'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
                    'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
                ]
            logger.debug(f"Learning areas for grade {grade}: {areas}")
            return areas
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching learning areas for grade {grade}: {str(e)}")
        return [
            'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
            'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
        ]

def get_performance_levels(marks, type_='learning_area'):
    """Determine performance level based on marks and type."""
    try:
        if not isinstance(marks, (int, float)) or marks < 0 or marks > 100:
            logger.warning(f"Invalid marks for performance level: {marks}, type={type_}")
            return 'N/A'
        marks = float(marks)
        if type_ in ('learning_area', 'class_teacher', 'principal'):
            if marks >= 90:
                return 'EE1'
            elif marks >= 80:
                return 'EE2'
            elif marks >= 70:
                return 'ME1'
            elif marks >= 60:
                return 'ME2'
            elif marks >= 50:
                return 'AE1'
            else:
                return 'AE2'
        return 'N/A'
    except (ValueError, TypeError) as e:
        logger.error(f"Error in get_performance_levels: marks={marks}, type={type_}, error={str(e)}")
        return 'N/A'

def get_points(marks, type_='learning_area'):
    """Calculate points based on marks."""
    try:
        if not isinstance(marks, (int, float)) or marks < 0 or marks > 100:
            logger.warning(f"Invalid marks for points calculation: {marks}, type={type_}")
            return 0.0
        marks = float(marks)
        if type_ in ('learning_area', 'class_teacher', 'principal'):
            if marks >= 90:
                return 4.0  # EE1
            elif marks >= 80:
                return 3.5  # EE2
            elif marks >= 70:
                return 3.0  # ME1
            elif marks >= 60:
                return 2.5  # ME2
            elif marks >= 50:
                return 2.0  # AE1
            else:
                return 1.0  # AE2
        return 0.0
    except (ValueError, TypeError) as e:
        logger.error(f"Error in get_points: marks={marks}, type={type_}, error={str(e)}")
        return 0.0

def get_teacher_comment(level):
    """Return teacher comment based on performance levels."""
    comments = {
        'EE1': 'Outstanding performance! Keep it up.',
        'EE2': 'Excellent work, consistently strong.',
        'ME1': 'Good effort, meets expectations.',
        'ME2': 'Satisfactory, room for improvement.',
        'AE1': 'Approaching expectations, needs more focus.',
        'AE2': 'Below expectations, seek extra support.',
        'N/A': 'No comment available.'
    }
    return comments.get(level, 'No comment available.')

def get_teacher_name(learning_area, grade):
    """Fetch teacher name for a learning area and grade."""
    try:
        if not isinstance(learning_area, str) or not isinstance(grade, str):
            logger.warning(f"Invalid input: learning_area={learning_area}, grade={grade}")
            return 'N/A'
        learning_area = learning_area.strip()
        grade = grade.strip()
        db_session = next(get_db())
        try:
            teacher = db_session.query(User.username).join(
                TeacherAssignments, User.id == TeacherAssignments.teacher_id
            ).filter(
                func.lower(TeacherAssignments.learning_area) == learning_area.lower(),
                func.lower(TeacherAssignments.grade) == grade.lower()
            ).first()
            return teacher[0].strip() if teacher and teacher[0] else 'N/A'
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching teacher for {learning_area}, {grade}: {str(e)}")
        return 'N/A'

def get_class_teacher_name(grade):
    """Fetch class teacher name for a grade."""
    try:
        if not isinstance(grade, str):
            logger.warning(f"Invalid grade: {grade}")
            return 'N/A'
        grade = grade.strip()
        db_session = next(get_db())
        try:
            teacher = db_session.query(User.username).join(
                ClassTeachers, User.id == ClassTeachers.teacher_id
            ).filter(
                func.lower(ClassTeachers.grade) == grade.lower()
            ).first()
            return teacher[0].strip() if teacher and teacher[0] else 'N/A'
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching class teacher for {grade}: {str(e)}")
        return 'N/A'

def get_principal_name():
    """Fetch principal name from TermInfo."""
    try:
        db_session = next(get_db())
        try:
            principal = db_session.query(TermInfo.principal).filter_by(id=1).first()
            return principal[0].strip() if principal and principal[0] else 'Mr. Principal'
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching principal name: {str(e)}")
        return 'Mr. Principal'

def get_rank(admission_no, grade, term, year, exam_type):
    """Calculate student rank based on total marks."""
    try:
        if not all(isinstance(x, str) for x in [admission_no, grade, term, exam_type, year]):
            logger.warning(f"Invalid inputs for rank: admission_no={admission_no}, grade={grade}, term={term}, year={year}, exam_type={exam_type}")
            return 'N/A'
        admission_no = admission_no.strip()
        grade = grade.strip()
        term = term.strip()
        exam_type = exam_type.strip()
        year = year.strip()
        db_session = next(get_db())
        try:
            ranks = db_session.query(
                Marks.admission_no,
                func.sum(Marks.total_marks).label('total')
            ).filter(
                func.lower(Marks.grade) == grade.lower(),
                func.lower(Marks.term) == term.lower(),
                Marks.year == year,
                func.lower(Marks.exam_type) == exam_type.lower()
            ).group_by(Marks.admission_no).order_by(func.sum(Marks.total_marks).desc()).all()
            rank_list = [r[0].strip() for r in ranks]
            rank = rank_list.index(admission_no) + 1 if admission_no in rank_list else 'N/A'
            logger.debug(f"Rank for {admission_no} in {grade}, {term}, {year}, {exam_type}: {rank}")
            return str(rank)
        finally:
            db_session.close()
    except (SQLAlchemyError, ValueError) as e:
        logger.error(f"Error calculating rank for {admission_no}: {str(e)}")
        return 'N/A'

def get_total_fee(admission_no):
    """Fetch total fee for a student for the latest term."""
    try:
        if not isinstance(admission_no, str):
            logger.warning(f"Invalid admission_no: {admission_no}")
            return 0.0
        admission_no = admission_no.strip()
        db_session = next(get_db())
        try:
            fee = db_session.query(Fee.total_fee).filter(
                func.lower(Fee.admission_no) == admission_no.lower()
            ).order_by(Fee.term.desc()).first()
            return float(fee[0]) if fee and fee[0] is not None else 0.0
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching total fee for {admission_no}: {str(e)}")
        return 0.0

def get_amount_paid(admission_no):
    """Fetch amount paid for a student for the latest term."""
    try:
        if not isinstance(admission_no, str):
            logger.warning(f"Invalid admission_no: {admission_no}")
            return 0.0
        admission_no = admission_no.strip()
        db_session = next(get_db())
        try:
            paid = db_session.query(Fee.amount_paid).filter(
                func.lower(Fee.admission_no) == admission_no.lower()
            ).order_by(Fee.term.desc()).first()
            return float(paid[0]) if paid and paid[0] is not None else 0.0
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching amount paid for {admission_no}: {str(e)}")
        return 0.0

def get_balance(admission_no):
    """Fetch fee balance for a student for the latest term."""
    try:
        if not isinstance(admission_no, str):
            logger.warning(f"Invalid admission_no: {admission_no}")
            return 0.0
        admission_no = admission_no.strip()
        db_session = next(get_db())
        try:
            balance = db_session.query(Fee.balance).filter(
                func.lower(Fee.admission_no) == admission_no.lower()
            ).order_by(Fee.term.desc()).first()
            return float(balance[0]) if balance and balance[0] is not None else 0.0
        finally:
            db_session.close()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching balance for {admission_no}: {str(e)}")
        return 0.0

def get_class_teacher_comment(total_marks, grade):
    """Generate class teacher comment based on total marks."""
    try:
        if not isinstance(grade, str) or not isinstance(total_marks, (int, float)):
            logger.warning(f"Invalid inputs: total_marks={total_marks}, grade={grade}")
            return 'No comment available.'
        level = get_performance_levels(total_marks, 'class_teacher')
        return get_teacher_comment(level)
    except Exception as e:
        logger.error(f"Error fetching class teacher comment: total_marks={total_marks}, grade={grade}, error={str(e)}")
        return 'No comment available.'

def get_principal_comment(total_marks):
    """Generate principal comment based on total marks."""
    try:
        if not isinstance(total_marks, (int, float)):
            logger.warning(f"Invalid total_marks: {total_marks}")
            return 'No comment available.'
        level = get_performance_levels(total_marks, 'principal')
        return get_teacher_comment(level)
    except Exception as e:
        logger.error(f"Error fetching principal comment: total_marks={total_marks}, error={str(e)}")
        return 'No comment available.'

def generate_individual_report_card(admission_no, term, year, exam_type):
    """Generate a PDF report card for a single student."""
    try:
        if not all(isinstance(x, str) for x in [admission_no, term, year, exam_type]):
            logger.error(f"Invalid input types: admission_no={type(admission_no)}, term={type(term)}, year={type(year)}, exam_type={type(exam_type)}")
            raise ValueError("All inputs must be strings")

        # Normalize exam_type
        exam_type = exam_type.lower().replace('cat 1', 'cat1').replace('cat 2', 'cat2').replace('cat 3', 'cat3') \
            .replace('rat 1', 'rat1').replace('rat 2', 'rat2').replace('rat 3', 'rat3') \
            .replace('mid term', 'midterm').replace('end term', 'endterm') \
            .replace('project 1', 'project1').replace('project 2', 'project2').replace('project 3', 'project3')

        logger.debug(f"Generating individual report card for admission_no={admission_no}, term={term}, year={year}, exam_type={exam_type}")

        db_session = next(get_db())
        try:
            # Fetch term_info
            term_data = db_session.query(TermInfo).filter_by(id=1).first()
            term_info = {
                'term': term_data.term if term_data and term_data.term else term,
                'year': term_data.year if term_data and term_data.year else year,
                'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
                'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
                'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
            }

            # Fetch student details
            student = db_session.query(User.username, User.grade).filter(
                func.lower(User.admission_no) == admission_no.lower(),
                User.role == 'student'
            ).first()
            if not student:
                logger.warning(f"No student found for admission_no={admission_no}")
                return None
            name, grade = student

            # Fetch marks
            marks = db_session.query(
                Marks.admission_no, Marks.learning_area, Marks.total_marks, Marks.term, Marks.year, Marks.grade, Marks.exam_type
            ).filter(
                func.lower(Marks.admission_no) == admission_no.lower(),
                func.lower(Marks.grade) == grade.lower(),
                func.lower(Marks.term) == term.lower(),
                Marks.year == year,
                func.lower(Marks.exam_type) == exam_type.lower()
            ).all()
            marks = [
                (m.admission_no, m.learning_area, m.total_marks, m.term, m.year, m.grade, m.exam_type)
                for m in marks
            ]
            logger.debug(f"Fetched {len(marks)} marks for {admission_no}: {[m[:3] for m in marks]}")

            # Fetch fees
            fees = db_session.query(
                Fee.admission_no, Fee.total_fee, Fee.amount_paid, Fee.balance, Fee.grade, Fee.term, Fee.year
            ).filter(
                func.lower(Fee.admission_no) == admission_no.lower(),
                func.lower(Fee.grade) == grade.lower(),
                func.lower(Fee.term) == term.lower(),
                Fee.year == year
            ).all()
            fees = [
                (f.admission_no, f.total_fee, f.amount_paid, f.balance, f.grade, f.term, f.year)
                for f in fees
            ]
            logger.debug(f"Fetched {len(fees)} fees for {admission_no}: {[f[:4] for f in fees]}")

            # Calculate rank
            rank = get_rank(admission_no, grade, term, year, exam_type)

            # Count total students
            total_students = db_session.query(func.count(func.distinct(Marks.admission_no))).filter(
                func.lower(Marks.grade) == grade.lower(),
                func.lower(Marks.term) == term.lower(),
                Marks.year == year,
                func.lower(Marks.exam_type) == exam_type.lower()
            ).scalar()

            # Prepare student data
            student_data = [(None, name, None, None, grade, admission_no)]

            # Generate the report card
            pdf_buffer = generate_report_card(
                students=student_data,
                marks=marks,
                fees=fees,
                term=term,
                year=year,
                exam_type=exam_type,
                rank=rank,
                total_students=total_students,
                grade=grade,
                term_info=term_info
            )

            if pdf_buffer is None or len(pdf_buffer.getvalue()) == 0:
                logger.error(f"Failed to generate PDF for admission_no={admission_no}")
                return None

            pdf_content = pdf_buffer.getvalue()
            if not pdf_content.startswith(b'%PDF-'):
                logger.error(f"Invalid PDF generated for admission_no={admission_no}, starts with: {pdf_content[:10]}")
                pdf_buffer.close()
                return None

            logger.info(f"Individual report card generated for admission_no={admission_no}, size={len(pdf_content)} bytes")
            return pdf_buffer
        finally:
            db_session.close()
    except Exception as e:
        logger.error(f"Error in generate_individual_report_card for admission_no={admission_no}: {str(e)}\n{traceback.format_exc()}")
        if 'pdf_buffer' in locals():
            pdf_buffer.close()
        return None

def create_zipped_report_cards(students, marks, fees, term, year, exam_type, rank, total_students, grade):
    """Generate a ZIP file containing report cards for multiple students."""
    try:
        logger.debug(f"Creating zipped report cards for grade: {grade}, term: {term}, year: {year}, exam_type: {exam_type}")
        db_session = next(get_db())
        try:
            term_data = db_session.query(TermInfo).filter_by(id=1).first()
            term_info = {
                'term': term_data.term if term_data and term_data.term else term,
                'year': term_data.year if term_data and term_data.year else year,
                'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
                'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
                'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
            }
        finally:
            db_session.close()

        zip_buffer = io.BytesIO()
        report_cards_added = 0
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for student in students:
                try:
                    if isinstance(student, (list, tuple)) and len(student) >= 6:
                        admission_no = str(student[5]).strip()
                        name = str(student[1]).strip()
                        student_grade = str(student[4]).strip()
                    elif isinstance(student, dict):
                        admission_no = str(student.get('admission_no', '')).strip()
                        name = str(student.get('name', '')).strip()
                        student_grade = str(student.get('grade', '')).strip()
                    else:
                        logger.warning(f"Invalid student data: {student}")
                        continue

                    if not all([admission_no, name, student_grade]):
                        logger.warning(f"Invalid student data: admission_no={admission_no}, name={name}, grade={student_grade}")
                        continue
                    if student_grade.lower() != grade.lower():
                        logger.debug(f"Skipping student {admission_no} with grade {student_grade} (expected {grade})")
                        continue

                    student_marks = [
                        m for m in marks
                        if str(m[0] if isinstance(m, (list, tuple)) else m.get('admission_no', '')).strip().lower() == admission_no.lower()
                    ]
                    student_fees = [
                        f for f in fees
                        if str(f[0] if isinstance(f, (list, tuple)) else f.get('admission_no', '')).strip().lower() == admission_no.lower()
                    ] or [(admission_no, 0, 0, 0, grade, term, year)]
                    logger.debug(f"Processing {admission_no}: {len(student_marks)} marks, {len(student_fees)} fees")

                    pdf_buffer = generate_report_card(
                        students=[(None, name, None, None, student_grade, admission_no)],
                        marks=student_marks,
                        fees=student_fees,
                        term=term,
                        year=year,
                        exam_type=exam_type,
                        rank=rank.get(admission_no, 'N/A') if isinstance(rank, dict) else rank,
                        total_students=total_students,
                        grade=grade,
                        term_info=term_info
                    )
                    if pdf_buffer is None or len(pdf_buffer.getvalue()) == 0:
                        logger.warning(f"Failed to generate PDF for {admission_no}")
                        continue
                    pdf_content = pdf_buffer.getvalue()
                    if not pdf_content.startswith(b'%PDF-'):
                        logger.warning(f"Invalid PDF content for {admission_no}, starts with: {pdf_content[:10]}")
                        pdf_buffer.close()
                        continue

                    zip_file.writestr(
                        f'Report_Card_{admission_no}_{grade.replace(" ", "_")}_{term}_{year}_{exam_type}.pdf',
                        pdf_content
                    )
                    report_cards_added += 1
                    logger.debug(f"Added report card for {admission_no}, size={len(pdf_content)} bytes")
                    pdf_buffer.close()
                except Exception as e:
                    logger.error(f"Error generating report card for {admission_no}: {str(e)}\n{traceback.format_exc()}")
                    continue

            if report_cards_added == 0:
                missing_students = [
                    s[5] if isinstance(s, (list, tuple)) else s.get('admission_no', '')
                    for s in students
                    if (s[5] if isinstance(s, (list, tuple)) else s.get('admission_no', '')) not in
                       [m[0] if isinstance(m, (list, tuple)) else m.get('admission_no', '') for m in marks]
                ]
                error_message = f"No report cards generated for {grade}, {term}, {year}, {exam_type}.\n"
                if missing_students:
                    error_message += f"Missing marks for students: {', '.join(map(str, missing_students))}.\n"
                error_message += "Please ensure marks are entered for the selected grade, term, year, and exam type in the database."
                zip_file.writestr('no_report_cards.txt', error_message.encode('utf-8'))
                report_cards_added += 1
                logger.warning(error_message)

        zip_buffer.seek(0)
        if report_cards_added == 0:
            logger.error(f"No files added to ZIP for grade={grade}")
            zip_buffer.close()
            return None

        try:
            with zipfile.ZipFile(zip_buffer, 'r') as zip_test:
                bad_file = zip_test.testzip()
                if bad_file is not None:
                    logger.error(f"Corrupted ZIP file, bad file: {bad_file}")
                    zip_buffer.close()
                    return None
        except zipfile.BadZipFile as e:
            logger.error(f"Invalid ZIP file generated: {str(e)}\n{traceback.format_exc()}")
            zip_buffer.close()
            return None

        logger.info(f"Zipped report cards generated for {grade}, {report_cards_added} files, size: {len(zip_buffer.getvalue())} bytes")
        return zip_buffer
    except Exception as e:
        logger.error(f"Error in create_zipped_report_cards: {str(e)}\n{traceback.format_exc()}")
        if 'zip_buffer' in locals():
            zip_buffer.close()
        return None

def generate_report_card(students, marks, fees, term, year, exam_type, rank=None, total_students=None, grade=None, term_info=None):
    """Generate a PDF report card for multiple students."""
    try:
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        processed_students = set()
        has_valid_content = False

        if not isinstance(students, (list, tuple)):
            logger.error(f"Invalid students type: {type(students)}")
            raise ValueError("Students must be a list or tuple")
        if not isinstance(marks, (list, tuple)):
            logger.error(f"Invalid marks type: {type(marks)}")
            raise ValueError("Marks must be a list or tuple")
        if not isinstance(fees, (list, tuple)):
            logger.error(f"Invalid fees type: {type(fees)}")
            raise ValueError("Fees must be a list or tuple")
        if not all(isinstance(x, str) for x in [term, year, exam_type] if x is not None):
            logger.error(f"Invalid string inputs: term={type(term)}, year={type(year)}, exam_type={type(exam_type)}")
            raise ValueError("Term, year, and exam_type must be strings")

        term = term.strip().lower()
        year = year.strip()
        exam_type = exam_type.strip().lower().replace('cat 1', 'cat1').replace('cat 2', 'cat2').replace('cat 3', 'cat3') \
            .replace('rat 1', 'rat1').replace('rat 2', 'rat2').replace('rat 3', 'rat3') \
            .replace('mid term', 'midterm').replace('end term', 'endterm') \
            .replace('project 1', 'project1').replace('project 2', 'project2').replace('project 3', 'project3')
        grade = grade.strip() if grade else None

        term_info = term_info or {
            'term': term.capitalize(),
            'year': year,
            'principal': get_principal_name(),
            'start_date': '2025-01-01',
            'end_date': '2025-04-01'
        }

        # Fetch learning areas
        learning_areas = get_learning_areas(grade) if grade else [
            'Mathematics', 'English', 'Kiswahili', 'Integrated Science', 'Pre-technical',
            'Social Studies', 'Agriculture and Nutrition', 'Creative Arts', 'CRE'
        ]
        logger.debug(f"Learning areas for grade {grade}: {learning_areas}")

        logger.debug(f"Processing students: {len(students)} students, {len(marks)} marks, term={term}, year={year}, exam_type={exam_type}, grade={grade}")

        for student in students:
            try:
                if isinstance(student, (list, tuple)) and len(student) >= 6:
                    admission_no = str(student[5]).strip().lower()
                    name = str(student[1]).strip()
                    student_grade = str(student[4]).strip()
                elif isinstance(student, dict):
                    admission_no = str(student.get('admission_no', '')).strip().lower()
                    name = str(student.get('name', '')).strip()
                    student_grade = str(student.get('grade', '')).strip()
                else:
                    logger.warning(f"Invalid student data: {student}")
                    continue

                if not all([admission_no, name, student_grade]):
                    logger.warning(f"Empty or invalid student data: admission_no={admission_no}, name={name}, grade={student_grade}")
                    continue
                if grade and student_grade.lower() != grade.lower():
                    logger.debug(f"Skipping student {admission_no} with grade {student_grade} (expected {grade})")
                    continue
                if admission_no in processed_students:
                    logger.debug(f"Skipping duplicate admission_no: {admission_no}")
                    continue

                student_marks = [
                    m for m in marks
                    if (isinstance(m, dict) and str(m.get('admission_no', '')).strip().lower() == admission_no) or
                       (isinstance(m, (list, tuple)) and len(m) >= 3 and str(m[0]).strip().lower() == admission_no)
                ]
                logger.debug(f"Marks for {admission_no}: {len(student_marks)} entries: {[m[:3] for m in student_marks]}")

                processed_students.add(admission_no)
                has_valid_content = True

                # Header
                c.setFont("Helvetica-Bold", 16)
                c.drawCentredString(300, 750, "JONYO JUNIOR SECONDARY SCHOOL")
                c.setFont("Helvetica", 14)
                c.drawCentredString(300, 730, "REPORT CARD")
                c.setFont("Helvetica", 12)
                c.drawString(50, 700, f"Name: {name}")
                c.drawString(50, 680, f"Admission No: {admission_no}")
                c.drawString(50, 660, f"Grade: {student_grade}")
                formatted_exam_type = exam_type.replace('cat1', 'CAT 1').replace('cat2', 'CAT 2').replace('cat3', 'CAT 3') \
                    .replace('rat1', 'RAT 1').replace('rat2', 'RAT 2').replace('rat3', 'RAT 3') \
                    .replace('midterm', 'Mid Term').replace('endterm', 'End Term') \
                    .replace('project1', 'Project 1').replace('project2', 'Project 2').replace('project3', 'Project 3')
                c.drawString(50, 640, f"Term: {term_info['term']} {year} ({formatted_exam_type})")
                c.drawString(50, 620, f"School Year: {term_info['year']}")

                # Create table data
                table_data = [['Learning Area', 'Marks', 'Perf. Level', 'Points', 'Teacher Comment', 'Teacher']]
                total_marks = 0
                total_points = 0
                subjects_covered = set()

                for learning_area in learning_areas:
                    mark_found = False
                    learning_area_normalized = learning_area.lower().strip()
                    for mark in student_marks:
                        mark_learning_area = str(mark[1] if isinstance(mark, (list, tuple)) else mark.get('learning_area', '')).strip().lower()
                        if mark_learning_area == learning_area_normalized and mark_learning_area not in subjects_covered:
                            try:
                                marks_value = float(mark[2] if isinstance(mark, (list, tuple)) else mark.get('total_marks', 0)) if (mark[2] if isinstance(mark, (list, tuple)) else mark.get('total_marks')) is not None and str(mark[2] if isinstance(mark, (list, tuple)) else mark.get('total_marks', '')).strip() != '' else None
                                if marks_value is not None and 0 <= marks_value <= 100:
                                    total_marks += marks_value
                                    points = get_points(marks_value, 'learning_area')
                                    total_points += points
                                    level = get_performance_levels(marks_value, 'learning_area')
                                    table_data.append([
                                        learning_area,
                                        f"{int(marks_value)}",
                                        level,
                                        f"{points:.2f}",
                                        get_teacher_comment(level)[:20],
                                        get_teacher_name(learning_area, student_grade)[:15]
                                    ])
                                    subjects_covered.add(mark_learning_area)
                                    mark_found = True
                                    logger.debug(f"Processed mark for {admission_no}, {learning_area}: {marks_value}, Points: {points}")
                                else:
                                    logger.warning(f"Invalid marks value {marks_value} for {admission_no}, {learning_area}")
                            except (ValueError, TypeError) as e:
                                logger.error(f"Error processing mark for {admission_no}, {learning_area}: {str(e)}")
                    if not mark_found:
                        logger.debug(f"No marks found for {admission_no} in {learning_area}")
                        table_data.append([learning_area, '-', 'N/A', '0.00', 'No marks available', 'N/A'])

                # Create table
                col_widths = [2.0*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1.5*inch, 1.2*inch]
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                table.wrapOn(c, 500, 400)
                table_height = table._height
                table_x = 30
                table_y = 600 - table_height
                if table_y < 100:
                    logger.warning(f"Table y-position {table_y} too low for {admission_no}, splitting table")
                    rows_per_page = max(5, int((600 - table_y) / 20))
                    for i in range(0, len(table_data), rows_per_page):
                        sub_table = Table(table_data[:1] + table_data[i+1:i+rows_per_page+1], colWidths=col_widths)
                        sub_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 1), (-1, -1), 9),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('LEFTPADDING', (0, 0), (-1, -1), 4),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                            ('TOPPADDING', (0, 0), (-1, -1), 4),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ]))
                        sub_table.wrapOn(c, 500, 400)
                        sub_table.drawOn(c, table_x, 600 - sub_table._height)
                        c.showPage()
                        c.setFont("Helvetica-Bold", 16)
                        c.drawCentredString(300, 750, "JONYO JUNIOR SECONDARY SCHOOL")
                        c.setFont("Helvetica", 14)
                        c.drawCentredString(300, 730, "REPORT CARD")
                        c.setFont("Helvetica", 12)
                        c.drawString(50, 700, f"Name: {name}")
                        c.drawString(50, 680, f"Admission No: {admission_no}")
                        c.drawString(50, 660, f"Grade: {student_grade}")
                        c.drawString(50, 640, f"Term: {term_info['term']} {year} ({formatted_exam_type})")
                        c.drawString(50, 620, f"School Year: {term_info['year']}")
                    table = None
                else:
                    table.drawOn(c, table_x, table_y)

                # Footer details
                y = max(table_y - 20, 100) if table else 580
                student_fees = [
                    f for f in fees
                    if (isinstance(f, dict) and str(f.get('admission_no', '')).strip().lower() == admission_no) or
                       (isinstance(f, (list, tuple)) and len(f) >= 7 and str(f[0]).strip().lower() == admission_no)
                ]
                fee_info = student_fees[0] if student_fees else (admission_no, 0, 0, 0, grade, term, year)

                total_fee = float(fee_info[1] if isinstance(fee_info, (list, tuple)) else fee_info.get('total_fee', 0) or 0)
                balance = float(fee_info[3] if isinstance(fee_info, (list, tuple)) else fee_info.get('balance', 0) or 0)

                student_rank = rank.get(admission_no, 'N/A') if isinstance(rank, dict) else rank if rank else get_rank(admission_no, student_grade, term, year, exam_type)

                c.setFont("Helvetica", 10)
                c.drawString(30, y-20, f"Rank: {student_rank} out of {total_students if total_students else 'N/A'}")
                c.drawString(30, y-40, f"Total Marks: {int(total_marks)}" if total_marks > 0 else "Total Marks: N/A")
                c.drawString(30, y-60, f"Total Points: {total_points:.2f}" if total_points > 0 else "Total Points: N/A")
                c.drawString(30, y-80, f"Performance Level: {get_performance_levels(total_marks, 'class_teacher') if total_marks > 0 else 'N/A'}")
                c.drawString(30, y-100, f"Total Fee: {total_fee:,.2f}")
                c.drawString(30, y-120, f"Balance: {balance:,.2f}")
                c.drawString(30, y-140, f"Class Teacher Comment: {get_class_teacher_comment(total_marks, student_grade)[:50] if total_marks > 0 else 'N/A'}")
                c.drawString(30, y-160, f"Principal Comment: {get_principal_comment(total_marks)[:50] if total_marks > 0 else 'N/A'}")
                c.drawString(30, y-180, f"Class Teacher: {get_class_teacher_name(student_grade)[:20]}")
                c.drawString(30, y-200, f"Principal: {get_principal_name()[:20]}")
                c.drawString(30, y-220, f"Start Date: {term_info['start_date']}")
                c.drawString(30, y-240, f"End Date: {term_info['end_date']}")
                c.drawString(30, y-260, "School Stamp: ____________________")

                # Watermark
                c.setFont("Helvetica", 50)
                c.setFillColor(colors.grey, alpha=0.2)
                c.rotate(45)
                c.drawCentredString(400, 200, "JONYO JSS")
                c.rotate(-45)
                c.setFillColor(colors.black)
                c.showPage()
            except Exception as e:
                logger.error(f"Error processing student {admission_no}: {str(e)}\n{traceback.format_exc()}")
                continue

        if not has_valid_content:
            logger.warning(f"No valid report cards generated for {grade}, {term}, {year}, {exam_type}")
            c.setFont("Helvetica", 12)
            c.drawString(50, 700, "No valid student data or marks available.")
            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        c.save()
        buffer.seek(0)
        pdf_content = buffer.getvalue()
        if not pdf_content.startswith(b'%PDF-'):
            logger.error(f"Invalid PDF generated for {grade}, starts with: {pdf_content[:10]}")
            buffer.close()
            return None
        logger.debug(f"PDF report card generated for {grade}, size: {len(pdf_content)} bytes")
        return buffer
    except Exception as e:
        logger.error(f"Error in generate_report_card: {str(e)}\n{traceback.format_exc()}")
        if 'buffer' in locals():
            buffer.close()
        return None

def generate_excel_results(students, marks, grade, term=None, year=None, exam_type=None):
    """Generate an Excel file with student results."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Results_{grade.replace(' ', '_')}"
        learning_areas = sorted(set(
            m.learning_area if isinstance(m, (Marks, dict)) else m[1]
            for m in marks
            if (isinstance(m, (Marks, dict)) and str(m.get('grade', m.grade if isinstance(m, Marks) else '')).lower().strip() == grade.lower().strip()) or
               (isinstance(m, (list, tuple)) and len(m) >= 7 and str(m[5]).lower().strip() == grade.lower().strip())
        ))
        headers = ['Admission No', 'Name']
        for la in learning_areas:
            headers.extend([f"{la} Marks", f"{la} Level"])
        headers.extend(['Total Marks', 'Total Points', 'Performance Level', 'Rank'])
        ws.append(headers)
        logger.debug(f"Excel headers for {grade}: {headers}")
        processed_students = set()
        total_students = 0

        for student in students:
            try:
                if isinstance(student, (list, tuple)) and len(student) >= 6:
                    admission_no = str(student[5]).strip()
                    name = str(student[1]).strip()
                    student_grade = str(student[4]).strip()
                elif isinstance(student, dict):
                    admission_no = str(student.get('admission_no', '')).strip()
                    name = str(student.get('name', '')).strip()
                    student_grade = str(student.get('grade', '')).strip()
                elif isinstance(student, (User, Student)):
                    admission_no = str(student.admission_no).strip()
                    name = str(student.name if isinstance(student, Student) else student.username).strip()
                    student_grade = str(student.grade).strip()
                else:
                    logger.warning(f"Invalid student data: {student}")
                    continue

                if grade != 'all' and student_grade.lower().strip() != grade.lower().strip():
                    continue
                if admission_no in processed_students:
                    continue
                processed_students.add(admission_no)
                total_students += 1

                row = [admission_no, name]
                student_marks = {
                    (m.learning_area if isinstance(m, (Marks, dict)) else m[1]).strip().lower(): (m.total_marks if isinstance(m, (Marks, dict)) else m[2])
                    for m in marks
                    if ((isinstance(m, (Marks, dict)) and str(m.get('admission_no', m.admission_no if isinstance(m, Marks) else '')).strip().lower() == admission_no.lower()) or
                        (isinstance(m, (list, tuple)) and len(m) >= 7 and str(m[0]).strip().lower() == admission_no.lower())) and
                       ((isinstance(m, (Marks, dict)) and str(m.get('grade', m.grade if isinstance(m, Marks) else '')).lower().strip() == grade.lower().strip()) or
                        (isinstance(m, (list, tuple)) and len(m) >= 7 and str(m[5]).lower().strip() == grade.lower().strip()))
                }
                total_marks = sum(float(m) for m in student_marks.values() if m is not None)
                total_points = sum(
                    get_points(float(m.total_marks if isinstance(m, (Marks, dict)) else m[2]), 'learning_area')
                    for m in marks
                    if ((isinstance(m, (Marks, dict)) and str(m.get('admission_no', m.admission_no if isinstance(m, Marks) else '')).strip().lower() == admission_no.lower()) or
                        (isinstance(m, (list, tuple)) and len(m) >= 7 and str(m[0]).strip().lower() == admission_no.lower())) and
                       ((isinstance(m, (Marks, dict)) and str(m.get('grade', m.grade if isinstance(m, Marks) else '')).lower().strip() == grade.lower().strip()) or
                        (isinstance(m, (list, tuple)) and len(m) >= 7 and str(m[5]).lower().strip() == grade.lower().strip()))
                )
                for la in learning_areas:
                    marks_value = student_marks.get(la.lower().strip(), None)
                    marks_display = int(marks_value) if marks_value is not None else ''
                    level = get_performance_levels(marks_value, 'learning_area') if marks_value is not None else ''
                    row.extend([marks_display, level])
                performance_level = get_performance_levels(total_marks, 'class_teacher') if total_marks > 0 else 'N/A'
                first_mark = next((m for m in marks if (isinstance(m, (Marks, dict)) and str(m.get('admission_no', m.admission_no if isinstance(m, Marks) else '')).strip().lower() == admission_no.lower()) or
                                  (isinstance(m, (list, tuple)) and len(m) >= 7 and str(m[0]).strip().lower() == admission_no.lower())), None)
                rank_term = term if term else (first_mark.term if isinstance(first_mark, (Marks, dict)) else first_mark[3] if first_mark else 'Term 1')
                rank_year = year if year else (first_mark.year if isinstance(first_mark, (Marks, dict)) else first_mark[4] if first_mark else '2025')
                rank_exam_type = exam_type if exam_type else (first_mark.exam_type if isinstance(first_mark, (Marks, dict)) else first_mark[6] if first_mark else 'endterm')
                rank = get_rank(admission_no, grade, rank_term, rank_year, rank_exam_type) if first_mark else 'N/A'
                row.extend([int(total_marks) if total_marks > 0 else '', f"{total_points:.2f}" if total_points > 0 else '', performance_level, rank])
                ws.append(row)
                logger.debug(f"Excel row for {admission_no}: {row}")
            except (ValueError, TypeError) as e:
                logger.error(f"Error processing student {admission_no}: {str(e)}")
                continue

        if not processed_students:
            logger.warning(f"No students processed for {grade}")
            ws.append(['No students found for this grade.'])

        header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.column > 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        if len(buffer.getvalue()) == 0:
            logger.error(f"Generated empty Excel file for grade={grade}")
            return None
        logger.debug(f"Excel file generated for {grade}, size: {len(buffer.getvalue())} bytes")
        return buffer
    except Exception as e:
        logger.error(f"Error generating Excel results for grade={grade}: {str(e)}\n{traceback.format_exc()}")
        return None

def generate_fee_statement(fees, grade=None, term=None, year=None):
    """Generate a PDF fee statement."""
    try:
        if not fees:
            logger.warning("No fee data received.")
            return None

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        processed_adm_nos = set()
        has_valid_content = False

        # Fetch term_info
        db_session = next(get_db())
        try:
            term_data = db_session.query(TermInfo).filter_by(id=1).first()
            term_info = {
                'term': term_data.term if term_data and term_data.term else term or 'Term 1',
                'year': term_data.year if term_data and term_data.year else year or '2025',
                'principal': term_data.principal if term_data and term_data.principal else 'Mr. Principal',
                'start_date': term_data.start_date if term_data and term_data.start_date else '2025-01-01',
                'end_date': term_data.end_date if term_data and term_data.end_date else '2025-04-01'
            }
        finally:
            db_session.close()

        # Header
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(300, 750, "JONYO JUNIOR SECONDARY SCHOOL")
        c.setFont("Helvetica", 14)
        c.drawCentredString(300, 730, "FEE STATEMENT")
        c.setFont("Helvetica", 12)
        c.drawString(50, 710, f"School Year: {term_info['year']}")
        c.drawString(50, 700, "Admission No")
        c.drawString(150, 700, "Student")
        c.drawString(250, 700, "Total Fee")
        c.drawString(350, 700, "Amount Paid")
        c.drawString(450, 700, "Balance")
        c.drawString(550, 700, "Term")
        c.drawString(650, 700, "Year")
        y = 680

        db_session = next(get_db())
        try:
            for fee in fees:
                try:
                    if isinstance(fee, Fee):
                        admission_no = str(fee.admission_no).strip()
                        total_fee = float(fee.total_fee or 0)
                        amount_paid = float(fee.amount_paid or 0)
                        balance = float(fee.balance or 0)
                        fee_grade = str(fee.grade).strip()
                        fee_term = str(fee.term).strip()
                        fee_year = str(fee.year).strip()
                    elif isinstance(fee, dict):
                        admission_no = str(fee.get('admission_no', '')).strip()
                        total_fee = float(fee.get('total_fee', 0) or 0)
                        amount_paid = float(fee.get('amount_paid', 0) or 0)
                        balance = float(fee.get('balance', 0) or 0)
                        fee_grade = str(fee.get('grade', '')).strip()
                        fee_term = str(fee.get('term', 'N/A')).strip()
                        fee_year = str(fee.get('year', 'N/A')).strip()
                    elif isinstance(fee, (list, tuple)) and len(fee) >= 7:
                        admission_no = str(fee[0]).strip()
                        total_fee = float(fee[1] or 0)
                        amount_paid = float(fee[2] or 0)
                        balance = float(fee[3] or 0)
                        fee_grade = str(fee[4]).strip()
                        fee_term = str(fee[5]).strip()
                        fee_year = str(fee[6]).strip()
                    else:
                        logger.warning(f"Invalid fee data: {fee}")
                        continue

                    if grade and fee_grade.lower() != grade.lower():
                        logger.debug(f"Skipping fee for admission_no={admission_no}, grade={fee_grade} (expected {grade})")
                        continue
                    if admission_no in processed_adm_nos:
                        logger.debug(f"Duplicate admission_no skipped: {admission_no}")
                        continue

                    processed_adm_nos.add(admission_no)

                    name = db_session.query(User.username).filter(
                        func.lower(User.admission_no) == admission_no.lower(),
                        User.role == 'student'
                    ).scalar() or 'Unknown'

                    if y < 100:
                        c.showPage()
                        c.setFont("Helvetica-Bold", 16)
                        c.drawCentredString(300, 750, "JONYO JUNIOR SECONDARY SCHOOL")
                        c.setFont("Helvetica", 14)
                        c.drawCentredString(300, 730, "FEE STATEMENT")
                        c.setFont("Helvetica", 12)
                        c.drawString(50, 710, f"School Year: {term_info['year']}")
                        c.drawString(50, 700, "Admission No")
                        c.drawString(150, 700, "Student")
                        c.drawString(250, 700, "Total Fee")
                        c.drawString(350, 700, "Amount Paid")
                        c.drawString(450, 700, "Balance")
                        c.drawString(550, 700, "Term")
                        c.drawString(650, 700, "Year")
                        y = 680

                    c.drawString(50, y, admission_no)
                    c.drawString(150, y, name[:15])
                    c.drawString(250, y, f"{total_fee:,.2f}")
                    c.drawString(350, y, f"{amount_paid:,.2f}")
                    c.drawString(450, y, f"{balance:,.2f}")
                    c.drawString(550, y, fee_term)
                    c.drawString(650, y, fee_year)
                    y -= 20

                    has_valid_content = True
                    logger.debug(f"Processed fee: {admission_no}, {name}, {total_fee}, {amount_paid}, {balance}, {fee_term}, {fee_year}")
                except (ValueError, TypeError) as e:
                    logger.error(f"Error processing fee entry: {fee}, Error: {str(e)}")
                    continue
        finally:
            db_session.close()

        if not has_valid_content:
            c.setFont("Helvetica", 12)
            c.drawString(50, y, "No fee records found.")
            logger.warning(f"No fee entries matched for grade: {grade if grade else 'All'}")

        # Watermark
        c.setFont("Helvetica", 50)
        c.setFillColor(colors.grey, alpha=0.2)
        c.rotate(45)
        c.drawCentredString(400, 200, "JONYO JSS")
        c.rotate(-45)
        c.setFillColor(colors.black)

        c.showPage()
        c.save()

        buffer.seek(0)
        pdf_content = buffer.getvalue()
        if not pdf_content.startswith(b'%PDF-'):
            logger.error(f"Invalid PDF generated: starts with {pdf_content[:20]}")
            buffer.close()
            return None

        logger.debug(f"PDF fee statement generated for grade='{grade if grade else 'All'}', records={len(processed_adm_nos)}, size={len(pdf_content)} bytes")
        return buffer
    except Exception as e:
        logger.error(f"Error in generate_fee_statement: {str(e)}\n{traceback.format_exc()}")
        if 'buffer' in locals():
            buffer.close()
        return None

def generate_fee_statement_excel(fees, grade=None, term=None, year=None, highlight_duplicates=False):
    """Generate an Excel file with fee statements."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Fee_Statement_{grade.replace(' ', '_') if grade else 'All_Grade'}"
        headers = ['Admission No', 'Name', 'Total Fee', 'Amount Paid', 'Balance']
        ws.append(headers)
        processed_adm_nos = set()
        duplicate_adm_nos = set()
        row_num = 2

        db_session = next(get_db())
        try:
            for fee in fees:
                try:
                    if isinstance(fee, Fee):
                        admission_no = str(fee.admission_no).strip()
                        total_fee = float(fee.total_fee or 0)
                        amount_paid = float(fee.amount_paid or 0)
                        balance = float(fee.balance or 0)
                        fee_grade = str(fee.grade).strip()
                    elif isinstance(fee, dict):
                        admission_no = str(fee.get('admission_no', '')).strip()
                        total_fee = float(fee.get('total_fee', 0) or 0)
                        amount_paid = float(fee.get('amount_paid', 0) or 0)
                        balance = float(fee.get('balance', 0) or 0)
                        fee_grade = str(fee.get('grade', '')).strip()
                    elif isinstance(fee, (list, tuple)) and len(fee) >= 5:
                        admission_no = str(fee[0]).strip()
                        total_fee = float(fee[1] or 0)
                        amount_paid = float(fee[2] or 0)
                        balance = float(fee[3] or 0)
                        fee_grade = str(fee[4]).strip()
                    else:
                        logger.warning(f"Invalid fee data: {fee}")
                        continue

                    if grade and fee_grade.lower().strip() != grade.lower().strip():
                        continue
                    if admission_no in processed_adm_nos:
                        logger.debug(f"Skipped duplicate admission_no={admission_no} in Excel fee statement")
                        duplicate_adm_nos.add(admission_no)
                        continue
                    processed_adm_nos.add(admission_no)

                    name = db_session.query(User.username).filter(
                        func.lower(User.admission_no) == admission_no.lower(),
                        User.role == 'student'
                    ).scalar() or 'Unknown'
                    ws.append([admission_no, name, total_fee, amount_paid, balance])
                    logger.debug(f"Excel fee row for {admission_no}: {name}, {total_fee}, {amount_paid}, {balance}")
                    row_num += 1
                except (IndexError, TypeError, ValueError) as e:
                    logger.error(f"Error processing fee for admission_no={admission_no}: {str(e)}")
                    continue
        finally:
            db_session.close()

        header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=row_num-1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.column > 2:
                    cell.number_format = '#,##0.00'
        if highlight_duplicates:
            for row in ws.iter_rows(min_row=2, max_row=row_num-1):
                admission_no = row[0].value
                if admission_no in duplicate_adm_nos:
                    for cell in row:
                        cell.fill = yellow_fill
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        if not processed_adm_nos:
            ws.append(['No fee records found.'])
            logger.warning(f"No fees processed for Excel: {grade}")
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.debug(f"Excel fee statement generated for {grade}, size: {len(buffer.getvalue())} bytes")
        return buffer
    except Exception as e:
        logger.error(f"Error in generate_fee_statement_excel: {str(e)}\n{traceback.format_exc()}")
        return None